"""
matching_ap_albaran.py — Yve.01 Modulo AP
Cruza cada factura de proveedor con los albaranes de entrega que la respaldan.

Responde a la pregunta que mas dinero mueve: **¿te estan facturando algo que
nunca llego?** Y su reverso, que sale gratis del mismo cruce: **¿hay mercancia
entregada que nadie ha facturado?** (pasivo no contabilizado).

DOS NIVELES DE CONFIANZA, en este orden:

  1. REFERENCIA EXPLICITA — el albaran cita el numero de factura
     (`referencia_factura`), o el numero de albaran aparece escrito en el
     concepto de la factura. Es exacto, sin heuristica.
  2. PROVEEDOR + VENTANA DE 45 DIAS — el nombre del proveedor casa y la entrega
     cae dentro de la ventana anterior a la fecha de factura. Una factura
     mensual agrupa VARIOS albaranes, asi que el emparejamiento es 1↔N.

EL HOTEL MANDA POR ENCIMA DE LOS DOS NIVELES (ver `_mismo_hotel`):
  Antes este modulo no miraba el hotel, y con dos hoteles en el mismo tenant
  cruzaba la factura de uno con el albaran del otro y la daba por buena. Un
  numero equivocado presentado como correcto, que es la peor forma de fallar.
  Reproducido en el banco de pruebas antes de arreglarlo.

  El modulo NO filtra la carga y NO lee `YVE_HOTEL`: sigue corriendo en lote
  sobre todo lo que hay, y la regla del hotel vive DENTRO del emparejamiento.
  Es a proposito, por dos motivos medidos:
    - el informe se reescribe ENTERO en cada pasada, asi que una carga filtrada
      por el hotel A borraria del informe de hoy las filas del B — y esa hoja
      `Albaranes` es la etapa que GANA en `almacen_datos.albaranes()`;
    - `asignador_cuentas` abre el nombre EXACTO `matching_albaran_<HOY>.xlsx`,
      asi que un informe por hotel desconectaria el cruce del panel en silencio.
  Ademas asi el resultado no depende de la sesion, igual que `agregador_grupo`.

QUE SE COMPARA (y una trampa que costaria caro):
  Un albaran NO lleva IVA y una factura SI. Comparar `total_factura` contra
  `total_albaran` daria una falsa discrepancia del ~21% en TODOS los cruces
  (medido: 55,9% en el caso de prueba). Se compara la **BASE IMPONIBLE** de la
  factura contra la suma de los totales de sus albaranes.

LO QUE ESTE MODULO TODAVIA NO PUEDE HACER:
  Comparar CANTIDADES linea a linea. El albaran tiene lineas; la factura, hoy,
  no — su esquema solo guarda `descripcion_concepto` y los totales. Añadir
  lineas a la factura toca el esquema compartido por los tres caminos de
  entrada, y por eso va en su propia fase (3c). Hasta entonces, la tolerancia
  de cantidad no se puede aplicar: aqui solo se cruzan importes.

Oracle NO interviene: esto genera un informe, no contabiliza nada.
Ejecutar: python matching_ap_albaran.py
"""

import os
import glob
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Multi-tenant desde el primer dia (a diferencia de matching_ap_otras/fb, que
# nacieron con las rutas fijas). tenant_dirs lee YVE_TENANT del entorno cuando
# el dashboard lo lanza por subprocess.
try:
    from tenant_dirs import reportes_dir as _t_rep, procesadas_dir as _t_proc
    REPORTES_DIR = _t_rep()
    PROCESADAS_DIR = _t_proc()
except Exception:
    REPORTES_DIR = os.path.join(BASE_DIR, "reportes")
    PROCESADAS_DIR = os.path.join(BASE_DIR, "facturas-procesadas")
os.makedirs(REPORTES_DIR, exist_ok=True)

FECHA_HOY = date.today().strftime("%Y%m%d")
SALIDA = os.path.join(REPORTES_DIR, f"matching_albaran_{FECHA_HOY}.xlsx")

NF = "NO_ENCONTRADO"

# Ventana hacia atras desde la fecha de factura, en dias. 45 cubre la
# facturacion mensual con holgura (decidido con el usuario).
VENTANA_DIAS = 45

# Tolerancias. Van juntas y se imprimen en el informe para que se vea con que
# criterio se cruzo. La de CANTIDAD es deliberadamente mas laxa: en alimentacion
# el peso servido difiere del pedido por naturaleza, y una alerta que grita por
# 800 gramos de merluza consigue que el AP Manager deje de mirar las alertas.
# TOL_CANTIDAD se aplica en el nivel 3 (Fase 3c), linea a linea.
TOL_IMPORTE = 0.02    # 2 % sobre la base imponible
TOL_CANTIDAD = 0.10   # 10 % — nivel 3. Mas laxa que la de importe A PROPOSITO:
                      # en alimentacion el peso servido difiere del pedido por
                      # naturaleza, y una alerta que grita por 800 g de merluza
                      # consigue que se dejen de mirar las alertas.

VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
AZUL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")


# ── utilidades ────────────────────────────────────────────────────────────

def _txt(v):
    """Texto comparable. NaN, NO_ENCONTRADO y demas vacios -> ''.

    En Python plano, nunca con el accesor `.str`: en pandas 3 `astype(str)` deja
    los nulos como NaN y los propaga (ver las reglas del proyecto).
    """
    s = "" if v is None else str(v)
    s = " ".join(s.split()).strip()
    return "" if s.lower() in ("", "nan", "none", "nat", "<na>", "no_encontrado", "null") else s


def _num(v):
    """Float tolerante con '1.234,56', '450 EUR', '€' y los vacios."""
    s = _txt(v)
    if not s:
        return None
    s = s.replace("EUR", "").replace("€", "").replace(" ", "").strip()
    if "," in s and "." in s:
        s = s.replace(",", "") if s.rfind(".") > s.rfind(",") else s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _fecha(v):
    """Fecha -> date. El mismo dato llega como Timestamp desde un fichero y
    como '18/07/2026' desde otro."""
    s = _txt(v)
    if not s:
        return None
    if isinstance(v, (datetime, pd.Timestamp)):
        return v.date() if hasattr(v, "date") else None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%y",
                "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _clave_prov(nombre):
    """Nombre de proveedor comparable: minusculas, sin puntuacion ni forma
    societaria. 'Pescados Rias, S.L.' y 'PESCADOS RIAS SL' son el mismo."""
    s = _txt(nombre).lower()
    for ch in (".", ",", "-", "&", "'", '"'):
        s = s.replace(ch, " ")
    fuera = {"sl", "s", "l", "sa", "slu", "sau", "sccl", "cb", "sociedad",
             "limitada", "anonima", "y", "de", "del", "la", "el", "los", "las"}
    return " ".join(p for p in s.split() if p not in fuera)


def _hotel(fila):
    """El hotel de un documento, comparable. '' = sin asignar.

    Pasa por `_txt` A PROPOSITO: la columna vuelve de Excel como NaN cuando se
    guardo vacia, y `str(float('nan'))` es la cadena 'nan', que NO es un vacio
    para Python. Sin esto, "sin asignar" seria unas veces '' y otras 'nan', y
    dos documentos igual de huerfanos no cruzarian entre ellos. No es teorico:
    el informe del banco de pruebas imprimia literalmente `hotel=nan`.
    """
    return _txt(fila.get("hotel_id"))


def _mismo_hotel(fila_f, alb):
    """La regla del hotel, decidida con el usuario: IGUALDAD ESTRICTA.

    Una factura solo cruza con un albaran del MISMO hotel, y lo que esta sin
    asignar solo cruza con lo que esta sin asignar. **El vacio NO es comodin.**

    Las dos propiedades que la hacen segura, y que hay que conservar al tocarla:
      - con 0 hoteles en el censo TODO esta sin asignar, asi que todo cruza con
        todo: exactamente el comportamiento de siempre. Es el tenant recien
        creado, o sea la mayoria.
      - con 1 hotel todo lleva la misma etiqueta, asi que tambien sale igual.
    Solo con 2 o mas cambia algo, y lo unico que cambia es que dejan de cruzarse
    documentos de hoteles DISTINTOS.

    NO se consulta el censo a proposito: se comparan las etiquetas entre si. Un
    hotel dado de baja, o un id que ya no este en el censo, sigue cruzando
    consigo mismo en vez de desaparecer del cruce sin avisar.
    """
    return _hotel(fila_f) == _hotel(alb)


_NOMBRES_HOTEL = {}


def _nombre_hotel(hid):
    """El nombre del hotel para ENSEÑARLO. Cae al id si no se puede resolver.

    Es el UNICO sitio del modulo que toca el censo, y solo para escribir un
    texto: la regla del cruce compara etiquetas entre si y no depende de esto
    (ver `_mismo_hotel`). Asi un censo vacio —que en Render pasa despues de cada
    despliegue— empeora el mensaje pero NO cambia ni un estado.
    """
    hid = _txt(hid)
    if not hid:
        return "sin hotel asignado"
    if hid not in _NOMBRES_HOTEL:
        try:
            import censo_hoteles
            _NOMBRES_HOTEL[hid] = censo_hoteles.nombre_de(hid) or hid
        except Exception:
            _NOMBRES_HOTEL[hid] = hid
    return _NOMBRES_HOTEL[hid]


def _mismo_proveedor(a, b):
    """Coincidencia parcial en los dos sentidos, como hace buscar_po."""
    ca, cb = _clave_prov(a), _clave_prov(b)
    if not ca or not cb:
        return False
    return ca == cb or (len(ca) > 3 and ca in cb) or (len(cb) > 3 and cb in ca)


def _base_factura(fila):
    """Lo que hay que comparar con el albaran: la base SIN IVA.

    Si no se extrajo la base, se deriva del total y el porcentaje; si tampoco
    hay porcentaje, se usa el total tal cual y se avisa en el detalle, porque
    entonces la comparacion puede llevar el IVA dentro.
    """
    base = _num(fila.get("base_imponible"))
    if base:
        return base, ""
    total = _num(fila.get("total_factura"))
    if not total:
        return None, "sin importe en la factura"
    pct = _num(fila.get("porcentaje_iva"))
    if pct:
        return round(total / (1 + pct / 100.0), 2), ""
    return total, "sin base imponible: se compara el total, que puede llevar IVA"


# ── carga ─────────────────────────────────────────────────────────────────

def cargar_facturas():
    """Todas las facturas AP del tenant, TODOS los dias.

    Por `almacen_datos` desde el principio: la mercancia llega ANTES que la
    factura, asi que el caso normal es cruzar una factura de hoy con un albaran
    de la semana pasada. Con el 'coge el mas reciente' de siempre, la mitad de
    los cruces no encontraria nada.
    """
    import almacen_datos as _alm
    df = _alm.facturas_ap(PROCESADAS_DIR, REPORTES_DIR)
    return df if df is not None else pd.DataFrame()


def cargar_albaranes():
    import almacen_datos as _alm
    df = _alm.albaranes(PROCESADAS_DIR, REPORTES_DIR)
    return df if df is not None else pd.DataFrame()


def cargar_lineas():
    """Las lineas de las dos partes, para el nivel 3. Vacias = nivel 3 apagado."""
    import almacen_datos as _alm
    lf = _alm.lineas_factura(PROCESADAS_DIR, REPORTES_DIR)
    la = _alm.lineas_albaran(PROCESADAS_DIR, REPORTES_DIR)
    return (lf if lf is not None else pd.DataFrame(),
            la if la is not None else pd.DataFrame())


# ── el cruce ──────────────────────────────────────────────────────────────

def _cita_explicita(fila_f, alb):
    """¿Se citan el uno al otro? Devuelve el motivo, o '' si no."""
    num_f = _txt(fila_f.get("numero_factura"))
    num_a = _txt(alb.get("numero_albaran"))
    ref_f = _txt(alb.get("referencia_factura"))
    if num_f and ref_f and ref_f.lower() == num_f.lower():
        return f"el albarán cita la factura {num_f}"
    if num_a:
        # El esquema de FACTURA no tiene campo para citar un albaran, pero una
        # factura mensual suele listarlos en el concepto ("albaranes 7781,
        # 7782..."). Buscarlo ahi sale gratis y no toca el esquema compartido.
        texto = " ".join(_txt(fila_f.get(c)) for c in
                         ("descripcion_concepto", "numero_factura")).lower()
        if num_a.lower() in texto:
            return f"la factura menciona el albarán {num_a}"
    return ""


def _encaja_nivel2(fila_f, alb):
    """El motivo por el que este albaran encaja por proveedor y ventana, o ''.

    NO mira el hotel: eso lo decide quien llama.

    Vive en su propia funcion para que el EMPAREJAMIENTO y la EXPLICACION de por
    que algo NO ha emparejado salgan de la MISMA comprobacion. Con dos copias
    acabarian diciendo cosas distintas — ya paso en F&B, donde el aviso
    normalizaba los nombres por su cuenta y contradecia al calculo que tenia al
    lado.
    """
    if not _mismo_proveedor(fila_f.get("nombre_proveedor"),
                            alb.get("nombre_proveedor")):
        return ""
    f_fact = _fecha(fila_f.get("fecha"))
    f_ent = _fecha(alb.get("fecha_entrega"))
    if f_fact and f_ent:
        # la entrega tiene que ser ANTERIOR a la factura (o del mismo dia) y
        # caer dentro de la ventana
        if not (f_fact - timedelta(days=VENTANA_DIAS) <= f_ent <= f_fact):
            return ""
        return f"mismo proveedor, entrega del {f_ent.strftime('%d/%m/%Y')}"
    if f_fact or f_ent:
        # falta una de las dos fechas: se acepta por proveedor, pero se deja
        # dicho, porque la ventana no se ha podido comprobar
        return "mismo proveedor (sin fecha para comprobar la ventana)"
    return "mismo proveedor (sin fechas)"


def emparejar(df_fact, df_alb):
    """Empareja facturas y albaranes. Devuelve (empare, porque, bloqueados).

    `empare` es {indice_factura: [indices_alb]}. `bloqueados` es
    {indice_factura: {"citan": [...], "encajan": [...]}}: los albaranes que
    habrian cruzado con esa factura si no fuera por el HOTEL.

    POR QUE HACE FALTA `bloqueados`, y no es un adorno:
    sin el, arreglar el cruce entre hoteles cambia un falso positivo (cuadra y
    esta mal) por un **falso negativo mudo**: la factura sale como "no se ha
    encontrado ninguna entrega", el AP Manager va a reclamarle al proveedor una
    mercancia que si llego, y la informacion para verlo —la entrega esta
    registrada en el hotel de al lado— existe y se la estamos escondiendo.
    Un hueco explicado y un hueco no son lo mismo.

    Un albaran solo puede consumirse UNA vez: si dos facturas se lo repartieran,
    la misma entrega estaria justificando dos cobros, que es justo lo que este
    modulo existe para detectar.

    Primero se resuelven TODAS las referencias explicitas y solo despues se
    reparte por proveedor y fecha, para que una coincidencia debil no se lleve
    un albaran que otra factura reclama por su numero.
    """
    asignados = {}          # indice de albaran -> indice de factura
    porque = {}             # (indice factura, indice albaran) -> motivo
    empare = {i: [] for i in df_fact.index}
    bloqueados = {i: {"citan": [], "encajan": []} for i in df_fact.index}

    def _apuntar(i_f, alb, donde):
        num = _txt(alb.get("numero_albaran")) or "s/n"
        bloqueados[i_f][donde].append((num, _hotel(alb)))

    # ── nivel 1 · referencia explicita ────────────────────────────────────
    for i_f, fila_f in df_fact.iterrows():
        for i_a, alb in df_alb.iterrows():
            motivo = _cita_explicita(fila_f, alb)
            if not motivo:
                continue
            if not _mismo_hotel(fila_f, alb):
                # Se citan por su numero y estan en hoteles distintos. La regla
                # del hotel manda igual, pero una referencia EXACTA que no cruza
                # huele a etiqueta mal puesta y es lo mas accionable que hay.
                _apuntar(i_f, alb, "citan")
                continue
            if i_a in asignados:
                continue
            asignados[i_a] = i_f
            empare[i_f].append(i_a)
            porque[(i_f, i_a)] = motivo

    # ── nivel 2 · proveedor + ventana ─────────────────────────────────────
    for i_f, fila_f in df_fact.iterrows():
        for i_a, alb in df_alb.iterrows():
            motivo = _encaja_nivel2(fila_f, alb)
            if not motivo:
                continue
            if not _mismo_hotel(fila_f, alb):
                _apuntar(i_f, alb, "encajan")
                continue
            if i_a in asignados:
                continue
            asignados[i_a] = i_f
            empare[i_f].append(i_a)
            porque[(i_f, i_a)] = motivo
    return empare, porque, bloqueados


def registro_por_hotel(df_alb):
    """Desde cuando registra albaranes cada hotel. Devuelve (cortes, con_albaran).

    EL CORTE: la entrega mas antigua registrada. Una factura anterior a ese dia
    NO puede tener albaran, asi que marcarla como "sin albaran" es una alerta que
    nadie puede accionar. Y una pantalla que grita en todo lo que mira consigue
    que se deje de mirar: entonces se pierde la alerta de verdad.

    POR QUE POR HOTEL Y NO UNO GLOBAL. El corte era la entrega mas antigua de
    TODOS los hoteles juntos. Con eso, un hotel que empezo a registrar en julio
    recibia alertas por sus facturas de marzo — porque OTRO hotel llevaba
    registrando desde enero. Medido en el banco: de 5 incidencias, 2 no eran
    accionables; con el corte por hotel se quedan en 3, y las 3 son reales.

    LAS DOS COSAS QUE DEVUELVE SON DISTINTAS Y NO SE PUEDEN DEDUCIR UNA DE OTRA:
      cortes[h]    la entrega mas antigua CON FECHA LEGIBLE de ese hotel
      con_albaran  los hoteles que han registrado ALGUN albaran, con fecha o sin

    Si "este hotel no registra albaranes" se dedujera de "no tiene corte", un
    tenant cuyos albaranes no traen fecha legible pasaria de alertar de todo a
    no alertar de NADA, en silencio — y ese tenant es justo el caso de 0 hoteles,
    que es el que no se puede mover. Por eso van separadas.

    Un hotel que no esta en `con_albaran` no ha registrado ni un albaran, y sus
    facturas no generan incidencia (ver `analizar_factura`). Sin esa excepcion el
    cambio seria PEOR que el corte global: con el corte a None, las facturas de
    ese hotel pasaban de ANTERIOR_AL_REGISTRO a FACTURA_SIN_ALBARAN. Es el caso
    del hotel que entra nuevo, que es el que mas veces va a pasar.

    Degenera solo: con 0 hoteles todo cae en la caja '' y con 1 hotel en la suya,
    asi que en los dos casos el corte vuelve a ser exactamente el de siempre.
    """
    cortes, con_albaran = {}, set()
    for _, a in df_alb.iterrows():
        h = _hotel(a)
        con_albaran.add(h)
        f = _fecha(a.get("fecha_entrega"))
        if f and (h not in cortes or f < cortes[h]):
            cortes[h] = f
    return cortes, con_albaran


def _clave_desc(v):
    """Descripcion comparable: sin acentos, sin mayusculas, sin espacios de mas.

    Misma tecnica que _clave_plato/_txt_ing en F&B, y por el mismo motivo: la
    factura escribe "Solomillo de Ternera" y el albaran "SOLOMILLO DE TERNERA".
    En Python plano, nunca con el accesor `.str`.
    """
    import unicodedata
    s = _txt(v).lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


def _lineas_de(df_lin, columna, valores, hotel=None):
    """Las lineas cuyo `columna` esta en `valores`. Sin .str (pandas 3).

    `hotel` acota ademas por hotel, y hace falta de verdad: las lineas de la
    factura se buscan por `archivo`, o sea **por el nombre del fichero**, y dos
    hoteles del mismo grupo pueden subir cada uno su `factura_julio.pdf`. Sin
    acotar, las lineas de un hotel se comparan contra el albaran del otro y sale
    una diferencia de precio inventada sobre una factura que esta bien.

    Medido en el banco antes de arreglarlo: DOS facturas correctas, una de cada
    hotel, salian las dos como DIFERENCIA_LINEA. Sin `hotel` la funcion se
    comporta como siempre, que es lo que mantiene intacto el caso de 0 hoteles.
    """
    if df_lin is None or df_lin.empty or columna not in df_lin.columns:
        return []
    vals = {_txt(v) for v in valores if _txt(v)}
    if not vals:
        return []
    return [r for _, r in df_lin.iterrows()
            if _txt(r.get(columna)) in vals
            and (hotel is None or _hotel(r) == hotel)]


def comparar_lineas(lin_f, lin_a):
    """NIVEL 3 · linea a linea. Devuelve (avisos, euros_en_juego, sin_pareja).

    Empareja por descripcion normalizada y compara precio unitario y cantidad.
    Lo que no se puede emparejar NO se inventa: se cuenta y se dice.
    """
    avisos, euros, sin_pareja = [], 0.0, 0
    if not lin_f or not lin_a:
        return avisos, euros, sin_pareja

    por_desc = {}
    for la in lin_a:
        por_desc.setdefault(_clave_desc(la.get("descripcion")), []).append(la)

    for lf in lin_f:
        k = _clave_desc(lf.get("descripcion"))
        cand = por_desc.get(k)
        if not cand:
            sin_pareja += 1
            continue
        # si el albaran repite el producto (dos lotes), se suma la cantidad y se
        # usa el precio medio ponderado: fusionar dos lotes en uno seria perder
        # la mercancia de uno de ellos
        cant_a = sum(_num(c.get("cantidad")) or 0.0 for c in cand)
        imp_a = sum(_num(c.get("importe")) or 0.0 for c in cand)
        prec_a = (imp_a / cant_a) if cant_a else (_num(cand[0].get("precio_unitario")))
        prec_f = _num(lf.get("precio_unitario"))
        cant_f = _num(lf.get("cantidad"))
        desc = _txt(lf.get("descripcion")) or "(sin descripción)"

        if prec_f is not None and prec_a:
            dif = abs(prec_f - prec_a) / prec_a
            if dif > TOL_IMPORTE:
                de_mas = (prec_f - prec_a) * (cant_f if cant_f is not None else 0.0)
                euros += de_mas
                signo = "más" if prec_f > prec_a else "menos"
                avisos.append(f"{desc}: {prec_f:.2f} EUR/ud en factura vs "
                              f"{prec_a:.2f} en albarán ({dif*100:.1f}% {signo}"
                              + (f", {abs(de_mas):.2f} EUR" if cant_f else "") + ")")
        if cant_f is not None and cant_a:
            difc = abs(cant_f - cant_a) / cant_a
            if difc > TOL_CANTIDAD:
                signo = "más" if cant_f > cant_a else "menos"
                avisos.append(f"{desc}: factura {cant_f:g} vs albarán {cant_a:g} "
                              f"({difc*100:.1f}% {signo})")
    return avisos, round(euros, 2), sin_pareja


def _lista_alb(pares):
    """'ALB-7781 (Hotel Sol Playa), ALB-7782 (sin hotel asignado)'.

    Se nombra el hotel, no solo el numero: "está en otro hotel" no le sirve de
    nada a quien tiene cuatro. Se agrupa por hotel para no repetir el nombre.
    """
    por_hotel = {}
    for num, hid in pares:
        por_hotel.setdefault(_nombre_hotel(hid), []).append(num)
    return "; ".join(f"{', '.join(nums)} ({nom})" for nom, nums in por_hotel.items())


def _exige_albaran(fila):
    try:
        from cuentas_proveedor import exige_albaran
        return exige_albaran(fila)
    except Exception:
        return True


def analizar_factura(fila_f, indices_alb, df_alb, porque, i_f, cortes=None,
                     df_lin_f=None, df_lin_a=None, hay_hoteles=False,
                     bloqueados=None, con_albaran=None):
    hot = _hotel(fila_f)
    base, aviso = _base_factura(fila_f)
    albs = [df_alb.loc[i] for i in indices_alb]
    nums = [_txt(a.get("numero_albaran")) or "s/n" for a in albs]
    suma = sum(_num(a.get("total_albaran")) or 0.0 for a in albs)

    if not albs:
        f_fact = _fecha(fila_f.get("fecha"))
        diff = dif_pct = NF
        corte = (cortes or {}).get(hot)
        if not _exige_albaran(fila_f):
            # Un servicio o suministro no lleva albaran, este el hotel donde
            # este y sea la factura de cuando sea: va ANTES que "anterior al
            # registro" para que la luz no salga con dos estados distintos
            # segun la fecha del primer albaran del hotel.
            estado = "NO_REQUIERE_ALBARAN"
            detalle = "servicio o suministro: no lleva albarán de entrega"
        elif con_albaran is not None and hot not in con_albaran:
            # Este hotel no ha registrado NI UN albaran. Es el equivalente por
            # hotel del "no hay albaranes todavia: nada que cruzar" con el que
            # sale el modulo entero, y por el mismo motivo: no se puede reclamar
            # una entrega en un sitio donde todavia no se registran entregas.
            # Sin esto, un hotel que entra nuevo se llena de alertas el dia que
            # OTRO hotel sube su primer albaran.
            estado = "ANTERIOR_AL_REGISTRO"
            detalle = ("no hay ningún albarán sin hotel asignado con el que esta "
                       "factura pueda cruzar" if not hot else
                       f"{_nombre_hotel(hot)} todavía no ha registrado ningún "
                       "albarán: no se puede esperar que esta factura tenga uno")
        elif corte and f_fact and f_fact < corte:
            # no es una incidencia: cuando se emitio esta factura todavia no se
            # registraban albaranes EN ESTE HOTEL, asi que no hay nada que reclamar
            estado = "ANTERIOR_AL_REGISTRO"
            detalle = (f"factura del {f_fact.strftime('%d/%m/%Y')}, anterior al primer "
                       f"albarán registrado" + (f" en {_nombre_hotel(hot)}" if hot else "")
                       + f" ({corte.strftime('%d/%m/%Y')}): no se puede esperar que "
                       "tenga uno")
        else:
            estado = "FACTURA_SIN_ALBARAN"
            detalle = ("no se ha encontrado ninguna entrega que respalde esta factura "
                       f"(mismo proveedor, {VENTANA_DIAS} días antes)")
            if not f_fact:
                # sin fecha no se puede saber si deberia tener albaran. Se deja
                # como incidencia A PROPOSITO: una factura sin fecha legible ya
                # es motivo para mirarla.
                detalle += " · además no se ha podido leer su fecha"
    elif base is None:
        estado = "SIN_IMPORTE"
        detalle = f"{len(albs)} albarán(es) encontrado(s), pero {aviso}"
        diff = dif_pct = NF
    elif suma <= 0:
        estado = "SIN_IMPORTE"
        detalle = f"albarán(es) {', '.join(nums)} sin importe extraíble"
        diff = dif_pct = NF
    else:
        diff = round(base - suma, 2)
        dif_pct = abs(diff) / suma
        motivos = "; ".join(porque[(i_f, i)] for i in indices_alb)
        if dif_pct <= TOL_IMPORTE:
            estado = "MATCH_ALBARAN_OK"
            detalle = (f"{len(albs)} albarán(es) {', '.join(nums)} cuadran "
                       f"({dif_pct*100:.2f}% de diferencia) · {motivos}")
        else:
            estado = "DIFERENCIA_IMPORTE"
            signo = "MÁS" if diff > 0 else "menos"
            detalle = (f"la factura cobra {abs(diff):.2f} EUR {signo} de lo entregado: "
                       f"base {base:.2f} vs albaranes {suma:.2f} "
                       f"({', '.join(nums)}) · {motivos}")
        if aviso:
            detalle += f" · OJO: {aviso}"

    # ── NIVEL 3 · linea a linea (Fase 3c) ────────────────────────────────
    # Lo que el nivel 2 no puede ver: suben un producto y bajan otro, el total
    # cuadra y la factura pasa. Solo se mira si hay albaranes emparejados.
    n_comp = n_sin_pareja = 0
    euros_linea = NF
    avisos_l = []
    if albs and estado not in ("ANTERIOR_AL_REGISTRO", "FACTURA_SIN_ALBARAN", "NO_REQUIERE_ALBARAN"):
        lin_f = _lineas_de(df_lin_f, "archivo", [fila_f.get("archivo")], hot)
        if not lin_f:
            lin_f = _lineas_de(df_lin_f, "numero_factura",
                               [fila_f.get("numero_factura")], hot)
        lin_a = _lineas_de(df_lin_a, "numero_albaran",
                           [a.get("numero_albaran") for a in albs], hot)
        avisos_l, euros_linea, n_sin_pareja = comparar_lineas(lin_f, lin_a)
        n_comp = len(lin_f)
        if avisos_l:
            if estado == "MATCH_ALBARAN_OK":
                # el total cuadraba: esto es lo que 3c existe para pillar
                estado = "DIFERENCIA_LINEA"
                detalle = ("el total cuadra pero las líneas NO: "
                           + " · ".join(avisos_l))
                if euros_linea:
                    signo = "de más" if euros_linea > 0 else "de menos"
                    detalle += f" · {abs(euros_linea):.2f} EUR {signo} en total"
            else:
                # ya habia incidencia: el estado NO cambia, se dice cual es
                detalle += " · por líneas: " + " · ".join(avisos_l)
        if n_sin_pareja and n_comp:
            detalle += (f" · {n_sin_pareja} de {n_comp} línea(s) de la factura sin "
                        "pareja en el albarán: no se han podido comparar")
        if not lin_f:
            euros_linea = NF

    # ── Por que este documento no cruza con mas cosas ─────────────────────
    # Los dos avisos van al final, DESPUES de que el estado este decidido: no
    # cambian ningun estado, solo explican un hueco. Un hueco sin explicar en
    # una pantalla de finanzas se lee como "faltan albaranes" cuando lo que
    # falta es la etiqueta.
    if hay_hoteles and not hot:
        # Solo se avisa si hay hoteles EN JUEGO. Con 0 hoteles en el censo todo
        # esta sin asignar y el aviso saldria en TODAS las facturas, que es
        # ruido puro: ahi "sin hotel" no es una anomalia, es el estado normal.
        detalle += (" · sin hotel asignado: solo puede cruzar con albaranes "
                    "sin hotel")
    _citan = (bloqueados or {}).get("citan") or []
    _encajan = (bloqueados or {}).get("encajan") or []
    if _citan:
        detalle += (" · OJO: el albarán " + _lista_alb(_citan) + " cita esta "
                    "factura y está en otro hotel — revisa la etiqueta")
    elif _encajan and estado not in ("MATCH_ALBARAN_OK", "ANTERIOR_AL_REGISTRO"):
        # Solo cuando la factura se queda con un hueco o una diferencia: si ya
        # cuadra con lo suyo, esto seria ruido. Sin este aviso el arreglo del
        # hotel convierte un cruce erroneo en un "no hay entrega" mudo, y eso se
        # lee como una reclamacion al proveedor que no toca.
        detalle += (f" · hay {len(_encajan)} entrega(s) de este proveedor en la "
                    "ventana registradas en otro hotel (" + _lista_alb(_encajan)
                    + "): no se cruzan porque el hotel no coincide")

    return {
        "archivo":            _txt(fila_f.get("archivo")) or NF,
        "numero_factura":     _txt(fila_f.get("numero_factura")) or NF,
        "fecha":              _txt(fila_f.get("fecha")) or NF,
        "nombre_proveedor":   _txt(fila_f.get("nombre_proveedor")) or NF,
        # La etiqueta tiene que sobrevivir la cadena ENTERA. La hoja `Albaranes`
        # ya la llevaba porque arrastra la fila entera; esta se construye a mano
        # campo a campo y se quedaba sin ella. Es la misma leccion que costo cara
        # en AR, donde se estampo el hotel en la primera etapa y se dio por hecho
        # el resto. Vacio = sin asignar, no es un fallo de lectura.
        "hotel_id":           hot,
        "base_imponible":     base if base is not None else NF,
        "total_factura":      _num(fila_f.get("total_factura")) or NF,
        "n_albaranes":        len(albs),
        "albaranes":          ", ".join(nums) if nums else NF,
        "total_albaranes":    round(suma, 2) if albs else NF,
        "diferencia_importe": diff,
        "diferencia_pct":     f"{dif_pct*100:.2f}%" if isinstance(dif_pct, float) else NF,
        "lineas_comparadas":  n_comp if n_comp else NF,
        "lineas_con_aviso":    len(avisos_l) if avisos_l else 0,
        "desvio_por_lineas":   euros_linea,
        "estado_matching":    estado,
        "detalle_matching":   detalle,
    }


def analizar_albaranes(df_alb, asignados_por_alb, df_fact):
    """El reverso: mercancia entregada que nadie ha facturado."""
    filas = []
    for i_a, alb in df_alb.iterrows():
        i_f = asignados_por_alb.get(i_a)
        if i_f is None:
            estado = "ALBARAN_SIN_FACTURAR"
            detalle = "entregado, pero ninguna factura lo respalda todavía"
            num_f = NF
        else:
            estado = "ALBARAN_FACTURADO"
            num_f = _txt(df_fact.loc[i_f].get("numero_factura")) or NF
            detalle = f"facturado en {num_f}"
        # OJO: se arrastra la fila ENTERA del albaran, no una seleccion de
        # campos. Esta hoja es la etapa mas avanzada de _ETAPAS_ALB, asi que es
        # la que gana en almacen_datos.albaranes(): si aqui se perdiera
        # `referencia_factura`, el siguiente cruce ya no podria emparejar por
        # referencia explicita — el modulo se romperia a si mismo. Mismo patron
        # que usan matching_ap_otras/fb con **fila.to_dict().
        filas.append({
            **{k: v for k, v in alb.to_dict().items() if not str(k).startswith("_")},
            "numero_factura":   num_f,
            "estado":           estado,
            "detalle":          detalle,
        })
    return filas


# ── informe ───────────────────────────────────────────────────────────────

_COLORES = {
    "MATCH_ALBARAN_OK": VERDE, "ALBARAN_FACTURADO": VERDE,
    "DIFERENCIA_IMPORTE": ROJO,
    "FACTURA_SIN_ALBARAN": AMARILLO, "ALBARAN_SIN_FACTURAR": AMARILLO,
    "ANTERIOR_AL_REGISTRO": AZUL, "NO_REQUIERE_ALBARAN": AZUL,
    "SIN_IMPORTE": AZUL,
}


def aplicar_formato(ws, col_estado):
    try:
        idx = next((i + 1 for i, c in enumerate(ws[1]) if c.value == col_estado), None)
        if idx is None:
            return
        for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
            fill = _COLORES.get(ws.cell(ri, idx).value)
            if fill:
                for cell in row:
                    cell.fill = fill
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    except Exception:
        pass


def generar_resumen(df_f, df_a):
    filas = []
    if not df_f.empty:
        for est, n in df_f["estado_matching"].value_counts().items():
            filas.append({"Bloque": "Facturas", "Estado": est, "Cantidad": int(n),
                          "Pct": f"{n/len(df_f)*100:.1f}%"})
    if not df_a.empty:
        for est, n in df_a["estado"].value_counts().items():
            filas.append({"Bloque": "Albaranes", "Estado": est, "Cantidad": int(n),
                          "Pct": f"{n/len(df_a)*100:.1f}%"})
    # dejar por escrito CON QUE criterio se cruzo: un informe sin sus umbrales
    # no se puede auditar despues
    filas += [
        {"Bloque": "─" * 12, "Estado": "", "Cantidad": "", "Pct": ""},
        {"Bloque": "Criterio", "Estado": "ventana de emparejamiento",
         "Cantidad": f"{VENTANA_DIAS} días", "Pct": ""},
        {"Bloque": "Criterio", "Estado": "tolerancia de importe",
         "Cantidad": f"{TOL_IMPORTE*100:.0f}%", "Pct": ""},
        {"Bloque": "Criterio", "Estado": "tolerancia de cantidad (nivel 3, línea a línea)",
         "Cantidad": f"{TOL_CANTIDAD*100:.0f}%", "Pct": ""},
        {"Bloque": "Criterio", "Estado": "se compara la BASE IMPONIBLE (el albarán no lleva IVA)",
         "Cantidad": "", "Pct": ""},
        {"Bloque": "Criterio", "Estado": "solo se cruzan documentos del MISMO hotel; lo que no lleva hotel, solo con lo que no lleva hotel",
         "Cantidad": "", "Pct": ""},
        {"Bloque": "Criterio", "Estado": "no se alerta de facturas anteriores al primer albarán registrado EN SU HOTEL, ni de las de un hotel que aún no registra albaranes",
         "Cantidad": "", "Pct": ""},
    ]
    return pd.DataFrame(filas)


def main():
    print("=" * 60)
    print("  Yve.01 — Matching AP · Factura ↔ Albarán")
    print("=" * 60)

    df_fact = cargar_facturas()
    df_alb = cargar_albaranes()
    if df_fact.empty:
        print("\n  No hay facturas AP que cruzar.")
        return 0
    if df_alb.empty:
        print("\n  No hay albaranes todavía: nada que cruzar.")
        print("  (Sube los albaranes por Procesar Archivos y vuelve a ejecutarlo.)")
        return 0

    print(f"  Facturas: {len(df_fact)}  |  Albaranes: {len(df_alb)}"
          f"  |  ventana {VENTANA_DIAS} días · tolerancia {TOL_IMPORTE*100:.0f}%\n")

    # ¿Hay hoteles EN JUEGO? Basta con que UN documento lleve etiqueta. Sirve
    # para no llenar de avisos un tenant que no usa hoteles: alli "sin hotel" es
    # el estado normal de todo, no una anomalia de nadie.
    hay_hoteles = (any(_hotel(f) for _, f in df_fact.iterrows())
                   or any(_hotel(a) for _, a in df_alb.iterrows()))
    if hay_hoteles:
        _sin_h = sum(1 for _, f in df_fact.iterrows() if not _hotel(f))
        _sin_a = sum(1 for _, a in df_alb.iterrows() if not _hotel(a))
        print(f"  Separado por hotel: solo cruzan documentos del mismo hotel"
              + (f" · {_sin_h} factura(s) y {_sin_a} albarán(es) sin hotel asignado"
                 if (_sin_h or _sin_a) else "") + "\n")

    empare, porque, bloqueados = emparejar(df_fact, df_alb)
    por_alb = {i_a: i_f for i_f, idxs in empare.items() for i_a in idxs}

    df_lin_f, df_lin_a = cargar_lineas()
    if not df_lin_f.empty and not df_lin_a.empty:
        print(f"  Nivel 3 activo: {len(df_lin_f)} línea(s) de factura vs "
              f"{len(df_lin_a)} de albarán · tolerancia cantidad "
              f"{TOL_CANTIDAD*100:.0f}%\n")
    else:
        # honestidad: si no hay lineas por los dos lados, el nivel 3 no se
        # aplica y hay que decirlo, no callarlo
        _falta = "la factura" if df_lin_f.empty else "el albarán"
        print(f"  Nivel 3 (línea a línea) no se aplica: {_falta} no trae líneas.\n")

    cortes, con_albaran = registro_por_hotel(df_alb)
    res_f = [analizar_factura(df_fact.loc[i_f], idxs, df_alb, porque, i_f, cortes,
                              df_lin_f, df_lin_a, hay_hoteles,
                              bloqueados.get(i_f), con_albaran)
             for i_f, idxs in empare.items()]
    res_a = analizar_albaranes(df_alb, por_alb, df_fact)

    iconos = {"MATCH_ALBARAN_OK": "✓", "DIFERENCIA_IMPORTE": "✗",
              "FACTURA_SIN_ALBARAN": "?", "SIN_IMPORTE": "~",
              "ANTERIOR_AL_REGISTRO": "·", "DIFERENCIA_LINEA": "✗",
              "NO_REQUIERE_ALBARAN": "·"}
    for r in res_f:
        print(f"  [{iconos.get(r['estado_matching'], '·')}] {r['numero_factura']} → {r['estado_matching']}")
        if r["estado_matching"] not in ("MATCH_ALBARAN_OK", "ANTERIOR_AL_REGISTRO", "NO_REQUIERE_ALBARAN"):
            print(f"       {r['detalle_matching']}")
    sin_facturar = [r for r in res_a if r["estado"] == "ALBARAN_SIN_FACTURAR"]
    for r in sin_facturar:
        print(f"  [!] {r['numero_albaran']} ({r['nombre_proveedor']}) → ALBARAN_SIN_FACTURAR")

    df_res_f = pd.DataFrame(res_f)
    df_res_a = pd.DataFrame(res_a)
    df_sum = generar_resumen(df_res_f, df_res_a)

    with pd.ExcelWriter(SALIDA, engine="openpyxl") as w:
        df_res_f.to_excel(w, index=False, sheet_name="Facturas")
        # OJO: esta hoja se llama 'Albaranes' a proposito. almacen_datos lee esa
        # hoja para consolidar los albaranes de todos los dias, asi que el
        # estado del cruce viaja con ellos y no se pierde al cambiar de dia.
        df_res_a.to_excel(w, index=False, sheet_name="Albaranes")
        df_sum.to_excel(w, index=False, sheet_name="Resumen")
        aplicar_formato(w.sheets["Facturas"], "estado_matching")
        aplicar_formato(w.sheets["Albaranes"], "estado")
        for sn in ("Facturas", "Albaranes", "Resumen"):
            ws = w.sheets[sn]
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 4, 60)

    print("\n" + "─" * 60)
    print("  RESUMEN")
    print("─" * 60)
    for _, r in df_sum.iterrows():
        print(f"  {str(r['Bloque']):<12} {str(r['Estado']):<58} {str(r['Cantidad'])}")
    # Linea legible por maquina, para que quien lo lance no tenga que contar
    # lineas de la consola: contarlas se comia tambien la fila del resumen y el
    # lote cantaba 2 incidencias donde habia 1. Mismo patron que el "FALTAN:"
    # de lector_ota.py.
    _inc_f = int((df_res_f["estado_matching"].isin(
        ["FACTURA_SIN_ALBARAN", "DIFERENCIA_IMPORTE",
         "DIFERENCIA_LINEA"])).sum()) if not df_res_f.empty else 0
    _inc_a = len(sin_facturar)
    print(f"INCIDENCIAS: {_inc_f}|{_inc_a}")
    # y CUALES son, para que el log del dashboard no deje adivinar
    _det = [f"{r['numero_factura']}={r['estado_matching']}" for r in res_f
            if r["estado_matching"] in ("FACTURA_SIN_ALBARAN", "DIFERENCIA_IMPORTE", "DIFERENCIA_LINEA")]
    if _det:
        print("INCIDENCIAS_DETALLE: " + ";".join(_det))
    print(f"\n✅ Reporte: {SALIDA}")
    print("=" * 60)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
