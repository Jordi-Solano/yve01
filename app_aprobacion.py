"""
app_aprobacion.py — Yve.01
Dashboard Flask optimizado para movil.
Muestra facturas del reporte de verificacion/DI para que el
jefe de departamento las apruebe o rechace con comentario.
Guarda cada decision en aprobaciones/aprobaciones.xlsx.

Uso:
    py app_aprobacion.py
    Abrir en navegador: http://localhost:5000
"""

import os, glob, json
from datetime import datetime
from pathlib import Path
import pandas as pd
from flask import Flask, request, redirect, url_for, jsonify

# ── Config ─────────────────────────────────────────────────────────────────
BASE_DIR       = Path(__file__).parent
REPORTES_DIR   = BASE_DIR / "reportes"
APROB_DIR      = BASE_DIR / "aprobaciones"
APROB_EXCEL    = APROB_DIR / "aprobaciones.xlsx"
APROB_DIR.mkdir(exist_ok=True)

NF = "NO_ENCONTRADO"

app = Flask(__name__)
app.secret_key = "yve01-aprobaciones-2024"


# ── Carga de datos ──────────────────────────────────────────────────────────

def cargar_facturas():
    """Carga el reporte DI mas reciente (tiene todos los campos)."""
    excels = sorted(REPORTES_DIR.glob("doble_imposicion_*.xlsx"), reverse=True)
    if excels:
        df = pd.read_excel(excels[0], sheet_name="Detalle_DI")
    else:
        excels = sorted(REPORTES_DIR.glob("verificacion_*.xlsx"), reverse=True)
        if not excels:
            return pd.DataFrame()
        df = pd.read_excel(excels[0], sheet_name="Detalle")
        df["estado_di"] = NF

    # Convertir a tipos seguros para JSON
    for col in df.columns:
        df[col] = df[col].fillna(NF).astype(str)
    return df


def cargar_aprobaciones():
    """Carga el historial de aprobaciones ya registradas."""
    if APROB_EXCEL.exists():
        return pd.read_excel(APROB_EXCEL).fillna("").astype(str)
    return pd.DataFrame(columns=["fecha","hora","numero_factura","ota","hotel","accion","comentario","estado_verificacion","estado_di"])


def guardar_aprobacion(data: dict):
    """Agrega una fila al Excel de aprobaciones."""
    df = cargar_aprobaciones()
    nueva = pd.DataFrame([data])
    df = pd.concat([df, nueva], ignore_index=True)
    with pd.ExcelWriter(APROB_EXCEL, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Aprobaciones")
        ws = writer.sheets["Aprobaciones"]
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for row in ws.iter_rows(min_row=2):
                accion = ws.cell(row=row[0].row, column=7).value
                fill_color = "C6EFCE" if accion == "APROBADA" else "FFC7CE"
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                for cell in row:
                    cell.fill = fill
        except Exception:
            pass
        for col in ws.columns:
            ancho = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(ancho + 3, 35)


def facturas_ya_aprobadas():
    """Devuelve set de numeros de factura ya procesados."""
    df = cargar_aprobaciones()
    if df.empty or "numero_factura" not in df.columns:
        return set()
    return set(df["numero_factura"].astype(str).tolist())


# ── HTML ────────────────────────────────────────────────────────────────────

HTML_BASE = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1, maximum-scale=1">
<meta name="apple-mobile-web-app-capable" content="yes">
<title>Yve.01 — Aprobaciones</title>
<style>
  :root {
    --verde:   #16a34a; --verde-claro: #dcfce7; --verde-txt: #14532d;
    --rojo:    #dc2626; --rojo-claro:  #fee2e2; --rojo-txt:  #7f1d1d;
    --naranja: #ea580c; --naranja-claro:#ffedd5; --naranja-txt:#7c2d12;
    --azul:    #2563eb; --azul-claro:  #dbeafe; --azul-txt:  #1e3a8a;
    --gris:    #6b7280; --gris-claro:  #f3f4f6;
    --bg: #f8fafc; --card: #ffffff;
    --sombra: 0 1px 3px rgba(0,0,0,.12), 0 1px 2px rgba(0,0,0,.08);
    --radio: 14px;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; -webkit-tap-highlight-color: transparent; }
  body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
         background: var(--bg); color: #1e293b; min-height: 100vh; }

  /* Header */
  .header { background: #0f172a; color: white; padding: 16px 20px;
             position: sticky; top: 0; z-index: 100; box-shadow: 0 2px 8px rgba(0,0,0,.3); }
  .header h1 { font-size: 18px; font-weight: 700; letter-spacing: -.3px; }
  .header p  { font-size: 12px; color: #94a3b8; margin-top: 2px; }
  .header-stats { display: flex; gap: 12px; margin-top: 10px; }
  .stat { background: rgba(255,255,255,.1); border-radius: 8px;
          padding: 6px 10px; font-size: 12px; }
  .stat strong { display: block; font-size: 18px; font-weight: 700; }

  /* Nav tabs */
  .tabs { display: flex; background: white; border-bottom: 1px solid #e2e8f0; }
  .tab  { flex: 1; padding: 12px 8px; text-align: center; font-size: 13px; font-weight: 500;
           color: var(--gris); border: none; background: none; cursor: pointer;
           border-bottom: 2px solid transparent; transition: .2s; }
  .tab.active { color: #0f172a; border-bottom-color: #0f172a; }

  /* Main */
  main { padding: 16px; max-width: 480px; margin: 0 auto; }

  /* Factura card */
  .factura { background: var(--card); border-radius: var(--radio); box-shadow: var(--sombra);
              margin-bottom: 16px; overflow: hidden; }
  .factura-header { padding: 14px 16px 10px; }
  .factura-ota { font-size: 17px; font-weight: 700; }
  .factura-num { font-size: 12px; color: var(--gris); margin-top: 2px; }
  .factura-hotel { font-size: 13px; color: #475569; margin-top: 4px; }
  .badges { display: flex; gap: 6px; flex-wrap: wrap; margin-top: 10px; }
  .badge { padding: 4px 9px; border-radius: 20px; font-size: 11px; font-weight: 600; }
  .badge-verde    { background: var(--verde-claro);   color: var(--verde-txt); }
  .badge-rojo     { background: var(--rojo-claro);    color: var(--rojo-txt); }
  .badge-naranja  { background: var(--naranja-claro); color: var(--naranja-txt); }
  .badge-azul     { background: var(--azul-claro);    color: var(--azul-txt); }
  .badge-gris     { background: var(--gris-claro);    color: var(--gris); }

  /* Detalles */
  .factura-detalles { padding: 0 16px 12px; }
  .detalle-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; }
  .detalle-item { background: var(--gris-claro); border-radius: 8px; padding: 8px 10px; }
  .detalle-label { font-size: 10px; color: var(--gris); text-transform: uppercase; letter-spacing: .5px; }
  .detalle-val   { font-size: 14px; font-weight: 600; margin-top: 1px; }
  .disc-box { background: var(--rojo-claro); border-radius: 8px; padding: 10px 12px;
               margin-top: 8px; display: flex; justify-content: space-between; align-items: center; }
  .disc-label { font-size: 11px; color: var(--rojo); font-weight: 600; }
  .disc-val   { font-size: 20px; font-weight: 800; color: var(--rojo); }

  /* Acciones */
  .factura-acciones { padding: 0 16px 16px; }
  .ya-procesada { background: var(--gris-claro); border-radius: 10px; padding: 12px;
                   text-align: center; font-size: 13px; color: var(--gris); }
  .ya-procesada strong { display: block; font-size: 15px; margin-bottom: 2px; }
  .comentario-wrap { margin-bottom: 10px; }
  .comentario-wrap label { display: block; font-size: 12px; color: var(--gris);
                            font-weight: 500; margin-bottom: 4px; }
  .comentario-wrap textarea { width: 100%; border: 1.5px solid #e2e8f0; border-radius: 10px;
                               padding: 10px 12px; font-size: 14px; resize: none; min-height: 70px;
                               font-family: inherit; transition: border-color .2s; }
  .comentario-wrap textarea:focus { outline: none; border-color: #0f172a; }
  .btn-row { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; }
  .btn { border: none; border-radius: 10px; padding: 14px; font-size: 15px; font-weight: 700;
          cursor: pointer; transition: .15s; display: flex; align-items: center;
          justify-content: center; gap: 6px; }
  .btn:active { transform: scale(.97); }
  .btn-aprobar { background: var(--verde); color: white; }
  .btn-rechazar { background: var(--rojo);  color: white; }
  .btn-aprobar:hover  { background: #15803d; }
  .btn-rechazar:hover { background: #b91c1c; }

  /* Historial */
  #tab-historial { display: none; }
  .hist-item { background: white; border-radius: var(--radio); box-shadow: var(--sombra);
                margin-bottom: 10px; padding: 14px 16px; display: flex; gap: 12px; align-items: flex-start; }
  .hist-icon { width: 36px; height: 36px; border-radius: 50%; display: flex;
                align-items: center; justify-content: center; font-size: 18px; flex-shrink: 0; }
  .hist-icon.aprobada  { background: var(--verde-claro); }
  .hist-icon.rechazada { background: var(--rojo-claro); }
  .hist-body { flex: 1; min-width: 0; }
  .hist-titulo { font-size: 14px; font-weight: 600; }
  .hist-meta   { font-size: 12px; color: var(--gris); margin-top: 2px; }
  .hist-comentario { font-size: 13px; color: #475569; margin-top: 6px;
                      background: var(--gris-claro); border-radius: 8px; padding: 7px 10px; }
  .empty-state { text-align: center; padding: 48px 20px; color: var(--gris); }
  .empty-state .emoji { font-size: 48px; margin-bottom: 12px; }
  .empty-state p { font-size: 15px; }

  /* Toast */
  #toast { position: fixed; bottom: 24px; left: 50%; transform: translateX(-50%) translateY(80px);
            background: #0f172a; color: white; padding: 12px 20px; border-radius: 12px;
            font-size: 14px; font-weight: 500; transition: transform .3s ease; z-index: 999;
            white-space: nowrap; max-width: 90vw; text-align: center; }
  #toast.show { transform: translateX(-50%) translateY(0); }

  .spinner { display: none; width: 18px; height: 18px; border: 2px solid rgba(255,255,255,.4);
              border-top-color: white; border-radius: 50%; animation: spin .6s linear infinite; }
  @keyframes spin { to { transform: rotate(360deg); } }
</style>
</head>
<body>

<div class="header">
  <h1>Yve.01 — Aprobaciones AR</h1>
  <p id="subtitulo">Cargando facturas...</p>
  <div class="header-stats">
    <div class="stat"><strong id="stat-pendientes">—</strong>Pendientes</div>
    <div class="stat"><strong id="stat-aprobadas">—</strong>Aprobadas</div>
    <div class="stat"><strong id="stat-rechazadas">—</strong>Rechazadas</div>
  </div>
</div>

<div class="tabs">
  <button class="tab active" onclick="showTab('facturas')">Facturas</button>
  <button class="tab" onclick="showTab('historial')">Historial</button>
</div>

<main>
  <div id="tab-facturas"></div>
  <div id="tab-historial" style="display:none"></div>
</main>

<div id="toast"></div>

<script>
let facturas = [];
let aprobaciones = [];

async function init() {
  const [fRes, aRes] = await Promise.all([fetch('/api/facturas'), fetch('/api/aprobaciones')]);
  facturas = await fRes.json();
  aprobaciones = await aRes.json();
  renderFacturas();
  renderHistorial();
  updateStats();
  document.getElementById('subtitulo').textContent =
    facturas.length + ' factura(s) en el lote actual';
}

function badgeEstado(estado, estadoDi) {
  let html = '';
  const map = {
    'CORRECTO':   ['badge-verde',   '✓ Comision OK'],
    'DISCREPANCIA':['badge-rojo',   '⚠ Discrepancia'],
    'OTA_DESCONOCIDA':['badge-gris','? OTA desconocida'],
    'SIN_PORCENTAJE':['badge-gris', '? Sin porcentaje'],
  };
  const [cls, lbl] = map[estado] || ['badge-gris', estado];
  html += `<span class="badge ${cls}">${lbl}</span>`;
  if (estadoDi === 'FALTA_CERTIFICADO_DI') {
    html += '<span class="badge badge-naranja">⚠ Falta cert. DI</span>';
  } else if (estadoDi === 'CERTIFICADO_OK') {
    html += '<span class="badge badge-azul">✓ Cert. DI OK</span>';
  } else if (estadoDi === 'NO_APLICA') {
    html += '<span class="badge badge-gris">DI: no aplica</span>';
  }
  return html;
}

function yaProcesada(numFactura) {
  return aprobaciones.find(a => a.numero_factura === numFactura);
}

function renderFacturas() {
  const container = document.getElementById('tab-facturas');
  if (!facturas.length) {
    container.innerHTML = '<div class="empty-state"><div class="emoji">📂</div><p>No hay facturas cargadas.<br>Ejecuta primero el pipeline completo.</p></div>';
    return;
  }
  container.innerHTML = facturas.map((f, i) => {
    const proc = yaProcesada(f.numero_factura);
    const tieneDisc = f.estado === 'DISCREPANCIA' && f.discrepancia_euros !== 'NO_ENCONTRADO';

    const detallesHtml = `
      <div class="detalle-grid">
        <div class="detalle-item">
          <div class="detalle-label">Periodo</div>
          <div class="detalle-val" style="font-size:12px">${f.periodo_inicio || '—'}<br>al ${f.periodo_fin || '—'}</div>
        </div>
        <div class="detalle-item">
          <div class="detalle-label">Comision</div>
          <div class="detalle-val">${f.porcentaje_factura !== 'NO_ENCONTRADO' ? f.porcentaje_factura+'%' : '—'}</div>
        </div>
        <div class="detalle-item">
          <div class="detalle-label">Pactado</div>
          <div class="detalle-val">${f.porcentaje_pactado !== 'None' && f.porcentaje_pactado !== 'NO_ENCONTRADO' ? f.porcentaje_pactado+'%' : '—'}</div>
        </div>
        <div class="detalle-item">
          <div class="detalle-label">Importe bruto</div>
          <div class="detalle-val" style="font-size:13px">${f.importe_bruto !== 'NO_ENCONTRADO' ? f.importe_bruto+' EUR' : '—'}</div>
        </div>
      </div>
      ${tieneDisc ? `<div class="disc-box"><div><div class="disc-label">IMPORTE RECLAMABLE</div></div><div class="disc-val">${f.discrepancia_euros} EUR</div></div>` : ''}
    `;

    const accionesHtml = proc
      ? `<div class="ya-procesada">
           <strong>${proc.accion === 'APROBADA' ? '✅ Aprobada' : '❌ Rechazada'}</strong>
           ${proc.fecha} ${proc.hora} — ${proc.comentario}
         </div>`
      : `<div class="comentario-wrap">
           <label>Comentario (obligatorio)</label>
           <textarea id="comentario-${i}" placeholder="Ej: Comision verificada con contrato firmado 2024-01..." rows="3"></textarea>
         </div>
         <div class="btn-row">
           <button class="btn btn-aprobar" onclick="accion(${i},'APROBADA')">
             <span id="sp-a-${i}" class="spinner"></span>✅ Aprobar
           </button>
           <button class="btn btn-rechazar" onclick="accion(${i},'RECHAZADA')">
             <span id="sp-r-${i}" class="spinner"></span>❌ Rechazar
           </button>
         </div>`;

    return `
      <div class="factura" id="card-${i}">
        <div class="factura-header">
          <div class="factura-ota">${f.nombre_ota}</div>
          <div class="factura-num">Factura ${f.numero_factura} · ${f.fecha || ''}</div>
          <div class="factura-hotel">${f.nombre_hotel}</div>
          <div class="badges">${badgeEstado(f.estado, f.estado_di)}</div>
        </div>
        <div class="factura-detalles">${detallesHtml}</div>
        <div class="factura-acciones">${accionesHtml}</div>
      </div>`;
  }).join('');
}

function renderHistorial() {
  const container = document.getElementById('tab-historial');
  if (!aprobaciones.length) {
    container.innerHTML = '<div class="empty-state"><div class="emoji">📋</div><p>Aun no hay aprobaciones registradas.</p></div>';
    return;
  }
  const sorted = [...aprobaciones].reverse();
  container.innerHTML = sorted.map(a => {
    const es_aprob = a.accion === 'APROBADA';
    return `
      <div class="hist-item">
        <div class="hist-icon ${es_aprob ? 'aprobada' : 'rechazada'}">${es_aprob ? '✅' : '❌'}</div>
        <div class="hist-body">
          <div class="hist-titulo">${a.ota} · ${a.numero_factura}</div>
          <div class="hist-meta">${a.hotel} · ${a.fecha} ${a.hora}</div>
          ${a.comentario ? `<div class="hist-comentario">${a.comentario}</div>` : ''}
        </div>
      </div>`;
  }).join('');
}

function updateStats() {
  const procesadas = aprobaciones.map(a => a.numero_factura);
  const pendientes = facturas.filter(f => !procesadas.includes(f.numero_factura)).length;
  const aprobadas  = aprobaciones.filter(a => a.accion === 'APROBADA').length;
  const rechazadas = aprobaciones.filter(a => a.accion === 'RECHAZADA').length;
  document.getElementById('stat-pendientes').textContent = pendientes;
  document.getElementById('stat-aprobadas').textContent  = aprobadas;
  document.getElementById('stat-rechazadas').textContent = rechazadas;
}

async function accion(i, tipo) {
  const f = facturas[i];
  const ta = document.getElementById('comentario-' + i);
  const comentario = ta.value.trim();
  if (!comentario) {
    ta.style.borderColor = '#dc2626';
    ta.focus();
    toast('Debes escribir un comentario antes de ' + (tipo === 'APROBADA' ? 'aprobar' : 'rechazar'));
    return;
  }
  ta.style.borderColor = '';

  // Mostrar spinner
  const spId = tipo === 'APROBADA' ? 'sp-a-' : 'sp-r-';
  const sp = document.getElementById(spId + i);
  if (sp) sp.style.display = 'block';

  try {
    const res = await fetch('/api/accion', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ indice: i, accion: tipo, comentario })
    });
    const data = await res.json();
    if (data.ok) {
      aprobaciones.push(data.registro);
      renderFacturas();
      renderHistorial();
      updateStats();
      toast(tipo === 'APROBADA' ? '✅ Factura aprobada' : '❌ Factura rechazada');
    } else {
      toast('Error: ' + data.error);
    }
  } catch(e) {
    toast('Error de conexion');
  }
}

function showTab(tab) {
  document.querySelectorAll('.tab').forEach((t, i) => t.classList.toggle('active', ['facturas','historial'][i] === tab));
  document.getElementById('tab-facturas').style.display  = tab === 'facturas'  ? '' : 'none';
  document.getElementById('tab-historial').style.display = tab === 'historial' ? '' : 'none';
}

let toastTimer;
function toast(msg) {
  const el = document.getElementById('toast');
  el.textContent = msg;
  el.classList.add('show');
  clearTimeout(toastTimer);
  toastTimer = setTimeout(() => el.classList.remove('show'), 3000);
}

init();
</script>
</body></html>"""


# ── API routes ──────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return HTML_BASE


@app.route("/api/facturas")
def api_facturas():
    df = cargar_facturas()
    if df.empty:
        return jsonify([])
    return jsonify(df.to_dict(orient="records"))


@app.route("/api/aprobaciones")
def api_aprobaciones():
    df = cargar_aprobaciones()
    if df.empty:
        return jsonify([])
    return jsonify(df.to_dict(orient="records"))


@app.route("/api/accion", methods=["POST"])
def api_accion():
    data = request.get_json()
    indice    = int(data.get("indice", 0))
    accion    = data.get("accion", "")
    comentario = data.get("comentario", "").strip()

    if accion not in ("APROBADA", "RECHAZADA"):
        return jsonify({"ok": False, "error": "Accion invalida"})
    if not comentario:
        return jsonify({"ok": False, "error": "Comentario obligatorio"})

    df = cargar_facturas()
    if indice >= len(df):
        return jsonify({"ok": False, "error": "Indice fuera de rango"})

    fila = df.iloc[indice]
    ahora = datetime.now()

    registro = {
        "fecha":                ahora.strftime("%d/%m/%Y"),
        "hora":                 ahora.strftime("%H:%M:%S"),
        "numero_factura":       str(fila.get("numero_factura", NF)),
        "ota":                  str(fila.get("nombre_ota", NF)),
        "hotel":                str(fila.get("nombre_hotel", NF)),
        "accion":               accion,
        "comentario":           comentario,
        "estado_verificacion":  str(fila.get("estado", NF)),
        "estado_di":            str(fila.get("estado_di", NF)),
    }

    try:
        guardar_aprobacion(registro)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

    return jsonify({"ok": True, "registro": registro})


# ── Arranque ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import socket

    # Obtener IP local para acceso desde movil en la misma red WiFi
    try:
        ip_local = socket.gethostbyname(socket.gethostname())
    except Exception:
        ip_local = "localhost"

    print("=" * 55)
    print("  Yve.01 — Dashboard de Aprobaciones AR")
    print("=" * 55)
    print(f"  Abre en tu navegador:  http://localhost:5000")
    print(f"  Movil (misma WiFi):    http://{ip_local}:5000")
    print()
    print("  Ctrl+C para detener el servidor")
    print("=" * 55)

    app.run(host="0.0.0.0", port=5000, debug=False)
