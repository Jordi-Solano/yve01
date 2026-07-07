"""
exportador_asientos.py — Exportador de Libro Diario (asientos contables)
Formato universal compatible con A3 Asesor, Sage, ContaPlus, Holded y Excel manual.

Endpoint: GET /api/exportar/asientos?mes=YYYY-MM
Devuelve: Excel con formato Libro Diario PGC español (Debe/Haber)
"""
import json, os
from datetime import datetime
from pathlib import Path
from io import BytesIO
from flask import Blueprint, request, send_file, jsonify

asientos_bp = Blueprint('asientos', __name__)

from tenant_dirs import datos_dir as _t_ddir

class _TData:
    def __truediv__(self, other): return Path(_t_ddir()) / other
    def __str__(self): return _t_ddir()

DATA = _TData()

# ── Mapa cuenta PGC por tipo de gasto/ingreso ────────────────────────────────
CUENTAS = {
    'proveedor_fb':       ('600', 'Compras mercaderías F&B'),
    'proveedor_servicios':('623', 'Servicios profesionales independientes'),
    'proveedor_arrend':   ('621', 'Arrendamientos y cánones'),
    'proveedor_otros':    ('629', 'Otros servicios'),
    'comision_ota':       ('628', 'Comisiones agencias y OTAs'),
    'iva_soportado':      ('472', 'H.P. IVA soportado'),
    'proveedores':        ('400', 'Proveedores'),
    'acreedores':         ('410', 'Acreedores por prestaciones de servicios'),
    'banco':              ('572', 'Bancos e instituciones de crédito'),
    'ingresos_hab':       ('705', 'Prestaciones de servicios — Alojamiento'),
    'ingresos_fb':        ('700', 'Ventas de mercaderías F&B'),
    'clientes':           ('430', 'Clientes'),
    'iva_repercutido':    ('477', 'H.P. IVA repercutido'),
}

def _cuenta(tipo):
    return CUENTAS.get(tipo, ('999', tipo))


def _get_proveedores():
    """Devuelve dict nombre_proveedor → {cuenta, tipo, iva}"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(DATA / 'proveedores.xlsx')
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).lower() for h in rows[0]]
        result = {}
        for row in rows[1:]:
            d = dict(zip(headers, row))
            nombre = str(d.get('nombre_proveedor', ''))
            result[nombre] = {
                'cuenta': str(int(d.get('cuenta_contable', 623))),
                'tipo': str(d.get('tipo', 'OTRAS')),
                'iva': float(d.get('porcentaje_iva_habitual', 21)),
            }
        return result
    except Exception:
        return {}


def _get_facturas_ap():
    """Lee facturas AP del JSON de archivos procesados."""
    try:
        arch = json.loads((DATA / 'archivos_procesados.json').read_text())
        if isinstance(arch, dict):
            facturas = arch.get('facturas_ap', arch.get('ap', []))
        else:
            facturas = arch if isinstance(arch, list) else []
        return facturas
    except Exception:
        return []


def _get_facturas_ar():
    """Lee facturas AR (comisiones OTA) del Excel de demo."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(DATA / 'facturas_ota_demo.xlsx')
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).lower().replace(' ', '_') for h in rows[0]]
        result = []
        for row in rows[1:]:
            if not any(row):
                continue
            d = dict(zip(headers, row))
            result.append(d)
        return result
    except Exception:
        return []


def _num(v, default=0.0):
    try:
        return float(str(v).replace('€','').replace(',','.').replace(' ','')) if v else default
    except Exception:
        return default


def generar_libro_diario(mes_filtro=None):
    """
    Genera lista de asientos en formato Libro Diario PGC.
    Cada asiento: {num, fecha, cuenta, desc_cuenta, concepto, debe, haber, documento}
    """
    proveedores = _get_proveedores()
    asientos = []
    num = 1

    # ── 1. FACTURAS AP (proveedores) ──────────────────────────────────────────
    facturas_ap = _get_facturas_ap()
    for f in facturas_ap:
        fecha_raw = f.get('fecha') or f.get('date') or datetime.now().strftime('%Y-%m-%d')
        try:
            fecha = str(fecha_raw)[:10]
        except Exception:
            fecha = datetime.now().strftime('%Y-%m-%d')

        if mes_filtro and not fecha.startswith(mes_filtro):
            continue

        proveedor = str(f.get('proveedor') or f.get('supplier') or 'Proveedor')
        n_factura = str(f.get('numero_factura') or f.get('invoice_number') or f'AP-{num}')
        base = _num(f.get('base_imponible') or f.get('base') or f.get('importe') or f.get('amount'))
        iva_pct = _num(f.get('iva_porcentaje') or f.get('iva') or 21)
        iva_amt = round(base * iva_pct / 100, 2)
        total = round(base + iva_amt, 2)

        tipo = proveedores.get(proveedor, {}).get('tipo', 'OTRAS')
        if tipo == 'FB':
            cta_gasto, desc_gasto = _cuenta('proveedor_fb')
        else:
            cta_gasto, desc_gasto = _cuenta('proveedor_servicios')

        cta_prov, desc_prov = _cuenta('proveedores')
        cta_iva, desc_iva = _cuenta('iva_soportado')
        concepto = f'Fra. {n_factura} — {proveedor}'

        # DEBE: gasto
        asientos.append({
            'num': num, 'fecha': fecha, 'cuenta': cta_gasto,
            'desc_cuenta': desc_gasto, 'concepto': concepto,
            'debe': base, 'haber': 0, 'documento': n_factura
        })
        # DEBE: IVA soportado
        if iva_amt > 0:
            asientos.append({
                'num': num, 'fecha': fecha, 'cuenta': cta_iva,
                'desc_cuenta': desc_iva, 'concepto': concepto,
                'debe': iva_amt, 'haber': 0, 'documento': n_factura
            })
        # HABER: proveedor (total)
        asientos.append({
            'num': num, 'fecha': fecha, 'cuenta': cta_prov,
            'desc_cuenta': desc_prov, 'concepto': concepto,
            'debe': 0, 'haber': total, 'documento': n_factura
        })
        num += 1

    # ── 2. COMISIONES OTA (AR) ────────────────────────────────────────────────
    facturas_ar = _get_facturas_ar()
    for f in facturas_ar:
        fecha = str(f.get('fecha') or f.get('date') or datetime.now().strftime('%Y-%m-%d'))[:10]
        if mes_filtro and not fecha.startswith(mes_filtro):
            continue

        ota = str(f.get('ota') or f.get('agencia') or 'OTA')
        n_factura = str(f.get('numero_factura') or f.get('invoice') or f'AR-{num}')
        total = _num(f.get('comision') or f.get('importe') or f.get('amount') or f.get('total'))
        if total == 0:
            continue
        base = round(total / 1.21, 2)
        iva_amt = round(total - base, 2)

        cta_gasto, desc_gasto = _cuenta('comision_ota')
        cta_acr, desc_acr = _cuenta('acreedores')
        cta_iva, desc_iva = _cuenta('iva_soportado')
        concepto = f'Comisión {ota} — {n_factura}'

        asientos.append({
            'num': num, 'fecha': fecha, 'cuenta': cta_gasto,
            'desc_cuenta': desc_gasto, 'concepto': concepto,
            'debe': base, 'haber': 0, 'documento': n_factura
        })
        if iva_amt > 0:
            asientos.append({
                'num': num, 'fecha': fecha, 'cuenta': cta_iva,
                'desc_cuenta': desc_iva, 'concepto': concepto,
                'debe': iva_amt, 'haber': 0, 'documento': n_factura
            })
        asientos.append({
            'num': num, 'fecha': fecha, 'cuenta': cta_acr,
            'desc_cuenta': desc_acr, 'concepto': concepto,
            'debe': 0, 'haber': total, 'documento': n_factura
        })
        num += 1

    return asientos


def exportar_excel(asientos, mes_filtro=None):
    """Genera Excel con Libro Diario en formato importable por cualquier software contable."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Hoja 1: Libro Diario ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Libro Diario'

    AZUL = '0F172A'
    AZUL2 = '1E3A5F'
    GRIS = 'F1F5F9'
    VERDE = '22C55E'
    bold_white = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    normal = Font(name='Calibri', size=10)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    right = Alignment(horizontal='right', vertical='center')

    # Cabecera empresa
    ws.merge_cells('A1:H1')
    ws['A1'] = f'LIBRO DIARIO — Yve.01  |  Período: {mes_filtro or "Todos"}  |  Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    ws['A1'].fill = PatternFill(fill_type='solid', fgColor=AZUL)
    ws['A1'].alignment = left
    ws.row_dimensions[1].height = 22

    # Cabeceras columnas
    headers = ['Nº Asiento', 'Fecha', 'Cuenta', 'Descripción cuenta', 'Concepto', 'Debe (€)', 'Haber (€)', 'Documento']
    widths =  [12,            12,      10,       28,                   42,         14,          14,           20]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = bold_white
        cell.fill = PatternFill(fill_type='solid', fgColor=AZUL2)
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 18

    # Datos
    total_debe = total_haber = 0
    for i, a in enumerate(asientos):
        row = i + 3
        fill = PatternFill(fill_type='solid', fgColor=GRIS) if i % 2 == 0 else PatternFill()
        valores = [
            a['num'], a['fecha'], a['cuenta'], a['desc_cuenta'],
            a['concepto'], a['debe'] or '', a['haber'] or '', a['documento']
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = normal
            cell.fill = fill
            cell.alignment = right if col in (6, 7) else (center if col in (1, 2, 3) else left)
            if col in (6, 7) and isinstance(val, (int, float)) and val:
                cell.number_format = '#,##0.00'
        total_debe += float(a.get('debe') or 0)
        total_haber += float(a.get('haber') or 0)

    # Fila total
    total_row = len(asientos) + 3
    ws.cell(row=total_row, column=5, value='TOTAL').font = Font(name='Calibri', bold=True, size=10)
    for col, val in [(6, total_debe), (7, total_haber)]:
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='1D4ED8')
        cell.alignment = right
        cell.number_format = '#,##0.00'

    # Verificación cuadre
    cuadra = abs(total_debe - total_haber) < 0.01
    ws.cell(row=total_row + 1, column=5,
            value='✓ CUADRADO' if cuadra else f'⚠ DESCUADRE: {total_debe - total_haber:.2f}€'
    ).font = Font(
        name='Calibri', bold=True, size=10,
        color='22C55E' if cuadra else 'EF4444'
    )

    # ── Hoja 2: Balance de comprobación ──────────────────────────────────────
    ws2 = wb.create_sheet('Balance Comprobación')
    ws2['A1'] = 'BALANCE DE COMPROBACIÓN'
    ws2['A1'].font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    ws2['A1'].fill = PatternFill(fill_type='solid', fgColor=AZUL)
    ws2.merge_cells('A1:E1')

    headers2 = ['Cuenta', 'Descripción', 'Debe acum.', 'Haber acum.', 'Saldo']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = bold_white
        cell.fill = PatternFill(fill_type='solid', fgColor=AZUL2)
        cell.alignment = center
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 16

    # Agrupar por cuenta
    from collections import defaultdict
    por_cuenta = defaultdict(lambda: {'desc': '', 'debe': 0, 'haber': 0})
    for a in asientos:
        por_cuenta[a['cuenta']]['desc'] = a['desc_cuenta']
        por_cuenta[a['cuenta']]['debe'] += float(a.get('debe') or 0)
        por_cuenta[a['cuenta']]['haber'] += float(a.get('haber') or 0)

    for i, (cta, vals) in enumerate(sorted(por_cuenta.items()), 1):
        saldo = vals['debe'] - vals['haber']
        row = i + 2
        fill = PatternFill(fill_type='solid', fgColor=GRIS) if i % 2 == 0 else PatternFill()
        ws2.cell(row=row, column=1, value=cta).font = normal
        ws2.cell(row=row, column=2, value=vals['desc']).font = normal
        ws2.cell(row=row, column=3, value=round(vals['debe'], 2)).number_format = '#,##0.00'
        ws2.cell(row=row, column=4, value=round(vals['haber'], 2)).number_format = '#,##0.00'
        saldo_cell = ws2.cell(row=row, column=5, value=round(saldo, 2))
        saldo_cell.number_format = '#,##0.00'
        saldo_cell.font = Font(name='Calibri', size=10,
                               color='22C55E' if saldo >= 0 else 'EF4444')
        for col in range(1, 6):
            ws2.cell(row=row, column=col).fill = fill
            ws2.cell(row=row, column=col).alignment = right if col > 2 else left

    # ── Hoja 3: Instrucciones importación ────────────────────────────────────
    ws3 = wb.create_sheet('Importación')
    instrucciones = [
        ('INSTRUCCIONES DE IMPORTACIÓN', True, AZUL),
        ('', False, None),
        ('Este archivo es compatible con los principales softwares contables españoles:', False, None),
        ('', False, None),
        ('A3 ASESOR (Wolters Kluwer)', True, AZUL2),
        ('1. Abrir A3 Asesor → Contabilidad → Importar → Libro Diario', False, None),
        ('2. Seleccionar formato Excel (.xlsx)', False, None),
        ('3. Mapear columnas: Fecha→Fecha, Cuenta→Cuenta, Debe→Debe, Haber→Haber', False, None),
        ('', False, None),
        ('SAGE 50 / CONTAPLUS', True, AZUL2),
        ('1. Menú Fichero → Importar → Asientos contables', False, None),
        ('2. Usar hoja "Libro Diario" de este archivo', False, None),
        ('3. Columnas en el orden: Nº Asiento, Fecha, Cuenta, Concepto, Debe, Haber', False, None),
        ('', False, None),
        ('HOLDED', True, AZUL2),
        ('1. Contabilidad → Asientos → Importar CSV', False, None),
        ('2. Guardar la hoja "Libro Diario" como CSV y subir', False, None),
        ('', False, None),
        ('EXCEL MANUAL', True, AZUL2),
        ('Usar directamente la hoja "Libro Diario" — ya está formateada.', False, None),
        ('', False, None),
        ('Para dudas: jordi@yve01.com', False, None),
    ]
    ws3.column_dimensions['A'].width = 70
    for i, (texto, bold, color) in enumerate(instrucciones, 1):
        cell = ws3.cell(row=i, column=1, value=texto)
        cell.font = Font(name='Calibri', bold=bold, size=11,
                         color='FFFFFF' if color else '334155')
        if color:
            cell.fill = PatternFill(fill_type='solid', fgColor=color)
        cell.alignment = left

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@asientos_bp.route('/api/exportar/asientos')
def exportar_asientos():
    mes = request.args.get('mes', '')  # formato YYYY-MM
    try:
        asientos = generar_libro_diario(mes_filtro=mes or None)
        if not asientos:
            # Generar asientos de demo si no hay datos reales
            asientos = _demo_asientos()
        output = exportar_excel(asientos, mes_filtro=mes or None)
        nombre = f"Yve_LibroDiario_{mes or datetime.now().strftime('%Y%m')}.xlsx"
        return send_file(output, as_attachment=True,
                         download_name=nombre,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _demo_asientos():
    """Asientos de demo cuando no hay datos reales cargados."""
    hoy = datetime.now().strftime('%Y-%m-%d')
    mes = datetime.now().strftime('%Y-%m')
    return [
        # Factura proveedor F&B
        {'num':1,'fecha':hoy,'cuenta':'600','desc_cuenta':'Compras mercaderías F&B',
         'concepto':'Fra. MAKRO-2024-001 — Makro Cash & Carry SL','debe':850.00,'haber':0,'documento':'MAKRO-2024-001'},
        {'num':1,'fecha':hoy,'cuenta':'472','desc_cuenta':'H.P. IVA soportado',
         'concepto':'Fra. MAKRO-2024-001 — Makro Cash & Carry SL','debe':85.00,'haber':0,'documento':'MAKRO-2024-001'},
        {'num':1,'fecha':hoy,'cuenta':'400','desc_cuenta':'Proveedores',
         'concepto':'Fra. MAKRO-2024-001 — Makro Cash & Carry SL','debe':0,'haber':935.00,'documento':'MAKRO-2024-001'},
        # Comisión Booking
        {'num':2,'fecha':hoy,'cuenta':'628','desc_cuenta':'Comisiones agencias y OTAs',
         'concepto':'Comisión Booking.com — BK-JUN-2024','debe':2310.00,'haber':0,'documento':'BK-JUN-2024'},
        {'num':2,'fecha':hoy,'cuenta':'472','desc_cuenta':'H.P. IVA soportado',
         'concepto':'Comisión Booking.com — BK-JUN-2024','debe':484.83,'haber':0,'documento':'BK-JUN-2024'},
        {'num':2,'fecha':hoy,'cuenta':'410','desc_cuenta':'Acreedores por prestaciones de servicios',
         'concepto':'Comisión Booking.com — BK-JUN-2024','debe':0,'haber':2794.83,'documento':'BK-JUN-2024'},
        # Pago proveedor
        {'num':3,'fecha':hoy,'cuenta':'400','desc_cuenta':'Proveedores',
         'concepto':'Pago Fra. MAKRO-2024-001','debe':935.00,'haber':0,'documento':'MAKRO-2024-001'},
        {'num':3,'fecha':hoy,'cuenta':'572','desc_cuenta':'Bancos e instituciones de crédito',
         'concepto':'Pago Fra. MAKRO-2024-001','debe':0,'haber':935.00,'documento':'MAKRO-2024-001'},
    ]
