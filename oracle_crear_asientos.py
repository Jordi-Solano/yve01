"""
oracle_crear_asientos.py — Yve.01 Módulo Oracle
Envía journal batches a Oracle GL REST API.
En modo simulación guarda los asientos en reportes/oracle_simulacion_[fecha].xlsx
sin realizar ninguna llamada real a Oracle.
"""

import os
import json
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

from oracle_auth import is_simulation, get_headers, ORACLE_BASE_URL

# ─── Rutas ────────────────────────────────────────────────────────────────────
BASE_DIR     = Path(__file__).parent
REPORTES_DIR = BASE_DIR / "reportes"
REPORTES_DIR.mkdir(exist_ok=True)
HOY          = datetime.now().strftime("%Y%m%d")

ORACLE_API_BATCHES = (
    f"{ORACLE_BASE_URL}/fscmRestApi/resources/11.13.18.05/journalBatches"
)

# Contador simulación
_sim_counter = [0]


def _sim_id() -> str:
    _sim_counter[0] += 1
    return f"SIM-{datetime.now().strftime('%Y%m%d')}-{_sim_counter[0]:03d}"


def _post_batch_real(payload: dict) -> dict:
    """Envía un journal batch a Oracle GL REST API (producción)."""
    try:
        import requests
        resp = requests.post(
            ORACLE_API_BATCHES,
            headers=get_headers(),
            json=payload,
            timeout=60,
        )
        if resp.status_code in (200, 201):
            data = resp.json()
            oracle_id = data.get("JournalBatchId") or data.get("BatchId") or data.get("id")
            return {"success": True, "oracle_id": str(oracle_id), "raw": data}
        else:
            return {
                "success": False,
                "error":   f"HTTP {resp.status_code}: {resp.text[:200]}",
                "oracle_id": None,
            }
    except Exception as e:
        return {"success": False, "error": str(e), "oracle_id": None}


def _post_batch_simulation(batch: dict) -> dict:
    """Simula el envío de un journal batch — devuelve ID ficticio."""
    sim_id = _sim_id()
    # Pequeña pausa visual
    import time; time.sleep(0.05)
    return {"success": True, "oracle_id": sim_id, "raw": {"simulado": True}}


def enviar_batch(batch: dict) -> dict:
    """
    Envía un journal batch a Oracle (real o simulación).
    Returns: { success, oracle_id, error (si falla) }
    """
    if is_simulation():
        return _post_batch_simulation(batch)
    else:
        return _post_batch_real(batch["oracle_payload"])


def procesar_batches(batches: list) -> list:
    """
    Procesa todos los batches. Devuelve lista de resultados enriquecidos.
    Cada resultado incluye todos los datos del batch + oracle_id + estado.
    """
    resultados = []

    for batch in batches:
        num_fac   = batch["numero_factura"]
        proveedor = batch["nombre_proveedor"]
        total     = batch["total_factura"]

        print(f"  → {num_fac} ({proveedor}) — {total:,.2f} EUR", end=" ... ", flush=True)

        resultado = enviar_batch(batch)

        if resultado["success"]:
            print(f"✓ Oracle ID: {resultado['oracle_id']}")
            estado = "CONTABILIZADA_SIM" if is_simulation() else "CONTABILIZADA"
        else:
            print(f"✗ ERROR: {resultado.get('error','')[:60]}")
            estado = "ERROR_ORACLE"

        resultados.append({
            "numero_factura":   num_fac,
            "nombre_proveedor": proveedor,
            "fecha":            batch["fecha"],
            "total_factura":    total,
            "batch_name":       batch["batch_name"],
            "oracle_id":        resultado.get("oracle_id") or "",
            "estado":           estado,
            "modo":             "SIMULACION" if is_simulation() else "PRODUCCION",
            "timestamp":        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "error":            resultado.get("error") or "",
            "journal_lines":    batch["journal_lines"],
        })

    return resultados


def guardar_simulacion_excel(resultados: list) -> str:
    """
    Guarda los asientos simulados en reportes/oracle_simulacion_[fecha].xlsx
    con todas las columnas: batch_id, journal_entry, cuenta, debe, haber, estado.
    """
    if not resultados:
        return ""

    import openpyxl
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Asientos_Oracle"

    # ── Estilos ──
    FILL_H  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FILL_ER = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    FILL_SIM= PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    FILL_ALT= PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    FH = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    FN = Font(name="Calibri", size=10)
    FB = Font(bold=True, name="Calibri", size=10)

    # Título
    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1, value="Yve.01 — Asientos Oracle (Simulación)")
    c.fill = FILL_H; c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=2, column=1).font = FN
    modo_txt = "SIMULACIÓN — sin credenciales Oracle" if is_simulation() else "PRODUCCIÓN"
    ws.cell(row=2, column=7, value=f"Modo: {modo_txt}").font = FB

    # Cabeceras
    headers = [
        "Batch ID (Oracle)", "Factura", "Proveedor", "Fecha",
        "Total EUR", "Línea", "Tipo", "Combinación Cuenta",
        "Debe EUR", "Haber EUR", "Descripción", "Estado",
    ]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.fill = FILL_H; c.font = FH
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[4].height = 18

    row_num = 5
    for res in resultados:
        is_ok  = res["estado"] in ("CONTABILIZADA", "CONTABILIZADA_SIM")
        fill_f = FILL_SIM if res["modo"] == "SIMULACION" else (FILL_OK if is_ok else FILL_ER)

        for i, line in enumerate(res["journal_lines"]):
            tipo    = line["type"]
            importe = line["debit"] if tipo == "DEBE" else line["credit"]
            debe    = line["debit"]
            haber   = line["credit"]
            alt     = FILL_ALT if row_num % 2 == 0 else None
            fill    = fill_f if i == 0 else alt

            vals = [
                res["oracle_id"] if i == 0 else "",
                res["numero_factura"] if i == 0 else "",
                res["nombre_proveedor"] if i == 0 else "",
                res["fecha"] if i == 0 else "",
                res["total_factura"] if i == 0 else None,
                line["line_number"],
                tipo,
                line["combination"],
                debe if debe > 0 else None,
                haber if haber > 0 else None,
                line["description"],
                res["estado"] if i == 0 else "",
            ]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=row_num, column=j, value=v)
                c.font = FB if i == 0 and j in (1, 2) else FN
                if fill:
                    c.fill = fill
                if j in (5, 9, 10) and isinstance(v, (int, float)):
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right")
            row_num += 1

    # Resumen al final
    row_num += 1
    ok_count  = sum(1 for r in resultados if r["estado"] in ("CONTABILIZADA","CONTABILIZADA_SIM"))
    err_count = len(resultados) - ok_count
    total_eur = sum(r["total_factura"] for r in resultados if r["estado"] in ("CONTABILIZADA","CONTABILIZADA_SIM"))

    ws.cell(row=row_num, column=1, value="RESUMEN:").font = FB
    ws.cell(row=row_num, column=2, value=f"{ok_count} contabilizadas / {err_count} errores").font = FN
    ws.cell(row=row_num, column=5, value=total_eur).number_format = "#,##0.00"
    ws.cell(row=row_num, column=5).font = FB

    # Anchos de columna
    from openpyxl.utils import get_column_letter
    anchos = [22, 22, 30, 12, 14, 8, 8, 28, 14, 14, 40, 20]
    for j, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ruta = str(REPORTES_DIR / f"oracle_simulacion_{HOY}.xlsx")
    wb.save(ruta)
    return ruta


if __name__ == "__main__":
    from oracle_auth import print_status
    from oracle_lector_facturas import preparar_facturas_para_oracle

    print("=" * 65)
    print("  Yve.01 — Oracle Crear Asientos")
    print("=" * 65)
    print_status()
    print()

    batches, bloqueadas, _ = preparar_facturas_para_oracle()
    print(f"  Batches a procesar: {len(batches)}")

    if not batches:
        print("  No hay facturas que procesar.")
    else:
        print()
        resultados = procesar_batches(batches)
        ok  = sum(1 for r in resultados if "CONTABILIZADA" in r["estado"])
        err = len(resultados) - ok
        print(f"\n  Resultado: {ok} OK / {err} errores")

        if is_simulation():
            ruta = guardar_simulacion_excel(resultados)
            print(f"  Excel simulación: {ruta}")

    print("=" * 65)
