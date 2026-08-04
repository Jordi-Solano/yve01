"""
detector_doble_imposicion.py — Yve.01
Detecta si las facturas de OTAs extranjeras incluyen certificado de doble imposición.
Lee el Excel de verificación, clasifica cada factura y actualiza el reporte.

Estados posibles:
  - CERTIFICADO_OK        → OTA extranjera con certificado detectado en el PDF
  - FALTA_CERTIFICADO_DI  → OTA extranjera sin certificado en el PDF
  - NO_APLICA             → OTA española o mercado nacional (no requiere certificado)
  - OTA_DESCONOCIDA       → OTA no reconocida, no se puede determinar
"""

import os
import glob
import re
from datetime import date
import pdfplumber
import pandas as pd

# ── Rutas ──────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Las rutas salen del arbol del TENANT, no de la raiz. Estaban clavadas aqui:
# con un segundo cliente, sus informes de doble imposicion se escribian y se
# leian del arbol del primero. Para el tenant `default`, `tenant_dirs` devuelve
# EXACTAMENTE estas mismas rutas, asi que hoy no cambia nada.
#
# Se resuelven al importar, y aqui vale: este script solo se ejecuta como
# subproceso (`python3 detector_doble_imposicion.py`), asi que importar y
# ejecutar son el mismo momento y el YVE_TENANT del entorno es el bueno. Es el
# mismo patron que `verificador_comisiones.py`.
from tenant_dirs import reportes_dir as _t_rdir, entrada_dir as _t_edir
REPORTES_DIR = _t_rdir()
ENTRADA_DIR = _t_edir()
os.makedirs(REPORTES_DIR, exist_ok=True)

FECHA_HOY = date.today().strftime("%Y%m%d")
REPORTE_DI_SALIDA = os.path.join(REPORTES_DIR, f"doble_imposicion_{FECHA_HOY}.xlsx")

NF = "NO_ENCONTRADO"

# ── OTAs clasificadas por mercado ──────────────────────────────────────────
# "Nacional" = española → NO_APLICA
# "Internacional" o "Latinoamerica" = extranjera → requiere certificado

OTAS_NACIONALES = {
    "booking.es",
}

OTAS_EXTRANJERAS = {
    "booking.com",
    "expedia",
    "hotels.com",
    "despegar",
    "airbnb",
    "agoda",
    "trip.com",
    "trivago",
    "hrs",
}

# ── Palabras clave que indican la presencia de un certificado de DI ────────
KEYWORDS_CERTIFICADO = [
    # Español
    r"certificado\s+de\s+doble\s+imposici[oó]n",
    r"doble\s+imposici[oó]n",
    r"exenci[oó]n\s+fiscal",
    r"convenio\s+de\s+doble\s+imposici[oó]n",
    r"residencia\s+fiscal",
    # Inglés
    r"double\s+taxation",
    r"tax\s+certificate",
    r"certificate\s+of\s+residence",
    r"wht\s+certificate",
    r"withholding\s+tax\s+certificate",
    r"tax\s+residency\s+certificate",
    r"relief\s+at\s+source",
    r"double\s+tax\s+relief",
    r"tax\s+treaty",
    # Francés / alemán (algunas OTAs emiten en otros idiomas)
    r"certificat\s+de\s+r[eé]sidence",
    r"steuerbescheinigung",
]

PATRON_CERTIFICADO = re.compile(
    "|".join(KEYWORDS_CERTIFICADO),
    re.IGNORECASE | re.MULTILINE,
)


def clasificar_mercado(nombre_ota: str, mercado: str) -> str:
    """
    Determina si la OTA es nacional o extranjera.
    Devuelve: 'nacional', 'extranjera', o 'desconocida'
    """
    if not isinstance(nombre_ota, str) or nombre_ota in (NF, ""):
        return "desconocida"

    ota_norm = nombre_ota.strip().lower()

    # Primero verificar en listas explícitas
    if ota_norm in OTAS_NACIONALES:
        return "nacional"
    if ota_norm in OTAS_EXTRANJERAS:
        return "extranjera"

    # Fallback: usar columna Mercado del verificador si existe
    if isinstance(mercado, str) and mercado not in (NF, ""):
        mercado_norm = mercado.strip().lower()
        if mercado_norm == "nacional":
            return "nacional"
        elif mercado_norm in ("internacional", "latinoamerica", "latinoamérica"):
            return "extranjera"

    return "desconocida"


def extraer_texto_pdf(archivo: str) -> str | None:
    """Extrae el texto completo de un PDF de la carpeta facturas-entrada."""
    # Buscar el archivo PDF por nombre (sin ruta)
    posibles = glob.glob(os.path.join(ENTRADA_DIR, archivo))
    if not posibles:
        # Intentar buscar solo por nombre de archivo
        nombre_base = os.path.basename(archivo)
        posibles = glob.glob(os.path.join(ENTRADA_DIR, nombre_base))

    if not posibles:
        return None

    pdf_path = posibles[0]
    try:
        textos = []
        with pdfplumber.open(pdf_path) as pdf:
            for pagina in pdf.pages:
                texto = pagina.extract_text()
                if texto:
                    textos.append(texto)
        return "\n".join(textos)
    except Exception as e:
        print(f"    ⚠️  Error al leer {pdf_path}: {e}")
        return None


def detectar_certificado_en_texto(texto: str) -> bool:
    """Devuelve True si el texto del PDF contiene alguna keyword de certificado."""
    return bool(PATRON_CERTIFICADO.search(texto))


def analizar_factura(fila: pd.Series) -> dict:
    """Analiza una fila del reporte de verificación y determina el estado DI."""
    archivo = str(fila.get("archivo", NF))
    nombre_ota = str(fila.get("nombre_ota", NF))
    mercado = str(fila.get("mercado", NF))
    estado_verificacion = str(fila.get("estado", NF))

    # Si la OTA es desconocida en el verificador, mantener estado
    if estado_verificacion == "OTA_DESCONOCIDA":
        return {
            **fila.to_dict(),
            "tipo_mercado": "desconocida",
            "certificado_encontrado": NF,
            "estado_di": "OTA_DESCONOCIDA",
            "keywords_detectadas": NF,
        }

    tipo_mercado = clasificar_mercado(nombre_ota, mercado)

    if tipo_mercado == "nacional":
        return {
            **fila.to_dict(),
            "tipo_mercado": "nacional",
            "certificado_encontrado": "N/A",
            "estado_di": "NO_APLICA",
            "keywords_detectadas": "N/A",
        }

    if tipo_mercado == "desconocida":
        return {
            **fila.to_dict(),
            "tipo_mercado": "desconocida",
            "certificado_encontrado": NF,
            "estado_di": "OTA_DESCONOCIDA",
            "keywords_detectadas": NF,
        }

    # OTA extranjera → buscar certificado en el PDF
    texto_pdf = extraer_texto_pdf(archivo)

    if texto_pdf is None:
        return {
            **fila.to_dict(),
            "tipo_mercado": "extranjera",
            "certificado_encontrado": "PDF_NO_ENCONTRADO",
            "estado_di": "FALTA_CERTIFICADO_DI",
            "keywords_detectadas": NF,
        }

    tiene_certificado = detectar_certificado_en_texto(texto_pdf)

    # Identificar qué keywords se encontraron (para trazabilidad)
    keywords_encontradas = []
    for kw_patron in KEYWORDS_CERTIFICADO:
        if re.search(kw_patron, texto_pdf, re.IGNORECASE):
            # Simplificar el patrón para mostrarlo
            kw_limpia = kw_patron.replace(r"\s+", " ").replace(r"[oó]", "o").replace(r"[eé]", "e")
            keywords_encontradas.append(kw_limpia[:40])

    keywords_str = " | ".join(keywords_encontradas) if keywords_encontradas else "ninguna"

    return {
        **fila.to_dict(),
        "tipo_mercado": "extranjera",
        "certificado_encontrado": "SÍ" if tiene_certificado else "NO",
        "estado_di": "CERTIFICADO_OK" if tiene_certificado else "FALTA_CERTIFICADO_DI",
        "keywords_detectadas": keywords_str,
    }


def cargar_ultimo_reporte_verificacion() -> pd.DataFrame:
    """Carga el Excel de verificación más reciente."""
    excels = sorted(
        glob.glob(os.path.join(REPORTES_DIR, "verificacion_*.xlsx")),
        reverse=True,
    )
    if not excels:
        raise FileNotFoundError(
            "No se encontró ningún archivo verificacion_*.xlsx en la carpeta reportes/.\n"
            "Ejecuta primero verificador_comisiones.py"
        )
    ruta = excels[0]
    print(f"  Cargando reporte de verificación: {os.path.basename(ruta)}")
    return pd.read_excel(ruta, sheet_name="Detalle")


def aplicar_formato_di(ws):
    """Aplica colores al Excel de doble imposición."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment

        VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        AZUL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        GRIS = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

        # Encontrar columna estado_di
        col_estado = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "estado_di":
                col_estado = idx
                break

        if col_estado is None:
            return

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            estado = ws.cell(row=row_idx, column=col_estado).value
            fill = {
                "CERTIFICADO_OK": VERDE,
                "FALTA_CERTIFICADO_DI": ROJO,
                "NO_APLICA": AZUL,
                "OTA_DESCONOCIDA": AMARILLO if False else GRIS,
            }.get(estado)
            if fill:
                for cell in row:
                    cell.fill = fill

        # Encabezado
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    except Exception:
        pass  # Si falla el formato, el archivo sigue siendo válido


def generar_resumen_di(df: pd.DataFrame) -> pd.DataFrame:
    """Genera un resumen del análisis de doble imposición."""
    total = len(df)
    conteos = df["estado_di"].value_counts().reset_index()
    conteos.columns = ["Estado DI", "Cantidad"]
    conteos["Descripción"] = conteos["Estado DI"].map({
        "CERTIFICADO_OK": "OTA extranjera con certificado ✓",
        "FALTA_CERTIFICADO_DI": "OTA extranjera SIN certificado — ACCIÓN REQUERIDA",
        "NO_APLICA": "OTA española — no requiere certificado",
        "OTA_DESCONOCIDA": "OTA no reconocida",
    }).fillna("")

    extra = pd.DataFrame([
        {"Estado DI": "─" * 25, "Cantidad": "", "Descripción": ""},
        {"Estado DI": "TOTAL FACTURAS", "Cantidad": total, "Descripción": ""},
    ])
    return pd.concat([conteos, extra], ignore_index=True)


def main():
    print("=" * 65)
    print("  Yve.01 — Detector de Doble Imposición")
    print("=" * 65)

    try:
        df_verificacion = cargar_ultimo_reporte_verificacion()
    except FileNotFoundError as e:
        print(f"\n❌ Error: {e}")
        return

    print(f"  Facturas a analizar: {len(df_verificacion)}\n")

    resultados = []
    for _, fila in df_verificacion.iterrows():
        resultado = analizar_factura(fila)

        estado_di = resultado["estado_di"]
        iconos = {
            "CERTIFICADO_OK": "✓",
            "FALTA_CERTIFICADO_DI": "✗",
            "NO_APLICA": "–",
            "OTA_DESCONOCIDA": "?",
        }
        icon = iconos.get(estado_di, "·")

        print(f"  [{icon}] {resultado['archivo']} ({resultado['nombre_ota']}) → {estado_di}")
        resultados.append(resultado)

    df_resultado = pd.DataFrame(resultados)
    df_resumen = generar_resumen_di(df_resultado)

    # Ordenar: primero las que requieren acción
    prioridad_orden = {
        "FALTA_CERTIFICADO_DI": 0,
        "OTA_DESCONOCIDA": 1,
        "CERTIFICADO_OK": 2,
        "NO_APLICA": 3,
    }
    df_resultado["_orden"] = df_resultado["estado_di"].map(prioridad_orden).fillna(9)
    df_resultado = df_resultado.sort_values("_orden").drop(columns=["_orden"])

    # Guardar Excel
    with pd.ExcelWriter(REPORTE_DI_SALIDA, engine="openpyxl") as writer:
        df_resultado.to_excel(writer, index=False, sheet_name="Detalle_DI")
        df_resumen.to_excel(writer, index=False, sheet_name="Resumen_DI")

        ws = writer.sheets["Detalle_DI"]
        aplicar_formato_di(ws)

        for sheet_name in ["Detalle_DI", "Resumen_DI"]:
            ws2 = writer.sheets[sheet_name]
            for col in ws2.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    # Resumen en consola
    faltan = df_resultado[df_resultado["estado_di"] == "FALTA_CERTIFICADO_DI"]
    print("\n" + "─" * 65)
    print("  RESUMEN DOBLE IMPOSICIÓN")
    print("─" * 65)
    for _, row in df_resumen.iterrows():
        if row["Estado DI"].startswith("─"):
            print("  " + "─" * 40)
        else:
            print(f"  {str(row['Estado DI']):<30} {str(row['Cantidad']):<6} {row.get('Descripción','')}")

    if not faltan.empty:
        print(f"\n⚠️  ATENCIÓN: {len(faltan)} factura(s) extranjera(s) sin certificado DI:")
        for _, r in faltan.iterrows():
            print(f"   → {r['archivo']} ({r['nombre_ota']})")

    print(f"\n✅ Reporte guardado en: {REPORTE_DI_SALIDA}")
    print("=" * 65)


if __name__ == "__main__":
    main()
