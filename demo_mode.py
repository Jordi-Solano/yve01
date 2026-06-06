"""
Demo Mode — Datos realistas ficticios para presentaciones a inversores
Simula: hoteles activos, facturas procesadas, alertas, flujos completos
"""
import json
from datetime import datetime, timedelta
import random

def generar_datos_demo():
    """Genera datos ficticios realistas para demostración"""
    
    # Hoteles demo
    hoteles_demo = [
        {
            "id": "DEMO01",
            "nombre": "Hotel Demo Barcelona Premium",
            "ciudad": "Barcelona",
            "categoria": "5 estrellas",
            "habitaciones": 250,
            "activo": True,
            "contacto": "carlos.garcia@hotelsdemo.es",
            "modulos": ["AR", "AP", "DRR", "FB", "Banco"],
            "ocupacion": 88.5,
            "adr": 285.50,
            "revenue_mtd": 1850400,
            "gop_pct": 42.3,
            "fb_cost_pct": 28.5,
            "ap_pendientes": 2,
            "ar_pendientes": 1,
            "status": "ok"
        },
        {
            "id": "DEMO02",
            "nombre": "Hotel Demo Valencia Beachfront",
            "ciudad": "Valencia",
            "categoria": "4 estrellas",
            "habitaciones": 180,
            "activo": True,
            "contacto": "elena.ruiz@hotelsdemo.es",
            "modulos": ["AR", "AP", "DRR", "Banco"],
            "ocupacion": 81.2,
            "adr": 165.00,
            "revenue_mtd": 780300,
            "gop_pct": 35.8,
            "fb_cost_pct": 0,
            "ap_pendientes": 4,
            "ar_pendientes": 2,
            "status": "warning"
        },
        {
            "id": "DEMO03",
            "nombre": "Hotel Demo Sevilla Historic",
            "ciudad": "Sevilla",
            "categoria": "4 estrellas",
            "habitaciones": 95,
            "activo": True,
            "contacto": "juan.lopez@hotelsdemo.es",
            "modulos": ["AR", "AP", "DRR", "FB"],
            "ocupacion": 92.7,
            "adr": 215.75,
            "revenue_mtd": 589200,
            "gop_pct": 38.2,
            "fb_cost_pct": 31.2,
            "ap_pendientes": 1,
            "ar_pendientes": 0,
            "status": "ok"
        }
    ]
    
    # Facturas AR (OTA) demo
    facturas_ar_demo = [
        {
            "ota": "Booking.com",
            "factura": "BKG-2025-06-4521",
            "importe": 4250.50,
            "fecha": "2025-06-05",
            "estado": "Procesada",
            "di_cert": "SI",
            "hotel": "DEMO01"
        },
        {
            "ota": "Expedia",
            "factura": "EXP-2025-06-8834",
            "importe": 3120.75,
            "fecha": "2025-06-04",
            "estado": "Pendiente Aprobación",
            "di_cert": "SI",
            "hotel": "DEMO01"
        },
        {
            "ota": "Hotels.com",
            "factura": "HTL-2025-06-5621",
            "importe": 2840.00,
            "fecha": "2025-06-03",
            "estado": "Procesada",
            "di_cert": "NO - Solicitar",
            "hotel": "DEMO02"
        },
        {
            "ota": "Booking.com",
            "factura": "BKG-2025-06-4522",
            "importe": 1950.25,
            "fecha": "2025-06-05",
            "estado": "Procesada",
            "di_cert": "SI",
            "hotel": "DEMO02"
        },
        {
            "ota": "Despegar",
            "factura": "DSP-2025-06-3345",
            "importe": 820.60,
            "fecha": "2025-06-02",
            "estado": "Procesada",
            "di_cert": "SI",
            "hotel": "DEMO03"
        }
    ]
    
    # Facturas AP (Proveedores) demo
    facturas_ap_demo = [
        {
            "proveedor": "Food Supply Co",
            "factura": "FAC-FSC-001-2025",
            "importe": 2450.00,
            "tipo": "F&B",
            "estado": "3-Way OK",
            "aprobada": "SI",
            "hotel": "DEMO01"
        },
        {
            "proveedor": "Cleaning Services Ltd",
            "factura": "FAC-CSL-002-2025",
            "importe": 1800.50,
            "tipo": "Servicios",
            "estado": "Pendiente albarán",
            "aprobada": "NO",
            "hotel": "DEMO01"
        },
        {
            "proveedor": "Maintenance Pro",
            "factura": "FAC-MNT-003-2025",
            "importe": 3200.00,
            "tipo": "Mantenimiento",
            "estado": "3-Way OK",
            "aprobada": "SI",
            "hotel": "DEMO02"
        },
        {
            "proveedor": "Utilities Provider",
            "factura": "FAC-UTL-004-2025",
            "importe": 950.75,
            "tipo": "Servicios",
            "estado": "3-Way OK",
            "aprobada": "SI",
            "hotel": "DEMO03"
        }
    ]
    
    # Alertas demo
    alertas_demo = [
        {"hotel": "DEMO02", "tipo": "Ocupación baja", "severidad": "warning", "mensaje": "Ocupación 81.2%, bajo presupuesto (85%)"},
        {"hotel": "DEMO01", "tipo": "Missing DI Cert", "severidad": "warning", "mensaje": "Hotels.com: Falta certificado de doble imposición"},
        {"hotel": "DEMO01", "tipo": "AP Discrepancia", "severidad": "info", "mensaje": "Cleaning Services: Albarán pendiente de validación"}
    ]
    
    # Métricas consolidadas
    consolidado = {
        "grupo": "Demo Hotels Group",
        "num_hoteles": 3,
        "total_rooms": 525,
        "total_revenue_mtd": 3219900,
        "avg_ocupacion": 87.5,
        "avg_adr": 222.08,
        "avg_revpar": 194.32,
        "total_gop": 1313910,
        "avg_gop_pct": 38.8,
        "facturas_ar_procesadas": 5,
        "facturas_ap_procesadas": 4,
        "alertas_activas": 3,
        "di_certificates_pending": 1,
        "oracle_integration": "Simulación activa"
    }
    
    return {
        "hoteles": hoteles_demo,
        "facturas_ar": facturas_ar_demo,
        "facturas_ap": facturas_ap_demo,
        "alertas": alertas_demo,
        "consolidado": consolidado,
        "timestamp": datetime.now().isoformat(),
        "version": "1.0"
    }

def exportar_demo_excel():
    """Crea un Excel con los datos demo"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    data = generar_datos_demo()
    wb = openpyxl.Workbook()
    
    # Sheet 1: Consolidado
    ws = wb.active
    ws.title = "Consolidado"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'] = "YVE DEMO — Demo Hotels Group"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    ws[f'A{row}'] = "MÉTRICA"
    ws[f'B{row}'] = "VALOR"
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
    
    row = 4
    for metric, value in data["consolidado"].items():
        ws[f'A{row}'] = metric.replace("_", " ").title()
        ws[f'B{row}'] = value
        row += 1
    
    # Sheet 2: Hoteles
    ws = wb.create_sheet("Hoteles")
    headers = ["Nombre", "Ciudad", "Rooms", "Ocupación %", "ADR", "Revenue MTD", "GOP%", "AP Pend", "AR Pend"]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col).value = header
        ws.cell(1, col).font = header_font
        ws.cell(1, col).fill = header_fill
    
    for row_idx, hotel in enumerate(data["hoteles"], 2):
        ws.cell(row_idx, 1).value = hotel["nombre"]
        ws.cell(row_idx, 2).value = hotel["ciudad"]
        ws.cell(row_idx, 3).value = hotel["habitaciones"]
        ws.cell(row_idx, 4).value = hotel["ocupacion"]
        ws.cell(row_idx, 5).value = hotel["adr"]
        ws.cell(row_idx, 6).value = hotel["revenue_mtd"]
        ws.cell(row_idx, 7).value = hotel["gop_pct"]
        ws.cell(row_idx, 8).value = hotel["ap_pendientes"]
        ws.cell(row_idx, 9).value = hotel["ar_pendientes"]
    
    # Sheet 3: Facturas AR
    ws = wb.create_sheet("Facturas AR")
    headers = ["OTA", "Factura #", "Importe EUR", "Fecha", "Estado", "DI Cert"]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col).value = header
        ws.cell(1, col).font = header_font
        ws.cell(1, col).fill = header_fill
    
    for row_idx, fact in enumerate(data["facturas_ar"], 2):
        ws.cell(row_idx, 1).value = fact["ota"]
        ws.cell(row_idx, 2).value = fact["factura"]
        ws.cell(row_idx, 3).value = fact["importe"]
        ws.cell(row_idx, 4).value = fact["fecha"]
        ws.cell(row_idx, 5).value = fact["estado"]
        ws.cell(row_idx, 6).value = fact["di_cert"]
    
    # Sheet 4: Facturas AP
    ws = wb.create_sheet("Facturas AP")
    headers = ["Proveedor", "Factura #", "Importe EUR", "Tipo", "Estado 3-Way", "Aprobada"]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col).value = header
        ws.cell(1, col).font = header_font
        ws.cell(1, col).fill = header_fill
    
    for row_idx, fact in enumerate(data["facturas_ap"], 2):
        ws.cell(row_idx, 1).value = fact["proveedor"]
        ws.cell(row_idx, 2).value = fact["factura"]
        ws.cell(row_idx, 3).value = fact["importe"]
        ws.cell(row_idx, 4).value = fact["tipo"]
        ws.cell(row_idx, 5).value = fact["estado"]
        ws.cell(row_idx, 6).value = fact["aprobada"]
    
    for ws in wb.sheetnames:
        for col in range(1, 10):
            wb[ws].column_dimensions[get_column_letter(col)].width = 18
    
    # Guardar
    output_file = f"/tmp/yve01/reportes/YVE_DEMO_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_file)
    
    return output_file

if __name__ == "__main__":
    print("DEMO MODE")
    print("=" * 70)
    data = generar_datos_demo()
    print(f"\nGrupo: {data['consolidado']['grupo']}")
    print(f"Hoteles: {data['consolidado']['num_hoteles']}")
    print(f"Revenue MTD: €{data['consolidado']['total_revenue_mtd']:,.0f}")
    print(f"Facturas AR procesadas: {data['consolidado']['facturas_ar_procesadas']}")
    print(f"Facturas AP procesadas: {data['consolidado']['facturas_ap_procesadas']}")
    print(f"Alertas activas: {data['consolidado']['alertas_activas']}")
    print("\n" + "=" * 70)
    
    excel_file = exportar_demo_excel()
    print(f"✓ Demo Excel generado: {excel_file}")
