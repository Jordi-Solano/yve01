"""
lector_drr.py — Yve.01
Lee el DRR (Daily Revenue Report) del hotel en formato .xlsm y extrae:
  - DAILY_MASTER: métricas KPI clave (Today / MTD / Full Month Forecast)
  - Hojas 1-31: Trial Balance completo por día (ACCOUNT NAME / DEBITS / CREDITS / Total)
  - CtaCble: mapping Entity + Department + Account → Line Description
  - Detecta días Out of Balance
  - Genera reportes/drr_procesado_[fecha].xlsx con 3 hojas
Uso: python lector_drr.py [ruta_al_xlsm]
"""

import os
import sys
import glob
from datetime import datetime, date
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─── Rutas ────────────────────────────────────────────────────────────────────

BASE_DIR     = Path(__file__).parent
REPORTES_DIR = BASE_DIR / "reportes"
REPORTES_DIR.mkdir(exist_ok=True)
HOY          = datetime.now().strftime("%Y%m%d")

# ─── Estilos para Excel de salida ─────────────────────────────────────────────

FILL_HEADER   = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
FILL_SECTION  = PatternFill(start_color="2D6A9F", end_color="2D6A9F", fill_type="solid")
FILL_OOB      = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_OK       = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_ALERT    = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ALT      = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
FONT_HEADER   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
FONT_SECTION  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
FONT_BOLD     = Font(bold=True, name="Calibri", size=10)
FONT_NORMAL   = Font(name="Calibri", size=10)

def _sf(v):
    """Safe float conversion."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if not isinstance(v, bool) else None
    s = str(v).replace("%", "").replace(",", "").replace("€", "").strip()
    if s in ("#VALUE!", "#DIV/0!", "#N/A", "#REF!", "", "None"):
        return None
    try:
        return float(s)
    except Exception:
        return None

def _pct(v):
    """Convert to percentage string."""
    f = _sf(v)
    if f is None:
        return "N/D"
    return f"{f*100:.1f}%" if abs(f) <= 1.0 else f"{f:.1f}%"

def _eur(v):
    """Format as EUR."""
    f = _sf(v)
    if f is None:
        return "N/D"
    return f"{f:,.0f} EUR"

# ═══════════════════════════════════════════════════════════════════════════════
# 1. DAILY_MASTER — KPI Metrics
# ═══════════════════════════════════════════════════════════════════════════════

# Mapping: row index (0-based) → metric name
DAILY_MASTER_METRICS = {
    10:  "Occupancy %",
    11:  "Rooms Occupied",
    12:  "ADR",
    13:  "Revenue PAR",
    14:  "Rooms Revenue",
    17:  "Food Revenue",
    18:  "Beverage Revenue",
    19:  "F&B Other",
    20:  "F&B Revenue Total",
    23:  "Telephone / Other",
    27:  "Total Revenue",
    28:  "Spend PAR",
    30:  "GOP",
    31:  "GOP %",
}

# Column indices in DAILY_MASTER (0-based, col A=0)
COL_TODAY     = 2   # C
COL_MTD       = 3   # D
COL_FORECAST  = 5   # F = Full Month Forecast
COL_BUDGET    = 6   # G

def leer_daily_master(wb):
    """Lee la hoja DAILY_MASTER y devuelve un dict de métricas."""
    if "DAILY_MASTER" not in wb.sheetnames:
        print("  ⚠  Hoja DAILY_MASTER no encontrada")
        return {}

    ws = wb["DAILY_MASTER"]
    rows = list(ws.iter_rows(min_row=1, max_row=45, max_col=10, values_only=True))

    metricas = {}
    for row_idx, nombre in DAILY_MASTER_METRICS.items():
        if row_idx >= len(rows):
            continue
        row = rows[row_idx]
        metricas[nombre] = {
            "today":    row[COL_TODAY]    if len(row) > COL_TODAY    else None,
            "mtd":      row[COL_MTD]      if len(row) > COL_MTD      else None,
            "forecast": row[COL_FORECAST] if len(row) > COL_FORECAST else None,
            "budget":   row[COL_BUDGET]   if len(row) > COL_BUDGET   else None,
        }

    # Obtener fecha del informe
    fecha_row = rows[1] if len(rows) > 1 else []
    fecha_informe = None
    for v in fecha_row:
        if isinstance(v, datetime):
            fecha_informe = v.date()
            break
    metricas["__fecha_informe__"] = fecha_informe

    # ── GOP fallback: si Today/MTD son None pero Forecast está disponible ──────
    # Las celdas GOP Today/MTD en DAILY_MASTER a veces son fórmulas con
    # referencias cruzadas que openpyxl no puede evaluar (data_only=True).
    # Estimamos a partir del GOP% implícito en Forecast vs Total Revenue Forecast.
    _gop = metricas.get("GOP", {})
    _gop_pct = metricas.get("GOP %", {})
    _rev = metricas.get("Total Revenue", {})

    if _gop.get("today") is None or _gop.get("mtd") is None:
        # Intentar obtener GOP% del forecast
        pct_f = _sf(_gop_pct.get("forecast")) if _gop_pct.get("forecast") is not None else None
        if pct_f is not None and abs(pct_f) <= 1:
            pct_decimal = pct_f
        elif pct_f is not None:
            pct_decimal = pct_f / 100
        else:
            # Calcular de GOP Forecast / Total Revenue Forecast
            gop_f = _sf(_gop.get("forecast"))
            rev_f = _sf(_rev.get("forecast"))
            pct_decimal = (gop_f / rev_f) if gop_f and rev_f and rev_f != 0 else None

        if pct_decimal is not None:
            rev_today = _sf(_rev.get("today"))
            rev_mtd   = _sf(_rev.get("mtd"))
            if _gop.get("today") is None and rev_today:
                metricas["GOP"]["today"] = rev_today * pct_decimal
            if _gop.get("mtd") is None and rev_mtd:
                metricas["GOP"]["mtd"] = rev_mtd * pct_decimal
            if _gop_pct.get("today") is None:
                metricas.setdefault("GOP %", {})["today"] = pct_decimal
            if _gop_pct.get("mtd") is None:
                metricas.setdefault("GOP %", {})["mtd"] = pct_decimal

    return metricas


# ═══════════════════════════════════════════════════════════════════════════════
# 2. TRIAL BALANCE — Hojas diarias
# ═══════════════════════════════════════════════════════════════════════════════

# The Trial Balance data is in columns:
#   A (idx 0) = section header (ASSETS, LIABILITIES, EXPENSES, INCOME)
#   C (idx 2) = ACCOUNT NAME
#   H (idx 7) = DEBITS
#   I (idx 8) = CREDITS
#   J (idx 9) = Total
# Out of Balance indicator is in R01 cells D (idx 3) / E (idx 4)
# Real account rows end around row 244; room list starts ~row 1004

SECTIONS = {"ASSETS", "LIABILITIES", "EXPENSES", "INCOME", "REVENUE"}
MAX_TB_ROW = 300  # Read up to this row for the Trial Balance (before room list)

def leer_trial_balance_dia(wb, dia: int):
    """
    Lee la hoja de un día y devuelve dict con:
      {
        'dia': int,
        'fecha': date | None,
        'out_of_balance': bool,
        'imbalance': float,
        'total_debits': float,
        'total_credits': float,
        'cuentas': list[dict]  — cada cuenta con seccion, nombre, debits, credits, total
      }
    """
    nombre_hoja = str(dia)
    if nombre_hoja not in wb.sheetnames:
        return None

    ws = wb[nombre_hoja]
    rows = list(ws.iter_rows(min_row=1, max_row=MAX_TB_ROW, max_col=10, values_only=True))

    # Detectar Out of Balance desde R01 col D
    # El archivo marca "Out of Balance" o "OK" — sólo confiamos en el texto
    r01 = rows[0] if rows else []
    oob_text  = r01[3] if len(r01) > 3 else None   # col D = "Out of Balance" | "OK"
    imbalance = _sf(r01[4]) if len(r01) > 4 else 0.0  # col E = valor de diferencia
    out_of_balance = (
        isinstance(oob_text, str) and "out of balance" in oob_text.lower()
    )

    # Fecha del día — R03 col D (index 3) es la fecha real de la hoja
    fecha_dia = None
    if len(rows) >= 3:
        r03 = rows[2]  # row index 2 = row 3
        # col D = index 3
        v = r03[3] if len(r03) > 3 else None
        if isinstance(v, datetime):
            fecha_dia = v.date()
    # Fallback: any datetime in first 3 rows
    if fecha_dia is None:
        for r in rows[:3]:
            for v in r:
                if isinstance(v, datetime):
                    fecha_dia = v.date()
                    break
            if fecha_dia:
                break

    # Extraer cuentas del Trial Balance
    seccion_actual = "UNKNOWN"
    cuentas = []
    total_debits  = 0.0
    total_credits = 0.0

    for i, row in enumerate(rows):
        # Detectar encabezado de sección (col A)
        if row[0] and isinstance(row[0], str) and row[0].upper() in SECTIONS:
            seccion_actual = row[0].upper()
            continue

        account_name = row[2]  # col C
        debits       = _sf(row[7]) if len(row) > 7 else None   # col H
        credits      = _sf(row[8]) if len(row) > 8 else None   # col I
        total_val    = _sf(row[9]) if len(row) > 9 else None   # col J

        if not account_name or not isinstance(account_name, str):
            continue
        account_name = account_name.strip()
        if len(account_name) < 2 or account_name in ("ACCOUNT NAME",):
            continue

        # La fila TOTAL : es el total del balance
        if account_name.startswith("TOTAL"):
            if debits is not None:
                total_debits  = debits
            if credits is not None:
                total_credits = credits
            continue

        # Ignorar filas con valores no numéricos en debits/credits (lookup errors)
        if debits is None and credits is None:
            continue

        cuentas.append({
            "dia":          dia,
            "fecha":        fecha_dia.isoformat() if fecha_dia else f"Jul-{dia:02d}",
            "seccion":      seccion_actual,
            "cuenta":       account_name,
            "debits":       debits or 0.0,
            "credits":      credits or 0.0,
            "total":        total_val,
            "out_of_balance": out_of_balance,
        })

    # Si no se encontró fila TOTAL, calcular
    if total_debits == 0:
        total_debits  = sum(r["debits"]  for r in cuentas)
    if total_credits == 0:
        total_credits = sum(r["credits"] for r in cuentas)

    return {
        "dia":           dia,
        "fecha":         fecha_dia,
        "out_of_balance": out_of_balance,
        "imbalance":     imbalance or round(total_debits - total_credits, 4),
        "total_debits":  round(total_debits, 4),
        "total_credits": round(total_credits, 4),
        "num_cuentas":   len(cuentas),
        "cuentas":       cuentas,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 3. CtaCble — Account mapping
# ═══════════════════════════════════════════════════════════════════════════════

def leer_ctacble(wb):
    """
    Lee la hoja CtaCble y devuelve un dict:
      { (entity, dept, account) -> line_description }
    También devuelve un DataFrame con todas las filas.
    """
    if "CtaCble" not in wb.sheetnames:
        print("  ⚠  Hoja CtaCble no encontrada")
        return {}, pd.DataFrame()

    ws = wb["CtaCble"]
    rows = list(ws.iter_rows(max_col=8, values_only=True))

    # Buscar la fila de cabecera (contiene 'Entity')
    header_idx = None
    for i, row in enumerate(rows):
        if any(str(v).strip().lower() == "entity" for v in row if v is not None):
            header_idx = i
            break

    if header_idx is None:
        return {}, pd.DataFrame()

    header = [str(v).strip() if v is not None else f"Col{j}" for j, v in enumerate(rows[header_idx])]
    data_rows = []
    mapping = {}

    for row in rows[header_idx + 1:]:
        if all(v is None for v in row):
            continue
        entity  = str(row[0]).strip() if row[0] is not None else ""
        dept    = str(row[1]).strip() if row[1] is not None else ""
        account = str(row[2]).strip() if row[2] is not None else ""
        line_desc = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""

        if entity and account:
            key = (entity, dept, account)
            if key not in mapping:
                mapping[key] = line_desc

        row_dict = {header[j]: row[j] for j in range(min(len(header), len(row)))}
        data_rows.append(row_dict)

    df = pd.DataFrame(data_rows)
    return mapping, df


# ═══════════════════════════════════════════════════════════════════════════════
# 4. GENERACIÓN DEL EXCEL DE SALIDA
# ═══════════════════════════════════════════════════════════════════════════════

def escribir_hoja_resumen(ws_out, metricas):
    """Hoja 1: Resumen con métricas DAILY_MASTER."""

    def apply_header(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill   = FILL_HEADER
        c.font   = FONT_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")

    def apply_section(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = FILL_SECTION
        c.font = FONT_SECTION

    # Título
    ws_out.merge_cells("A1:E1")
    c = ws_out.cell(row=1, column=1, value="Yve.01 — Daily Revenue Report: Resumen KPIs")
    c.fill = FILL_HEADER
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_out.row_dimensions[1].height = 30

    fecha_inf = metricas.get("__fecha_informe__")
    ws_out.cell(row=2, column=1, value=f"Fecha informe: {fecha_inf or 'N/D'}")
    ws_out.cell(row=2, column=1).font = FONT_BOLD
    ws_out.cell(row=2, column=4, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws_out.cell(row=2, column=4).font = FONT_NORMAL

    # Cabeceras tabla
    headers = ["Métrica", "Today", "MTD", "Full Month Forecast", "Budget"]
    for j, h in enumerate(headers, 1):
        apply_header(ws_out, 4, j, h)
    ws_out.row_dimensions[4].height = 20

    # Filas de métricas
    FORMATO_PCT  = {"Occupancy %", "GOP %"}
    FORMATO_EUR  = {"Rooms Revenue", "Food Revenue", "Beverage Revenue", "F&B Revenue Total",
                    "Total Revenue", "GOP", "Telephone / Other", "F&B Other"}

    row_num = 5
    for nombre, vals in metricas.items():
        if nombre.startswith("__"):
            continue
        today    = vals.get("today")
        mtd      = vals.get("mtd")
        forecast = vals.get("forecast")
        budget   = vals.get("budget")

        # Formatear
        def fmt(v, nm):
            if v is None:
                return "N/D"
            if nm in FORMATO_PCT:
                f = _sf(v)
                return f"{f*100:.2f}%" if f is not None and abs(f) <= 1 else (f"{f:.1f}%" if f else "N/D")
            elif nm in FORMATO_EUR:
                f = _sf(v)
                return f"{f:,.0f} EUR" if f is not None else "N/D"
            else:
                f = _sf(v)
                return f"{f:,.2f}" if f is not None else str(v)

        fill = FILL_ALT if row_num % 2 == 0 else None

        for j, val_str in enumerate([nombre, fmt(today, nombre), fmt(mtd, nombre),
                                      fmt(forecast, nombre), fmt(budget, nombre)], 1):
            c = ws_out.cell(row=row_num, column=j, value=val_str)
            c.font = FONT_BOLD if j == 1 else FONT_NORMAL
            c.alignment = Alignment(horizontal="left" if j == 1 else "right")
            if fill:
                c.fill = fill

        row_num += 1

    # Ajustar anchos
    ws_out.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E"]:
        ws_out.column_dimensions[col].width = 22


def escribir_hoja_trial_balance(ws_out, todos_dias):
    """Hoja 2: Trial Balance completo de todos los días."""

    headers = ["Día", "Fecha", "Sección", "Cuenta", "Débitos", "Créditos", "Total", "Out of Balance"]
    for j, h in enumerate(headers, 1):
        c = ws_out.cell(row=1, column=j, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")

    row_num = 2
    prev_dia = None
    for dia_data in todos_dias:
        if dia_data is None:
            continue
        for cuenta in dia_data["cuentas"]:
            is_oob = dia_data["out_of_balance"]
            fill = FILL_OOB if is_oob else (FILL_ALT if row_num % 2 == 0 else None)

            values = [
                cuenta["dia"],
                cuenta["fecha"],
                cuenta["seccion"],
                cuenta["cuenta"],
                cuenta["debits"],
                cuenta["credits"],
                cuenta["total"],
                "⚠ OOB" if is_oob else "✓ OK",
            ]
            for j, v in enumerate(values, 1):
                c = ws_out.cell(row=row_num, column=j, value=v)
                c.font = FONT_NORMAL
                if fill:
                    c.fill = fill
                if j in (5, 6, 7):
                    c.alignment = Alignment(horizontal="right")
                    if isinstance(v, (int, float)):
                        c.number_format = '#,##0.00'
            row_num += 1

    ws_out.column_dimensions["A"].width  = 6
    ws_out.column_dimensions["B"].width  = 12
    ws_out.column_dimensions["C"].width  = 14
    ws_out.column_dimensions["D"].width  = 45
    ws_out.column_dimensions["E"].width  = 14
    ws_out.column_dimensions["F"].width  = 14
    ws_out.column_dimensions["G"].width  = 14
    ws_out.column_dimensions["H"].width  = 14


def escribir_hoja_alertas(ws_out, todos_dias, metricas):
    """Hoja 3: Alertas — días OOB + anomalías."""

    # Título
    ws_out.merge_cells("A1:G1")
    c = ws_out.cell(row=1, column=1, value="Alertas y Anomalías DRR")
    c.fill = FILL_HEADER
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    c.alignment = Alignment(horizontal="center")
    ws_out.row_dimensions[1].height = 25

    # Sección: Días Out of Balance
    ws_out.cell(row=3, column=1, value="DÍAS OUT OF BALANCE").fill = FILL_SECTION
    ws_out.cell(row=3, column=1).font = FONT_SECTION
    ws_out.merge_cells("A3:G3")

    headers_oob = ["Día", "Fecha", "Total Débitos", "Total Créditos", "Diferencia", "Estado", "Notas"]
    for j, h in enumerate(headers_oob, 1):
        c = ws_out.cell(row=4, column=j, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")

    row_num = 5
    oob_count = 0
    dias_revenue = []  # para análisis de anomalías de ingresos

    for dia_data in todos_dias:
        if dia_data is None:
            continue

        is_oob   = dia_data["out_of_balance"]
        imb      = dia_data["imbalance"] or 0.0
        estado   = "⚠ OUT OF BALANCE" if is_oob else "✓ IN BALANCE"
        fill_row = FILL_OOB if is_oob else FILL_OK
        nota     = f"Diferencia: {abs(imb):,.2f} EUR" if is_oob else ""

        row_vals = [
            dia_data["dia"],
            dia_data["fecha"].isoformat() if dia_data["fecha"] else f"Día {dia_data['dia']}",
            dia_data["total_debits"],
            dia_data["total_credits"],
            round(imb, 4),
            estado,
            nota,
        ]
        for j, v in enumerate(row_vals, 1):
            c = ws_out.cell(row=row_num, column=j, value=v)
            c.fill = fill_row
            c.font = FONT_BOLD if is_oob else FONT_NORMAL
            if j in (3, 4, 5):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right")
        if is_oob:
            oob_count += 1
        row_num += 1

    # Sección: Métricas clave del mes
    row_num += 2
    ws_out.cell(row=row_num, column=1, value="MÉTRICAS CLAVE DEL MES").fill = FILL_SECTION
    ws_out.cell(row=row_num, column=1).font = FONT_SECTION
    ws_out.merge_cells(f"A{row_num}:G{row_num}")
    row_num += 1

    metricas_clave = ["Total Revenue", "Occupancy %", "ADR", "GOP", "GOP %",
                      "Rooms Revenue", "F&B Revenue Total"]
    for nm in metricas_clave:
        if nm not in metricas:
            continue
        vals = metricas[nm]
        mtd  = vals.get("mtd")
        fore = vals.get("forecast")
        bud  = vals.get("budget")

        mtd_f  = _sf(mtd)
        fore_f = _sf(fore)
        bud_f  = _sf(bud)

        vs_budget = ""
        if mtd_f is not None and bud_f is not None and bud_f != 0:
            diff = (mtd_f - bud_f) / abs(bud_f)
            vs_budget = f"{'▲' if diff >= 0 else '▼'} {abs(diff)*100:.1f}% vs Budget"

        fill_alerta = FILL_ALERT if vs_budget.startswith("▼") else (FILL_OK if vs_budget.startswith("▲") else None)

        row_vals = [nm, "MTD", mtd, "Forecast", fore, "Budget", bud]
        for j, v in enumerate([nm, str(mtd or "N/D"), str(fore or "N/D"),
                                 str(bud or "N/D"), vs_budget], 1):
            c = ws_out.cell(row=row_num, column=j, value=v)
            c.font = FONT_BOLD if j == 1 else FONT_NORMAL
            if fill_alerta:
                c.fill = fill_alerta
        row_num += 1

    # Resumen final
    row_num += 2
    ws_out.cell(row=row_num, column=1, value="RESUMEN GENERAL")
    ws_out.cell(row=row_num, column=1).font = FONT_BOLD
    row_num += 1
    dias_procesados = sum(1 for d in todos_dias if d is not None)
    ws_out.cell(row=row_num, column=1, value=f"Días procesados: {dias_procesados}")
    row_num += 1
    ws_out.cell(row=row_num, column=1, value=f"Días Out of Balance: {oob_count}")
    if oob_count > 0:
        ws_out.cell(row=row_num, column=1).fill = FILL_OOB
        ws_out.cell(row=row_num, column=1).font = FONT_BOLD
    row_num += 1
    dias_ok = dias_procesados - oob_count
    ws_out.cell(row=row_num, column=1, value=f"Días In Balance: {dias_ok}")
    if dias_ok == dias_procesados:
        ws_out.cell(row=row_num, column=1).fill = FILL_OK

    ws_out.column_dimensions["A"].width = 28
    ws_out.column_dimensions["B"].width = 14
    ws_out.column_dimensions["C"].width = 18
    ws_out.column_dimensions["D"].width = 18
    ws_out.column_dimensions["E"].width = 16
    ws_out.column_dimensions["F"].width = 26
    ws_out.column_dimensions["G"].width = 30


# ═══════════════════════════════════════════════════════════════════════════════
# 5. MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 65)
    print("  Yve.01 — Lector DRR (Daily Revenue Report)")
    print("=" * 65)

    # Determinar ruta del archivo
    if len(sys.argv) > 1:
        ruta_drr = sys.argv[1]
    else:
        # Buscar automáticamente en el directorio del proyecto
        candidatos = sorted(
            glob.glob(str(BASE_DIR / "**/*.xlsm"), recursive=True) +
            glob.glob(str(BASE_DIR / "*.xlsm"))
        )
        if not candidatos:
            print("  ✗ No se encontró ningún .xlsm en el proyecto.")
            print("  Uso: python lector_drr.py ruta/al/DailyReport.xlsm")
            return
        ruta_drr = candidatos[0]

    ruta_drr = str(ruta_drr).strip('"').strip("'")
    if not os.path.exists(ruta_drr):
        print(f"  ✗ Archivo no encontrado: {ruta_drr}")
        return

    nombre_archivo = os.path.basename(ruta_drr)
    print(f"  Archivo: {nombre_archivo}")
    print(f"  Ruta:    {ruta_drr}")
    print()

    # Abrir workbook
    print("  Abriendo workbook (puede tardar unos segundos)...")
    wb = openpyxl.load_workbook(ruta_drr, read_only=True, keep_vba=True, data_only=True)
    print(f"  Hojas disponibles: {len(wb.sheetnames)}")

    # ── 1. DAILY_MASTER ──
    print("\n  [1/4] Leyendo DAILY_MASTER...")
    metricas = leer_daily_master(wb)
    fecha_inf = metricas.get("__fecha_informe__")
    print(f"        Fecha informe: {fecha_inf or 'N/D'}")
    print(f"        Métricas extraídas: {sum(1 for k in metricas if not k.startswith('__'))}")

    # ── 2. HOJAS DIARIAS (1-31) ──
    print("\n  [2/4] Leyendo Trial Balance diario...")
    dias_disponibles = [s for s in wb.sheetnames if s.isdigit() and 1 <= int(s) <= 31]
    dias_disponibles.sort(key=int)
    print(f"        Hojas de días encontradas: {dias_disponibles}")

    todos_dias = []
    dias_oob   = []
    total_cuentas_extraidas = 0

    for dia_str in dias_disponibles:
        dia_int  = int(dia_str)
        dia_data = leer_trial_balance_dia(wb, dia_int)
        if dia_data:
            todos_dias.append(dia_data)
            total_cuentas_extraidas += dia_data["num_cuentas"]
            estado = "⚠ OOB" if dia_data["out_of_balance"] else "✓"
            if dia_data["out_of_balance"]:
                dias_oob.append(dia_data)
                print(f"        Día {dia_int:2d}: {estado}  |  {dia_data['num_cuentas']:3d} cuentas  "
                      f"|  Dif: {dia_data['imbalance']:,.2f} EUR")
            else:
                print(f"        Día {dia_int:2d}: {estado}   |  {dia_data['num_cuentas']:3d} cuentas")

    print(f"\n        Total días: {len(todos_dias)} | "
          f"Out of Balance: {len(dias_oob)} | "
          f"Total líneas TB: {total_cuentas_extraidas:,}")

    # ── 3. CtaCble ──
    print("\n  [3/4] Leyendo CtaCble (Oracle account mapping)...")
    mapping_cta, df_cta = leer_ctacble(wb)
    print(f"        Mappings Entity+Dept+Account → Line Description: {len(mapping_cta)}")

    # ── 4. GENERAR EXCEL ──
    print("\n  [4/4] Generando Excel de salida...")
    nombre_salida = f"drr_procesado_{HOY}.xlsx"
    ruta_salida   = str(REPORTES_DIR / nombre_salida)

    wb_out = openpyxl.Workbook()

    # Hoja 1: Resumen
    ws_resumen = wb_out.active
    ws_resumen.title = "Resumen"
    escribir_hoja_resumen(ws_resumen, metricas)
    print(f"        ✓ Hoja 'Resumen' creada")

    # Hoja 2: Trial Balance completo
    ws_tb = wb_out.create_sheet("Trial_Balance_Completo")
    escribir_hoja_trial_balance(ws_tb, todos_dias)
    print(f"        ✓ Hoja 'Trial_Balance_Completo' creada ({total_cuentas_extraidas:,} filas)")

    # Hoja 3: Alertas
    ws_alertas = wb_out.create_sheet("Alertas")
    escribir_hoja_alertas(ws_alertas, todos_dias, metricas)
    print(f"        ✓ Hoja 'Alertas' creada")

    # Hoja 4: CtaCble mapping
    if not df_cta.empty:
        ws_cta = wb_out.create_sheet("CtaCble_Mapping")
        for j, col in enumerate(df_cta.columns, 1):
            c = ws_cta.cell(row=1, column=j, value=col)
            c.fill = FILL_HEADER
            c.font = FONT_HEADER
        for i, row in df_cta.iterrows():
            for j, val in enumerate(row, 1):
                ws_cta.cell(row=i+2, column=j, value=val).font = FONT_NORMAL
        print(f"        ✓ Hoja 'CtaCble_Mapping' creada ({len(df_cta)} filas)")

    wb_out.save(ruta_salida)
    print(f"\n  ✅ Excel guardado: {ruta_salida}")

    # ── RESUMEN EN CONSOLA ──────────────────────────────────────────────────
    print()
    print("=" * 65)
    print("  RESUMEN DEL DRR")
    print("=" * 65)
    print(f"  Días procesados:      {len(todos_dias)}")
    print(f"  Todos in balance:     {'✅ SÍ' if not dias_oob else f'❌ NO — {len(dias_oob)} día(s) OOB'}")
    if dias_oob:
        for d in dias_oob:
            print(f"    ⚠  Día {d['dia']:2d} ({d['fecha']}): diferencia {d['imbalance']:,.2f} EUR")
    print()

    # 3 métricas más importantes del mes
    top3 = [
        ("Total Revenue MTD",     "Total Revenue",     "mtd"),
        ("Occupancy % Full Month","Occupancy %",       "forecast"),
        ("GOP Full Month",        "GOP",               "forecast"),
    ]
    print("  MÉTRICAS CLAVE DEL MES:")
    for label, nm, col in top3:
        v = metricas.get(nm, {}).get(col)
        f = _sf(v)
        if nm == "Occupancy %" and f is not None:
            display = f"{f*100:.1f}%" if abs(f) <= 1 else f"{f:.1f}%"
        elif nm == "GOP %" and f is not None:
            display = f"{f*100:.1f}%" if abs(f) <= 1 else f"{f:.1f}%"
        elif f is not None:
            display = f"{f:,.0f} EUR"
        else:
            display = "N/D"
        print(f"    {label:<30} {display}")

    print("=" * 65)


if __name__ == "__main__":
    main()
