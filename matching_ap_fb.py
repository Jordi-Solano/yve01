"""
matching_ap_fb.py — Yve.01 Módulo AP
3-way matching para facturas F&B: Factura + PO + POS.
Estados: MATCH_3WAY_OK | ALERTA_CONSUMO | DISCREPANCIA_PO | SIN_PO | SIN_DATOS_POS
Ejecutar: python matching_ap_fb.py
"""

import os, glob, re
from datetime import date, datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
PROCESADAS_DIR = os.path.join(BASE_DIR, "facturas-procesadas")
REFERENCIA_DIR = os.path.join(BASE_DIR, "datos-referencia")
REPORTES_DIR   = os.path.join(BASE_DIR, "reportes")
os.makedirs(REPORTES_DIR, exist_ok=True)

ORDENES_FILE   = os.path.join(REFERENCIA_DIR, "pos_ordenes.xlsx")
POS_FILE       = os.path.join(REFERENCIA_DIR, "pos_ventas.xlsx")
FECHA_HOY      = date.today().strftime("%Y%m%d")
SALIDA         = os.path.join(REPORTES_DIR, f"matching_fb_{FECHA_HOY}.xlsx")

NF             = "NO_ENCONTRADO"
TOL_PO         = 0.01   # 1% — PO vs factura
TOL_POS        = 0.15   # 15% — POS coste vs factura

VERDE   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ROJO    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
AMARILLO= PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
AZUL    = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

def safe_float(v):
    try:
        if v is None or str(v).strip() in ("", NF, "nan", "None"):
            return None
        s = str(v).replace("EUR","").replace("€","").replace(" ","").strip()
        if "," in s and "." in s:
            s = s.replace(",","") if s.rfind(".") > s.rfind(",") else s.replace(".","").replace(",",".")
        elif "," in s:
            s = s.replace(",",".")
        return float(s)
    except Exception:
        return None

def parse_fecha(s):
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(s).strip(), fmt).date()
        except Exception:
            pass
    return None

def cargar_ultimo_ap_fb():
    excels = sorted(glob.glob(os.path.join(PROCESADAS_DIR, "facturas_ap_*.xlsx")), reverse=True)
    if not excels:
        raise FileNotFoundError("No hay facturas_ap_*.xlsx. Ejecuta lector_facturas_ap.py")
    df = pd.read_excel(excels[0])
    print(f"  Facturas AP: {os.path.basename(excels[0])}")
    return df[df.get("tipo_proveedor","") == "FB"].copy() if "tipo_proveedor" in df.columns else df

def cargar_ordenes_fb():
    if not os.path.exists(ORDENES_FILE):
        raise FileNotFoundError(f"No se encontró {ORDENES_FILE}. Ejecuta gestor_pos.py")
    df = pd.read_excel(ORDENES_FILE)
    df["proveedor_norm"] = df["proveedor"].str.strip().str.lower()
    return df[df["departamento"] == "F&B"].copy() if "departamento" in df.columns else df

def cargar_pos_ventas():
    if not os.path.exists(POS_FILE):
        raise FileNotFoundError(f"No se encontró {POS_FILE}. Ejecuta gestor_pos.py")
    df = pd.read_excel(POS_FILE)
    df["_fecha"] = df["fecha"].apply(parse_fecha)
    return df

def coste_pos_en_periodo(pos_df, fecha_factura):
    """Suma el coste estimado del POS en el mismo mes que la factura."""
    if pos_df.empty or fecha_factura is None:
        return None
    mask = pos_df["_fecha"].apply(
        lambda d: d is not None and d.year == fecha_factura.year and d.month == fecha_factura.month
    )
    sub = pos_df[mask]
    if sub.empty:
        return None
    return float(sub["coste_estimado"].sum())

def buscar_po_fb(nombre_prov, total, ordenes_df):
    if not nombre_prov or nombre_prov == NF:
        return None
    norm = str(nombre_prov).strip().lower()
    mask = ordenes_df["proveedor_norm"].apply(
        lambda p: (norm in p or p in norm) and len(norm) > 3
    )
    candidatos = ordenes_df[mask]
    if candidatos.empty:
        return None
    if total is None:
        return candidatos.iloc[0].to_dict()
    candidatos = candidatos.copy()
    candidatos["_diff"] = (candidatos["importe_aprobado"] - total).abs()
    return candidatos.sort_values("_diff").iloc[0].to_dict()

def analizar_fb(fila, ordenes_df, pos_df):
    nombre_prov = str(fila.get("nombre_proveedor", NF))
    total       = safe_float(fila.get("total_factura"))
    fecha_fac   = parse_fecha(fila.get("fecha"))

    po = buscar_po_fb(nombre_prov, total, ordenes_df)
    coste_pos = coste_pos_en_periodo(pos_df, fecha_fac)

    # ── 1. Verificar PO ───────────────────────────────────────────────────
    if po is None:
        return {
            **fila.to_dict(),
            "numero_po":          NF,
            "importe_po":         NF,
            "departamento_po":    "F&B",
            "coste_pos_periodo":  coste_pos,
            "diferencia_po":      NF,
            "diferencia_pos_pct": NF,
            "estado_matching":    "SIN_PO",
            "alerta_detalle":     "No se encontró PO para este proveedor F&B",
        }

    importe_po = float(po.get("importe_aprobado", 0) or 0)

    if total is not None and importe_po > 0:
        diff_po_abs = total - importe_po
        diff_po_pct = abs(diff_po_abs) / importe_po
    else:
        diff_po_abs = None
        diff_po_pct = None

    po_ok = diff_po_pct is not None and diff_po_pct <= TOL_PO

    if not po_ok:
        return {
            **fila.to_dict(),
            "numero_po":          po.get("numero_po", NF),
            "importe_po":         importe_po,
            "departamento_po":    "F&B",
            "coste_pos_periodo":  coste_pos,
            "diferencia_po":      round(diff_po_abs, 2) if diff_po_abs is not None else NF,
            "diferencia_pos_pct": NF,
            "estado_matching":    "DISCREPANCIA_PO",
            "alerta_detalle":     (f"PO {po['numero_po']}: factura={total:.2f} EUR "
                                   f"vs PO={importe_po:.2f} EUR "
                                   f"(diff={diff_po_abs:+.2f} EUR, {diff_po_pct*100:.1f}%)"),
        }

    # ── 2. PO OK — verificar POS ──────────────────────────────────────────
    if coste_pos is None:
        return {
            **fila.to_dict(),
            "numero_po":          po.get("numero_po", NF),
            "importe_po":         importe_po,
            "departamento_po":    "F&B",
            "coste_pos_periodo":  NF,
            "diferencia_po":      round(diff_po_abs, 2) if diff_po_abs is not None else NF,
            "diferencia_pos_pct": NF,
            "estado_matching":    "SIN_DATOS_POS",
            "alerta_detalle":     "PO cuadra pero no hay datos POS para el periodo de la factura",
        }

    if total is not None and total > 0:
        diff_pos_pct = abs(coste_pos - total) / total
    else:
        diff_pos_pct = 0.0

    if diff_pos_pct <= TOL_POS:
        estado  = "MATCH_3WAY_OK"
        detalle = (f"PO {po['numero_po']} OK ({diff_po_pct*100:.2f}%) "
                   f"| POS coste={coste_pos:.2f} EUR vs factura={total:.2f} EUR "
                   f"({diff_pos_pct*100:.1f}% diferencia)")
    else:
        estado  = "ALERTA_CONSUMO"
        detalle = (f"PO OK — pero POS coste={coste_pos:.2f} EUR vs factura={total:.2f} EUR "
                   f"difiere {diff_pos_pct*100:.1f}% (>15%). "
                   f"Posible merma, inventario incorrecto o error de registro.")

    return {
        **fila.to_dict(),
        "numero_po":          po.get("numero_po", NF),
        "importe_po":         importe_po,
        "departamento_po":    "F&B",
        "coste_pos_periodo":  round(coste_pos, 2),
        "diferencia_po":      round(diff_po_abs, 2) if diff_po_abs is not None else NF,
        "diferencia_pos_pct": f"{diff_pos_pct*100:.1f}%",
        "estado_matching":    estado,
        "alerta_detalle":     detalle,
    }

def aplicar_formato(ws):
    try:
        col_est = next((i+1 for i, c in enumerate(ws[1]) if c.value == "estado_matching"), None)
        if col_est is None: return
        colores = {
            "MATCH_3WAY_OK":   VERDE,
            "DISCREPANCIA_PO": ROJO,
            "SIN_PO":          AMARILLO,
            "ALERTA_CONSUMO":  AZUL,
            "SIN_DATOS_POS":   AMARILLO,
        }
        for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
            fill = colores.get(ws.cell(ri, col_est).value)
            if fill:
                for cell in row: cell.fill = fill
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    except Exception:
        pass

def generar_resumen(df):
    total = len(df)
    cnt = df["estado_matching"].value_counts().reset_index()
    cnt.columns = ["Estado","Cantidad"]
    cnt["Descripcion"] = cnt["Estado"].map({
        "MATCH_3WAY_OK":   "Factura + PO + POS cuadran",
        "ALERTA_CONSUMO":  "PO OK pero POS difiere >15% — revisar inventario",
        "DISCREPANCIA_PO": "Factura no cuadra con PO",
        "SIN_PO":          "No existe orden de compra",
        "SIN_DATOS_POS":   "Sin datos POS para el periodo",
    }).fillna("")
    return pd.concat([cnt, pd.DataFrame([
        {"Estado":"─"*22,"Cantidad":"","Descripcion":""},
        {"Estado":"TOTAL F&B","Cantidad":total,"Descripcion":""},
    ])], ignore_index=True)

def main():
    print("="*60)
    print("  Yve.01 — 3-Way Matching AP · Sub-camino F&B")
    print("="*60)
    try:
        df_fb   = cargar_ultimo_ap_fb()
        df_ord  = cargar_ordenes_fb()
        df_pos  = cargar_pos_ventas()
    except FileNotFoundError as e:
        print(f"\n❌ {e}"); return

    print(f"  Facturas F&B: {len(df_fb)}  |  POs F&B: {len(df_ord)}  |  Líneas POS: {len(df_pos)}\n")
    iconos = {"MATCH_3WAY_OK":"✓","DISCREPANCIA_PO":"✗","SIN_PO":"?",
              "ALERTA_CONSUMO":"⚠","SIN_DATOS_POS":"~"}

    resultados = []
    for _, fila in df_fb.iterrows():
        res = analizar_fb(fila, df_ord, df_pos)
        ic  = iconos.get(res["estado_matching"],"·")
        print(f"  [{ic}] {res['archivo']} → {res['estado_matching']}")
        if res["estado_matching"] != "MATCH_3WAY_OK":
            print(f"       {res['alerta_detalle']}")
        resultados.append(res)

    df_res = pd.DataFrame(resultados)
    df_sum = generar_resumen(df_res)

    with pd.ExcelWriter(SALIDA, engine="openpyxl") as w:
        df_res.to_excel(w, index=False, sheet_name="Detalle_FB")
        df_sum.to_excel(w, index=False, sheet_name="Resumen_FB")
        aplicar_formato(w.sheets["Detalle_FB"])
        for sn in ["Detalle_FB","Resumen_FB"]:
            ws = w.sheets[sn]
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col)+4, 50)

    print("\n" + "─"*60)
    print("  RESUMEN 3-WAY MATCHING F&B")
    print("─"*60)
    for _, r in df_sum.iterrows():
        desc = r.get("Descripcion","")
        print(f"  {str(r['Estado']):<24} {str(r['Cantidad']):<6} {desc}")
    print(f"\n✅ Reporte: {SALIDA}")
    print("="*60)

if __name__ == "__main__":
    main()
