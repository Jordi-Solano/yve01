"""
asignador_cuentas.py — Yve.01 Módulo AP
Asigna cuentas contables a cada factura y genera el asiento contable.
Ejecutar: python asignador_cuentas.py
"""

import os, glob, re
from datetime import date
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

BASE_DIR       = os.path.dirname(os.path.abspath(__file__))

# Multi-tenant: al lanzarse por subprocess desde el dashboard, tenant_dirs lee
# YVE_TENANT del entorno (lo pone _env_tenant()). Mismo patron que
# verificador_comisiones y lector_drr. Sin esto, las facturas de un cliente se
# leerian y el informe contable se escribiria en el arbol raiz, mezclando los
# datos financieros de dos hoteles. Para el tenant 'default' tenant_dirs
# devuelve BASE_DIR, asi que no cambia nada de lo que ya funciona.
try:
    from tenant_dirs import reportes_dir as _t_rep, procesadas_dir as _t_proc, datos_dir as _t_datos
    REPORTES_DIR   = _t_rep()
    PROCESADAS_DIR = _t_proc()
    REFERENCIA_DIR = _t_datos()
except Exception:
    REPORTES_DIR   = os.path.join(BASE_DIR, "reportes")
    PROCESADAS_DIR = os.path.join(BASE_DIR, "facturas-procesadas")
    REFERENCIA_DIR = os.path.join(BASE_DIR, "datos-referencia")
os.makedirs(REPORTES_DIR, exist_ok=True)

PROV_FILE  = os.path.join(REFERENCIA_DIR, "proveedores.xlsx")
CC_FILE    = os.path.join(REFERENCIA_DIR, "plan_cuentas.xlsx")
FECHA_HOY  = date.today().strftime("%Y%m%d")
SALIDA     = os.path.join(PROCESADAS_DIR, f"facturas_contabilizadas_{FECHA_HOY}.xlsx")
NF         = "NO_ENCONTRADO"

VERDE   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMARILLO= PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GRIS    = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

# ── Reglas de asignación ──────────────────────────────────────────────────
# Orden de evaluación: más específico primero

REGLAS_CUENTA_GASTO = [
    # F&B → 600
    ({"tipo_proveedor": "FB"},                                    "600"),
    # Por nombre de proveedor (palabras clave)
    ({"kw_proveedor": ["limpieza","cleaning","housekeeping"]},    "623"),
    ({"kw_proveedor": ["telefon","vodafone","orange","telecom",
                       "movil","internet","fibra"]},              "629"),
    ({"kw_proveedor": ["endesa","iberdrola","gas","electr",
                       "energia","suministro"]},                  "629"),
    ({"kw_proveedor": ["otis","schindler","kone","ascensor",
                       "elevador","mantenimiento","reparacion"]}, "622"),
    ({"kw_proveedor": ["securitas","prosegur","seguridad",
                       "alarma","vigilancia"]},                   "623"),
    ({"kw_proveedor": ["seguro","mapfre","axa","allianz",
                       "zurich","prima"]},                        "625"),
    ({"kw_proveedor": ["arrendamiento","alquiler","renting"]},    "621"),
    # Fallback
    ({"tipo_proveedor": "OTRAS"},                                 "629"),
]

def safe_float(v):
    try:
        if v is None or str(v).strip() in ("", NF, "nan", "None"):
            return None
        s = str(v).replace("EUR","").replace("€","").replace(" ","").strip()
        if "," in s and "." in s:
            s = s.replace(",","") if s.rfind(".") > s.rfind(",") else s.replace(".","").replace(",",".")
        elif "," in s:
            s = s.replace(",",".")
        return float(s)
    except Exception:
        return None

def cargar_proveedores():
    if not os.path.exists(PROV_FILE):
        return {}
    df = pd.read_excel(PROV_FILE)
    return {r["nombre_proveedor"].strip().lower(): r.to_dict() for _, r in df.iterrows()}

def cargar_plan_cuentas():
    if not os.path.exists(CC_FILE):
        return {}
    df = pd.read_excel(CC_FILE)
    return {str(r["codigo_cuenta"]).strip(): r.to_dict() for _, r in df.iterrows()}

def determinar_cuenta_gasto(fila, proveedores):
    tipo      = str(fila.get("tipo_proveedor","")).strip().upper()
    nombre    = str(fila.get("nombre_proveedor","")).strip().lower()
    ya_tiene  = str(fila.get("cuenta_contable","")).strip()

    # Si ya viene asignada desde proveedores.xlsx y no es NF
    if ya_tiene and ya_tiene not in (NF, "nan", "None", ""):
        return ya_tiene, "ASIGNADA_PROVEEDOR"

    for regla, cuenta in REGLAS_CUENTA_GASTO:
        if "tipo_proveedor" in regla and tipo == regla["tipo_proveedor"]:
            return cuenta, "REGLA_TIPO"
        if "kw_proveedor" in regla:
            if any(kw in nombre for kw in regla["kw_proveedor"]):
                return cuenta, "REGLA_KEYWORD"

    return "REVISAR_MANUAL", "SIN_REGLA"

def cuenta_iva(pct_iva):
    try:
        pct = float(str(pct_iva).replace("%","").strip())
        if abs(pct - 21) < 0.5:  return "472",  "H.P. IVA soportado 21%"
        if abs(pct - 10) < 0.5:  return "4720", "H.P. IVA soportado 10%"
        if abs(pct -  4) < 0.5:  return "4721", "H.P. IVA soportado 4%"
    except Exception:
        pass
    return "472", "H.P. IVA soportado"

def generar_asiento(fila, cuenta_gasto, plan_cc):
    num_fac  = fila.get("numero_factura", NF)
    prov     = fila.get("nombre_proveedor", NF)
    base     = safe_float(fila.get("base_imponible"))
    cuota    = safe_float(fila.get("cuota_iva"))
    total    = safe_float(fila.get("total_factura"))
    pct_iva  = fila.get("porcentaje_iva", NF)
    fecha    = fila.get("fecha", NF)

    # Descripción de la cuenta de gasto
    desc_gasto = plan_cc.get(cuenta_gasto, {}).get("descripcion", cuenta_gasto)

    # Cuenta IVA
    c_iva, desc_iva = cuenta_iva(pct_iva)

    # Asiento textual
    lineas = [
        f"FECHA: {fecha}  |  CONCEPTO: Fact. {num_fac} – {prov}",
        f"  DEBE  {cuenta_gasto} {desc_gasto:<40} {base or '?':>12} EUR",
        f"  DEBE  {c_iva:<6} {desc_iva:<40} {cuota or '?':>12} EUR",
        f"  HABER 400    Proveedores ({prov}){'':<20} {total or '?':>12} EUR",
    ]

    return {
        "cuenta_debe_gasto":  cuenta_gasto,
        "cuenta_debe_iva":    c_iva,
        "cuenta_haber":       "400",
        "asiento_contable":   " | ".join(lineas),
    }

# Columnas que escribe Oracle y que NUNCA se pueden perder al regenerar el
# informe: oracle_status es el unico candado que impide contabilizar dos veces.
_COLS_ORACLE = ("oracle_status", "oracle_journal_id", "oracle_batch_id",
                "oracle_fecha", "oracle_error")


_VACIOS_CRUCE = ("", "nan", "none", "nat", "<na>", "no_encontrado", "null")


def _txt_cruce(v):
    """Texto comparable para la clave del cruce. Los vacios, todos a ''.

    Hace falta porque los dos lados escriben el vacio DISTINTO:
    `matching_ap_albaran` pone la cadena literal `NO_ENCONTRADO`, el Excel de
    facturas devuelve `NaN`, y `str(float('nan'))` es `'nan'`, que no es vacio
    para Python. Sin unificarlo, la clave no casaria y la factura perderia su
    estado — que es la unica forma en que este cambio podria empeorar algo.

    NO se pasa a minusculas a proposito: `_guardar_factura_ap` deduplica con la
    cadena tal cual, y plegar mayusculas juntaria `Factura.pdf` con
    `factura.pdf`, que en un sistema sensible a mayusculas son dos ficheros.
    """
    s = "" if v is None else str(v)
    s = " ".join(s.split()).strip()
    return "" if s.lower() in _VACIOS_CRUCE else s


def _clave_cruce(fila):
    """Identidad de una factura para pegarle el estado de un cruce.

    Es `(archivo, hotel_id)`: **la MISMA pareja con la que
    `_guardar_factura_ap` deduplica** el Excel de facturas. Esa es toda la
    razon — quien lee tiene que usar la identidad de quien escribe, o los dos
    criterios se separan en el siguiente cambio.

    POR QUE NO ERA ASI, Y POR QUE FALLABA:
    cuando se escribio este merge, `archivo` SI era unico por factura, y el
    codigo lo dice en el docstring de `clave_factura` de la pantalla de
    aprobaciones. La separacion por hotel cambio la clave del escritor a
    `(archivo, hotel_id)` —para que dos hoteles que suban `factura_enero.pdf` no
    se borren el uno al otro— y este lado se quedo atras. Consecuencia medida:
    con dos hoteles y el mismo nombre de fichero sobrevivia UN estado y se le
    pegaba a las dos facturas, asi que **una DIFERENCIA_IMPORTE real llegaba al
    panel en verde**, lista para aprobar y contabilizar.

    POR QUE NO EL NUMERO DE FACTURA (que es lo primero que uno piensa):
      - puede venir VACIO, y el proyecto no fusiona nunca dos facturas sin
        numero; con el numero como clave volverian a compartirla todas;
      - no es unico por si solo: cada proveedor tiene su propia numeracion, asi
        que la `FA-001` de dos proveedores son dos facturas distintas;
      - y no es la clave del escritor, que es el criterio que importa aqui.
    """
    return _txt_cruce(fila.get("archivo")) + "|" + _txt_cruce(fila.get("hotel_id"))


def cargar_todas_facturas_ap():
    """Carga TODAS las facturas AP (todos los dias) conservando lo ya contabilizado.

    ANTES leia SOLO el facturas_ap_*.xlsx mas reciente y regeneraba el informe
    desde cero. Dos consecuencias, las dos malas:

      - al cambiar de dia, el informe PERDIA las facturas de los dias anteriores;
      - y perdia la columna `oracle_status`, que es el unico candado que impide
        que Oracle contabilice dos veces la misma factura.

    Y como `oracle_lector_facturas.py` abre SOLO el informe mas reciente, una
    factura ya contabilizada que reapareciera sin su marcador se contabilizaria
    OTRA VEZ en el libro mayor. Reproducido antes de arreglarlo.

    `almacen_datos.facturas_ap()` junta todos los dias y hace ganar la etapa mas
    avanzada (facturas_contabilizadas por encima de facturas_ap), asi que el
    marcador viaja solo. Es el mismo punto unico que ya usan el panel, el banco
    y la conciliacion.
    """
    import almacen_datos as _alm
    df = _alm.facturas_ap(PROCESADAS_DIR, REPORTES_DIR)
    if df is None or df.empty:
        raise FileNotFoundError("No hay facturas_ap_*.xlsx. Ejecuta lector_facturas_ap.py")
    ya = 0
    if "oracle_status" in df.columns:
        ya = int((df["oracle_status"].astype(str).str.upper() == "CONTABILIZADA").sum())
    print(f"  Facturas AP: {len(df)} de todos los dias"
          + (f" · {ya} ya contabilizada(s) en Oracle (marcador conservado)" if ya else ""))

    # ── Unir el estado de los cruces ──────────────────────────────────────
    # ANTES se hacia un merge POR INFORME dentro del bucle, y cada uno quitaba
    # la columna del anterior: con dos cruces activos, el ultimo pisaba al
    # primero y se perdia. Ahora se juntan todos en uno y se mezcla UNA vez.
    #
    # El orden importa: gana el PRIMERO que tenga estado para esa factura. El
    # cruce contra PO va delante del de albaranes porque los estados del panel
    # (SIN_PO, DISCREPANCIA_PO) se diseñaron para el. Meter dos cruces
    # independientes —¿estaba pedido? / ¿llego la mercancia?— en una sola
    # columna es una simplificacion; se revisa en la fase del PO.
    _cruces = []
    for patron in [f"matching_otras_{FECHA_HOY}.xlsx",
                   f"matching_fb_{FECHA_HOY}.xlsx",
                   f"matching_albaran_{FECHA_HOY}.xlsx"]:
        ruta = os.path.join(REPORTES_DIR, patron)
        if not os.path.exists(ruta):
            continue
        try:
            dm = pd.read_excel(ruta, sheet_name=0)
        except Exception:
            continue
        if "archivo" not in dm.columns or "estado_matching" not in dm.columns:
            continue
        # cada modulo llama al detalle de una manera
        _det = next((c for c in ("detalle_matching", "alerta_detalle", "detalle")
                     if c in dm.columns), None)
        _cols = ["archivo", "estado_matching"] + ([_det] if _det else [])
        if "hotel_id" in dm.columns:
            _cols.append("hotel_id")
        dm = dm[_cols].copy()
        if "hotel_id" not in dm.columns:
            # Un informe sin columna de hotel no es un error: puede venir de
            # antes de la separacion. Se trata como "sin asignar", y la red de
            # mas abajo se encarga de que no pierda el cruce por eso.
            dm["hotel_id"] = ""
        if _det:
            dm = dm.rename(columns={_det: "detalle_matching"})
        else:
            dm["detalle_matching"] = ""
        # una fila sin 'archivo' no puede unirse a nada, y ademas pandas 3
        # trata NaN==NaN al deduplicar: se irian todas menos una
        dm = dm[dm["archivo"].map(lambda v: str(v).strip().lower()
                                  not in ("", "nan", "none", "no_encontrado"))]
        if not dm.empty:
            _cruces.append(dm)
            print(f"  Matching unido: {patron}")
    if _cruces:
        dm_total = pd.concat(_cruces, ignore_index=True)
        # La clave es (archivo, hotel) — ver `_clave_cruce`. `keep="first"`
        # CONSERVA la prioridad de siempre: gana el primer informe de la lista
        # (PO > F&B > albaran), que esta puesta a proposito ahi arriba.
        dm_total["_clave_cruce"] = [_clave_cruce(f) for f in dm_total.to_dict("records")]
        dm_total = dm_total.drop_duplicates(subset=["_clave_cruce"], keep="first")
        _mt = dm_total[["_clave_cruce", "estado_matching", "detalle_matching"]]
        # quitar las de la pasada anterior: si no, el merge crea
        # estado_matching_x / estado_matching_y y se pierde la buena
        df = df.drop(columns=[c for c in ("estado_matching", "detalle_matching")
                              if c in df.columns])
        df["_clave_cruce"] = [_clave_cruce(f) for f in df.to_dict("records")]
        df = df.merge(_mt, on="_clave_cruce", how="left")

        # ── LA RED: nadie puede PERDER el estado que ya tenia ─────────────
        # Segunda pasada por `archivo` A SECAS para las facturas que se hayan
        # quedado sin estado, pero SOLO con los archivos que aparecen bajo UNA
        # sola clave. Si un archivo esta en dos hoteles es exactamente el caso
        # del bug, y ahi la red NO se aplica.
        #
        # Con esto la propiedad que pidio el usuario se cumple por
        # construccion: una factura solo puede CONSERVAR o GANAR estado, nunca
        # perderlo. Sirve para el informe que no trae columna de hotel mientras
        # la factura si lo tiene — el unico caso en que la clave fuerte falla.
        _por_arch = {}
        for fila in dm_total.to_dict("records"):
            _a = _txt_cruce(fila.get("archivo"))
            if _a:
                _por_arch.setdefault(_a, []).append(fila)
        _unicos = {a: v[0] for a, v in _por_arch.items() if len(v) == 1}
        if _unicos:
            # astype(object) antes de escribir texto: una columna que ha salido
            # entera vacia se tipa float64 y rechaza cadenas (nos paso en la
            # asignacion manual de conciliacion)
            for _c in ("estado_matching", "detalle_matching"):
                if _c in df.columns:
                    df[_c] = df[_c].astype(object)
            _rescatadas = 0
            for _i in df.index:
                if _txt_cruce(df.at[_i, "estado_matching"] if "estado_matching" in df.columns else ""):
                    continue
                _e = _unicos.get(_txt_cruce(df.at[_i, "archivo"]))
                if _e:
                    df.at[_i, "estado_matching"] = _e.get("estado_matching")
                    df.at[_i, "detalle_matching"] = _e.get("detalle_matching", "")
                    _rescatadas += 1
            if _rescatadas:
                print(f"  Cruzadas por nombre de fichero (el informe no traia hotel): {_rescatadas}")

        df = df.drop(columns=["_clave_cruce"])
    return df

def aplicar_formato(ws):
    try:
        col_asig = next((i+1 for i, c in enumerate(ws[1]) if c.value == "estado_asignacion"), None)
        if col_asig is None: return
        for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
            v = ws.cell(ri, col_asig).value
            fill = VERDE if v == "ASIGNADA_PROVEEDOR" else (AMARILLO if v == "REGLA_KEYWORD" else (GRIS if v == "SIN_REGLA" else None))
            if fill:
                for cell in row: cell.fill = fill
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    except Exception:
        pass

def main():
    print("="*60)
    print("  Yve.01 — Asignador de Cuentas Contables AP")
    print("="*60)
    try:
        df = cargar_todas_facturas_ap()
    except FileNotFoundError as e:
        print(f"\n❌ {e}"); return

    proveedores = cargar_proveedores()
    plan_cc     = cargar_plan_cuentas()
    print(f"  Plan de cuentas: {len(plan_cc)} cuentas\n")

    resultados = []
    manuales   = 0
    for _, fila in df.iterrows():
        cuenta_gasto, metodo = determinar_cuenta_gasto(fila, proveedores)
        asiento_dict = generar_asiento(fila, cuenta_gasto, plan_cc)

        if cuenta_gasto == "REVISAR_MANUAL":
            manuales += 1
            icono = "⚠"
        else:
            icono = "✓"
        print(f"  [{icono}] {fila.get('archivo',NF)} → {cuenta_gasto} [{metodo}]")

        # fila.to_dict() ya arrastra las columnas de Oracle si venian del
        # informe anterior; el orden importa: asiento_dict NO las pisa.
        resultados.append({
            **fila.to_dict(),
            **asiento_dict,
            "estado_asignacion": metodo,
        })

    df_res = pd.DataFrame(resultados)

    with pd.ExcelWriter(SALIDA, engine="openpyxl") as w:
        df_res.to_excel(w, index=False, sheet_name="Facturas_Contabilizadas")
        aplicar_formato(w.sheets["Facturas_Contabilizadas"])
        ws = w.sheets["Facturas_Contabilizadas"]
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col)+4, 50)

    print(f"\n  Asignadas automáticamente: {len(resultados)-manuales}")
    print(f"  Requieren revisión manual:  {manuales}")
    print(f"\n✅ Archivo: {SALIDA}")
    print("="*60)

if __name__ == "__main__":
    main()
