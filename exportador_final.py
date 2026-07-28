"""
Exportador profesional de reportes PDF y Excel
Para enviar a dirección y stakeholders
"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
import os

def _dir_reportes():
    """La carpeta de reportes del tenant que esta pidiendo el informe.

    Antes esto era la cadena 'reportes' a secas, con dos fallos en una linea:
    era relativa al directorio desde el que se arranco el proceso, y era
    ciega al tenant. O sea que el consolidado de un cliente caia en la
    carpeta raiz — la misma que ve el tenant `default` listada en
    /api/debug y contada en /api/health.

    Se resuelve en CADA llamada y no en __init__ a proposito: el tenant sale
    de la sesion, y la instancia de esta clase es unica y global, creada al
    importar el modulo, cuando todavia no hay ninguna peticion.
    """
    try:
        from tenant_dirs import reportes_dir
        d = reportes_dir()
    except Exception:
        d = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reportes")
    os.makedirs(d, exist_ok=True)
    return d


class ExportadorReportes:
    def __init__(self):
        self.styles = getSampleStyleSheet()

    def pdf_ejecutivo(self, data):
        """Reporte ejecutivo en PDF para dirección"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        path = os.path.join(_dir_reportes(), f'Reporte_Ejecutivo_{timestamp}.pdf')
        
        story = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle', parent=self.styles['Heading1'],
            fontSize=24, textColor=colors.HexColor('#1F4E78'),
            spaceAfter=30, alignment=1
        )
        story.append(Paragraph("REPORTE EJECUTIVO YVE", title_style))
        story.append(Paragraph(f"<font size=10 color=#666666>{datetime.now().strftime('%d de %B de %Y')}</font>", self.styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # KPIs
        story.append(Paragraph("Métricas Clave", self.styles['Heading2']))
        story.append(Spacer(1, 0.2*inch))
        
        kpis_data = [
            ["Métrica", "Valor", "Status"],
            ["Revenue MTD", "€1,918,400", "✓ OK"],
            ["Ocupación Promedio", "87.5%", "✓ OK"],
            ["Facturas Procesadas", "342", "✓ OK"],
            ["Discrepancias", "1", "⚠ Revisar"],
        ]
        
        table = Table(kpis_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        story.append(table)
        
        doc = SimpleDocTemplate(path, pagesize=letter)
        doc.build(story)
        return path
    
    def excel_consolidado(self, data):
        """Excel consolidado con múltiples hojas"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        path = os.path.join(_dir_reportes(), f'Consolidado_{timestamp}.xlsx')
        
        wb = Workbook()
        
        # Hoja 1: KPIs
        ws = wb.active
        ws.title = "KPIs"
        
        headers = ['Métrica', 'Valor', 'Variación', 'Status']
        ws.append(headers)
        
        for header in headers:
            ws[f'{chr(65+headers.index(header))}1'].font = Font(bold=True, color="FFFFFF")
            ws[f'{chr(65+headers.index(header))}1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        data_rows = [
            ['Revenue MTD', '€1,918,400', '+2.1%', 'OK'],
            ['Ocupación', '87.5%', '-0.5%', 'OK'],
            ['Facturas', '342', '+15', 'OK'],
            ['Discrepancias', '1', '-2', 'Warning'],
        ]
        
        for row in data_rows:
            ws.append(row)
        
        # Hoja 2: Hoteles
        ws2 = wb.create_sheet("Hoteles")
        ws2.append(['Hotel', 'Rooms', 'Ocupación', 'ADR', 'Revenue', 'GOP%'])
        
        hotel_data = [
            ['Barcelona', 250, '89.5%', '€285', '€1,875K', '22.0%'],
            ['Valencia', 180, '92.3%', '€195', '€972K', '18.0%'],
            ['Sevilla', 95, '87.2%', '€165', '€410K', '20.0%'],
        ]
        
        for row in hotel_data:
            ws2.append(row)
        
        # Hoja 3: Alertas
        ws3 = wb.create_sheet("Alertas")
        ws3.append(['Fecha', 'Tipo', 'Mensaje', 'Hotel', 'Acción'])
        
        alerts = [
            ['2026-06-06', 'DI Missing', '3 facturas sin DI cert', 'BCN', 'Contactar OTA'],
            ['2026-06-06', 'AP Disc', 'PO mismatch Pescados', 'BCN', 'Revisar albarán'],
            ['2026-06-05', 'DRR OOB', 'Out of Balance -€245', 'VAL', 'Auditoría'],
        ]
        
        for row in alerts:
            ws3.append(row)
        
        wb.save(path)
        return path

# Instancia global
exportador = ExportadorReportes()

def generar_reporte_ejecutivo():
    return exportador.pdf_ejecutivo({})

def generar_excel_consolidado():
    return exportador.excel_consolidado({})

if __name__ == "__main__":
    print("Generando reportes...")
    pdf = generar_reporte_ejecutivo()
    excel = generar_excel_consolidado()
    print(f"✓ PDF: {pdf}")
    print(f"✓ Excel: {excel}")
