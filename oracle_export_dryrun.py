"""
oracle_export_dryrun.py — Yve.01
Exporta a formato Oracle GL (GL_INTERFACE) los asientos que el pipeline de
Oracle ha PRODUCIDO DE VERDAD, para importarlos a mano cuando no hay conexion.

Fuente unica: reportes/oracle_asientos_producidos.json, que escribe
oracle_pipeline en cada ejecucion (simulacion o produccion). Si el pipeline no
ha corrido, aqui no hay nada que exportar y se dice.

Decision de Jordi (3 sep 2026): este blueprint solo se registra porque exporta
lo producido; NUNCA asientos inventados (la version anterior se inventaba
asientos desde proveedores.xlsx cuando no habia informes — por eso estuvo sin
registrar).
"""
import os
from datetime import date
from io import BytesIO

import pandas as pd
from flask import Blueprint, jsonify, request, send_file

oracle_export_bp = Blueprint('oracle_export', __name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Oracle GL column mapping (Fusion GL_INTERFACE)
GL_COLUMNS = [
    'STATUS', 'LEDGER_NAME', 'ACCOUNTING_DATE', 'CURRENCY_CODE',
    'USER_JE_CATEGORY_NAME', 'USER_JE_SOURCE_NAME', 'REFERENCE_DATE',
    'ACTUAL_FLAG', 'SEGMENT1', 'SEGMENT2', 'SEGMENT3', 'SEGMENT4',
    'ENTERED_DR', 'ENTERED_CR', 'ACCOUNTED_DR', 'ACCOUNTED_CR',
    'DESCRIPTION', 'ATTRIBUTE1', 'ATTRIBUTE2', 'CONVERSION_TYPE',
    'CONVERSION_RATE', 'PERIOD_NAME'
]
ESTADOS_EXPORTABLES = ('CONTABILIZADA', 'CONTABILIZADA_SIM')


def _ledger():
    try:
        from oracle_auth import ORACLE_LEDGER_NAME
        return ORACLE_LEDGER_NAME
    except Exception:
        return 'Yve Ledger'


def _period_name(dt):
    """Fecha -> periodo Oracle: MAY-26."""
    months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
    d = pd.Timestamp(dt)
    return f"{months[d.month - 1]}-{str(d.year)[2:]}"


def _fecha(v):
    s = str(v or '').strip()
    try:
        d = pd.to_datetime(s, dayfirst=not (len(s) >= 10 and s[4] == '-'), errors='coerce')
        if pd.isna(d):
            return date.today().isoformat()
        return d.date().isoformat()
    except Exception:
        return date.today().isoformat()


def _make_gl_row(ledger, date_str, cat, seg1, seg2, seg3, seg4, dr, cr, desc, ref1='', ref2=''):
    return {
        'STATUS': 'NEW',
        'LEDGER_NAME': ledger,
        'ACCOUNTING_DATE': date_str,
        'CURRENCY_CODE': 'EUR',
        'USER_JE_CATEGORY_NAME': cat,
        'USER_JE_SOURCE_NAME': 'YVE01',
        'REFERENCE_DATE': date_str,
        'ACTUAL_FLAG': 'A',
        'SEGMENT1': seg1, 'SEGMENT2': seg2, 'SEGMENT3': seg3, 'SEGMENT4': seg4,
        'ENTERED_DR': round(dr, 2) if dr else '',
        'ENTERED_CR': round(cr, 2) if cr else '',
        'ACCOUNTED_DR': round(dr, 2) if dr else '',
        'ACCOUNTED_CR': round(cr, 2) if cr else '',
        'DESCRIPTION': str(desc)[:240],
        'ATTRIBUTE1': ref1, 'ATTRIBUTE2': ref2,
        'CONVERSION_TYPE': 'User', 'CONVERSION_RATE': 1.0,
        'PERIOD_NAME': _period_name(date_str),
    }


def asientos_exportables(solo_estados=ESTADOS_EXPORTABLES):
    """Los asientos producidos, uno por factura (la ultima ejecucion manda)."""
    from oracle_actualizar_estado import asientos_producidos
    ultimo = {}
    for a in asientos_producidos():
        num = str(a.get('numero_factura') or '').strip()
        if not num:
            continue
        ultimo[num] = a                      # el mas reciente pisa al anterior
    out = [a for a in ultimo.values() if str(a.get('estado', '')).upper() in solo_estados]
    out.sort(key=lambda a: (a.get('timestamp', ''), a.get('numero_factura', '')))
    return out


def gl_rows(asientos):
    """Lineas GL_INTERFACE a partir de los asientos producidos. Sin inventar nada."""
    ledger = _ledger()
    rows = []
    for a in asientos:
        num = str(a.get('numero_factura', ''))
        fecha = _fecha(a.get('fecha'))
        for l in a.get('journal_lines') or []:
            comb = str(l.get('combination') or '')
            partes = comb.split('.') if comb else []
            seg1 = str(l.get('entity') or (partes[0] if len(partes) > 0 else ''))
            seg2 = str(l.get('department') or (partes[1] if len(partes) > 1 else ''))
            seg3 = str(l.get('account') or (partes[2] if len(partes) > 2 else ''))
            dr = float(l.get('debit') or 0)
            cr = float(l.get('credit') or 0)
            if not dr and not cr:
                continue
            rows.append(_make_gl_row(ledger, _fecha(l.get('accounting_date') or fecha),
                                     'Purchase Invoices', seg1, seg2, seg3, '0000',
                                     dr, cr, l.get('description') or '', num,
                                     str(a.get('oracle_id') or '')))
    return rows


@oracle_export_bp.route('/api/oracle/dryrun')
def api_oracle_dryrun():
    """Vista previa: que se exportaria (JSON)."""
    asientos = asientos_exportables()
    rows = gl_rows(asientos)
    total_dr = sum(float(r.get('ENTERED_DR') or 0) for r in rows)
    total_cr = sum(float(r.get('ENTERED_CR') or 0) for r in rows)
    return jsonify({
        'ok': True,
        'mode': 'producido_por_el_pipeline',
        'facturas': len(asientos),
        'entries': len(rows),
        'total_debe': round(total_dr, 2),
        'total_haber': round(total_cr, 2),
        'balanced': abs(total_dr - total_cr) < 0.01,
        'sample': rows[:5],
        'nota': ('Solo asientos que el pipeline de Oracle ha producido (simulacion o real). '
                 'Sin ejecucion del pipeline no hay nada que exportar.'),
    })


@oracle_export_bp.route('/api/oracle/export_excel')
def api_oracle_export_excel():
    """Excel GL_INTERFACE con lo producido por el pipeline, para importar a mano."""
    asientos = asientos_exportables()
    rows = gl_rows(asientos)
    if not rows:
        return jsonify({'ok': False,
                        'error': 'Sin asientos que exportar: el pipeline de Oracle no ha producido ninguno todavia.'}), 404

    df = pd.DataFrame(rows, columns=GL_COLUMNS)
    origen = pd.DataFrame([{
        'numero_factura': a.get('numero_factura'), 'proveedor': a.get('nombre_proveedor'),
        'fecha': a.get('fecha'), 'total_factura': a.get('total_factura'),
        'estado': a.get('estado'), 'modo': a.get('modo'), 'oracle_id': a.get('oracle_id'),
        'batch': a.get('batch_name'), 'producido': a.get('timestamp'),
    } for a in asientos])

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='GL_INTERFACE', index=False)
        origen.to_excel(writer, sheet_name='Origen', index=False)
        ws = writer.sheets['GL_INTERFACE']
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True, size=9)
            cell.alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A2'
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16

    buf.seek(0)
    fname = f"oracle_gl_journal_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)
