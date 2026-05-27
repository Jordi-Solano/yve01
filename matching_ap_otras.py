"""
matching_ap_otras.py — Yve.01 Módulo AP
Cruza facturas OTRAS (no F&B) con sus órdenes de compra (PO).
Estados: MATCH_CORRECTO | DISCREPANCIA | SIN_PO
Ejecutar: python matching_ap_otras.py
"""

import os, glob, re
from datetime import date
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
PROCESADAS_DIR = os.path.join(BASE_DIR, "facturas-procesadas")
REFERENCIA_DIR = os.path.join(BASE_DIR, "datos-referencia")
REPORTES_DIR   = os.path.join(BASE_DIR, "reportes")
os.makedirs(REPORTES_DIR, exist_ok=True)

ORDENES_FILE   = os.path.join(REFERENCIA_DIR, "pos_ordenes.xlsx")
FECHA_HOY      = date.today().strftime("%Y%m%d")
SALIDA         = os.path.join(REPORTES_DIR, f"matching_otras_{FECHA_HOY}.xlsx")

NF             = "NO_ENCONTRADO"
TOLERANCIA_PO  = 0.01   # 1 %

VERDE   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ROJO    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
AMARILLO= PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

def cargar_ultimo_ap():
    excels = sorted(glob.glob(os.path.join(PROCESADAS_DIR, "facturas_ap_*.xlsx")), reverse=True)
    if not excels:
        raise FileNotFoundError(f"No hay facturas_ap_*.xlsx en {PROCESADAS_DIR}.\n"
                                "Ejecuta primero lector_facturas_ap.py")
    ruta = excels[0]
    print(f"  Facturas AP: {os.path.basename(ruta)}")
    df = pd.read_excel(ruta)
    return df[df["tipo_proveedor"] == "OTRAS"].copy() if "tipo_proveedor" in df.columns else df

def cargar_ordenes():
    if not os.path.exists(ORDENES_FILE):
        raise FileNotFoundError(f"No se encontró: {ORDENES_FILE}\nEjecuta gestor_pos.py primero.")
    df = pd.read_excel(ORDENES_FILE)
    df["proveedor_norm"] = df["proveedor"].str.strip().str.lower()
    return df

def safe_float(v):
    try:
        if v is None or str(v).strip() in ("", NF, "nan", "None"):
            return None
        s = str(v).replace("EUR","").replace("€","").replace(" ","").strip()
        if "," in s and "." in s:
            s = s.replace(",","") if s.rfind(".") > s.rfind(",") else s.replace(".","").replace(",",".")
        elif "," in s:
            s = s.replace(",",".")
        return float(s)
    except Exception:
        return None

def buscar_po(nombre_prov, total_factura, ordenes_df):
    """Busca la PO más cercana por proveedor e importe."""
    if not nombre_prov or nombre_prov == NF:
        return None, "SIN_NOMBRE_PROVEEDOR"
    norm = str(nombre_prov).strip().lower()
    # Filtrar por proveedor (coincidencia parcial)
    mask = ordenes_df["proveedor_norm"].apply(
        lambda p: (norm in p or p in norm) and len(norm) > 3
    )
    candidatos = ordenes_df[mask]
    if candidatos.empty:
        return None, "PROVEEDOR_NO_EN_POs"
    if total_factura is None:
        return candidatos.iloc[0].to_dict(), "SIN_IMPORTE_FACTURA"
    # Ordenar por diferencia de importe
    candidatos = candidatos.copy()
    candidatos["_diff"] = (candidatos["importe_aprobado"] - total_factura).abs()
    mejor = candidatos.sort_values("_diff").iloc[0]
    return mejor.to_dict(), None

def analizar_factura(fila, ordenes_df):
    nombre_prov = str(fila.get("nombre_proveedor", NF))
    total       = safe_float(fila.get("total_factura"))
    po, err     = buscar_po(nombre_prov, total, ordenes_df)

    if po is None:
        return {
            **fila.to_dict(),
            "numero_po":          NF,
            "importe_po":         NF,
            "departamento_po":    NF,
            "diferencia_importe": NF,
            "diferencia_pct":     NF,
            "estado_matching":    "SIN_PO",
            "detalle":            err or "No se encontró PO para este proveedor",
        }

    importe_po = float(po.get("importe_aprobado", 0) or 0)
    if total is not None and importe_po > 0:
        diff_abs = total - importe_po
        diff_pct = abs(diff_abs) / importe_po
        if diff_pct <= TOLERANCIA_PO:
            estado   = "MATCH_CORRECTO"
            detalle  = f"PO {po['numero_po']} cuadra ({diff_pct*100:.2f}% diferencia)"
        else:
            estado   = "DISCREPANCIA"
            detalle  = (f"PO {po['numero_po']}: factura={total:.2f} EUR, "
                        f"PO={importe_po:.2f} EUR, diff={diff_abs:+.2f} EUR")
    else:
        diff_abs = NF
        diff_pct = NF
        estado   = "SIN_IMPORTE"
        detalle  = "Importe de factura no extraído"

    return {
        **fila.to_dict(),
        "numero_po":          po.get("numero_po", NF),
        "importe_po":         importe_po,
        "departamento_po":    po.get("departamento", NF),
        "diferencia_importe": round(diff_abs, 2) if isinstance(diff_abs, float) else NF,
        "diferencia_pct":     f"{diff_pct*100:.2f}%" if isinstance(diff_pct, float) else NF,
        "estado_matching":    estado,
        "detalle":            detalle,
    }

def aplicar_formato(ws):
    try:
        col_est = next((i+1 for i, c in enumerate(ws[1]) if c.value == "estado_matching"), None)
        if col_est is None:
            return
        colores = {"MATCH_CORRECTO":VERDE, "DISCREPANCIA":ROJO, "SIN_PO":AMARILLO,
                   "SIN_IMPORTE":AMARILLO, "SIN_NOMBRE_PROVEEDOR":AMARILLO}
        for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
            fill = colores.get(ws.cell(ri, col_est).value)
            if fill:
                for cell in row: cell.fill = fill
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    except Exception:
        pass

def generar_resumen(df):
    total = len(df)
    cnt = df["estado_matching"].value_counts().reset_index()
    cnt.columns = ["Estado","Cantidad"]
    cnt["Pct"] = (cnt["Cantidad"]/total*100).round(1).astype(str)+"%"
    extra = pd.DataFrame([
        {"Estado":"─"*20,"Cantidad":"","Pct":""},
        {"Estado":"TOTAL FACTURAS OTRAS","Cantidad":total,"Pct":"100%"},
    ])
    return pd.concat([cnt, extra], ignore_index=True)

def main():
    print("="*60)
    print("  Yve.01 — Matching AP · Sub-camino OTRAS")
    print("="*60)
    try:
        df_fact = cargar_ultimo_ap()
        df_ord  = cargar_ordenes()
    except FileNotFoundError as e:
        print(f"\n❌ {e}"); return

    print(f"  Facturas OTRAS: {len(df_fact)}  |  POs cargadas: {len(df_ord)}\n")
    resultados = []
    for _, fila in df_fact.iterrows():
        res = analizar_factura(fila, df_ord)
        icono = {"MATCH_CORRECTO":"✓","DISCREPANCIA":"✗","SIN_PO":"?","SIN_IMPORTE":"~"}.get(
            res["estado_matching"],"·")
        print(f"  [{icono}] {res['archivo']} → {res['estado_matching']}")
        if res["estado_matching"] != "MATCH_CORRECTO":
            print(f"       {res['detalle']}")
        resultados.append(res)

    df_res = pd.DataFrame(resultados)
    df_sum = generar_resumen(df_res)

    with pd.ExcelWriter(SALIDA, engine="openpyxl") as w:
        df_res.to_excel(w, index=False, sheet_name="Detalle_OTRAS")
        df_sum.to_excel(w, index=False, sheet_name="Resumen_OTRAS")
        aplicar_formato(w.sheets["Detalle_OTRAS"])
        for sn in ["Detalle_OTRAS","Resumen_OTRAS"]:
            ws = w.sheets[sn]
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col)+4, 45)

    print("\n" + "─"*60)
    print("  RESUMEN MATCHING OTRAS")
    print("─"*60)
    for _, r in df_sum.iterrows():
        print(f"  {str(r['Estado']):<28} {str(r['Cantidad']):<8} {r['Pct']}")
    print(f"\n✅ Reporte: {SALIDA}")
    print("="*60)

if __name__ == "__main__":
    main()
