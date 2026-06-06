"""
Exportador de reportes a PDF y Excel
Genera descargas de AR, AP, DRR, Multi-Hotel
"""
import os
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def crear_reporte_ar_excel():
    """Crea reporte AR (OTAs) en Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AR - OTAs"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Headers
    headers = ["OTA", "Factura #", "Fecha", "Importe EUR", "Estado", "DI Cert", "Acción"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Datos demo
    datos = [
        ["Booking.com", "BKG-2025-06-001", "2025-06-05", 2840.50, "Procesada", "SI", "—"],
        ["Expedia", "EXP-2025-06-002", "2025-06-05", 1950.75, "Pendiente", "SI", "Revisar"],
        ["Hotels.com", "HTL-2025-06-003", "2025-06-04", 3220.00, "Procesada", "NO", "Solicitar cert"],
        ["Despegar", "DSP-2025-06-004", "2025-06-04", 1645.25, "Procesada", "SI", "—"],
    ]
    
    for row_idx, row_data in enumerate(datos, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            if col_idx == 4:  # Importe
                cell.number_format = '€ #,##0.00'
    
    # Resumen
    summary_row = len(datos) + 3
    ws.cell(row=summary_row, column=1).value = "TOTAL"
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    ws.cell(row=summary_row, column=4).value = f"=SUM(D2:D{len(datos)+1})"
    ws.cell(row=summary_row, column=4).font = Font(bold=True)
    
    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    return wb

def crear_reporte_ap_excel():
    """Crea reporte AP (Proveedores) en Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AP - Proveedores"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ["Proveedor", "Factura #", "Importe EUR", "Estado 3-Way", "Aprobada", "Fecha Pago", "Notas"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    datos = [
        ["Proveedor A - F&B", "FAC-001-2025", 450.00, "OK", "SI", "2025-06-10", "—"],
        ["Proveedor B - Servicios", "FAC-002-2025", 1200.50, "Pendiente albarán", "NO", "—", "Solicitar doc"],
        ["Proveedor C - Mantenimiento", "FAC-003-2025", 850.25, "OK", "SI", "2025-06-12", "—"],
        ["Proveedor D - Suministros", "FAC-004-2025", 320.75, "OK (sin PO)", "SI", "2025-06-08", "Pequeño mto"],
    ]
    
    for row_idx, row_data in enumerate(datos, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            if col_idx == 3:
                cell.number_format = '€ #,##0.00'
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    return wb

def crear_reporte_drr_excel():
    """Crea reporte DRR (Daily Revenue Report) en Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DRR"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Título
    ws.merge_cells('A1:F1')
    title = ws['A1']
    title.value = "Daily Revenue Report - Junio 2025"
    title.font = Font(bold=True, size=14)
    
    ws.merge_cells('A2:F2')
    subtitle = ws['A2']
    subtitle.value = "Property: Hotel Demo | Date: 2025-06-05"
    subtitle.font = Font(italic=True, size=10)
    
    # KPIs principales
    row = 4
    ws[f'A{row}'] = "MÉTRICA"
    ws[f'B{row}'] = "HOY"
    ws[f'C{row}'] = "MTD"
    ws[f'D{row}'] = "BUDGET"
    ws[f'E{row}'] = "LY (Año Pasado)"
    ws[f'F{row}'] = "FORECAST"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
    
    row = 5
    kpis = [
        ["Total Revenue", "€145,230", "€1,651,590", "€1,550,000", "€1,632,400", "€1,680,000"],
        ["Occupancy %", "92.5%", "88.3%", "85.0%", "86.2%", "87.5%"],
        ["ADR", "€285.50", "€272.30", "€270.00", "€268.50", "€275.00"],
        ["RevPAR", "€264.08", "€240.54", "€229.50", "€231.15", "€240.31"],
        ["GOP", "€52,630", "€478,962", "€465,000", "€489,720", "€504,000"],
        ["GOP %", "36.2%", "29.0%", "30.0%", "30.0%", "30.0%"],
    ]
    
    for kpi_row, kpi_data in enumerate(kpis, row):
        for col_idx, value in enumerate(kpi_data, 1):
            cell = ws.cell(row=kpi_row, column=col_idx)
            cell.value = value
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="right")
    
    # F&B Analysis
    row = 12
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "F&B PERFORMANCE"
    ws[f'A{row}'].font = Font(bold=True)
    
    row = 13
    headers_fb = ["Concepto", "HOY", "MTD", "TARGET", "% Target", "Notas"]
    for col, header in enumerate(headers_fb, 1):
        ws.cell(row=row, column=col).value = header
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).fill = header_fill
    
    row = 14
    fb_data = [
        ["F&B Ventas", "€18,250", "€192,400", "€190,000", "101.3%", "OK"],
        ["F&B Costo", "€3,387", "€35,583", "€38,000", "93.6%", "OK"],
        ["F&B %", "18.6%", "18.5%", "20.0%", "92.5%", "Dentro target"],
    ]
    
    for fb_row, fb_item in enumerate(fb_data, row):
        for col_idx, value in enumerate(fb_item, 1):
            ws.cell(row=fb_row, column=col_idx).value = value
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    return wb

def crear_reporte_multihotel_excel():
    """Crea reporte Multi-Hotel consolidado en Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Multi-Hotel Summary"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Título
    ws.merge_cells('A1:H1')
    title = ws['A1']
    title.value = "Multi-Hotel Consolidado - Junio 2025"
    title.font = Font(bold=True, size=14)
    
    # Headers
    headers = ["Hotel", "Ciudad", "Rooms", "Occ%", "ADR", "RevPAR", "Revenue MTD", "GOP%"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    hoteles = [
        ["Premier London Mayfair", "London", 245, 89.6, 485.20, 434.74, 3194580, 43.7],
        ["Premier Paris Champs", "Paris", 290, 92.4, 412.50, 381.15, 3289400, 45.1],
        ["Premier Barcelona Diagonal", "Barcelona", 412, 90.1, 245.80, 221.46, 2754320, 42.8],
        ["Premier Madrid Recoletos", "Madrid", 387, 88.7, 232.40, 206.14, 2412780, 40.2],
        ["Sitges Promenade Resort", "Sitges", 210, 91.8, 158.20, 145.20, 842900, 41.5],
        ["Sitges Beach Hotel", "Sitges", 158, 87.3, 142.50, 124.40, 587450, 38.2],
    ]
    
    for row_idx, row_data in enumerate(hoteles, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            if col_idx in [4, 5, 6]:  # Porcentajes y precios
                cell.alignment = Alignment(horizontal="right")
            if col_idx == 7:  # Revenue
                cell.number_format = '€ #,##0'
    
    # Totales
    total_row = len(hoteles) + 3
    ws.cell(row=total_row, column=1).value = "TOTAL GRUPO"
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    # Cálculos
    ws.cell(row=total_row, column=3).value = f"=SUM(C3:C{total_row-1})"
    ws.cell(row=total_row, column=7).value = f"=SUM(G3:G{total_row-1})"
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = '€ #,##0'
    
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    return wb

def exportar_excel(tipo):
    """Exporta reporte específico como Excel"""
    if tipo == "ar":
        wb = crear_reporte_ar_excel()
        filename = f"AR_Reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    elif tipo == "ap":
        wb = crear_reporte_ap_excel()
        filename = f"AP_Reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    elif tipo == "drr":
        wb = crear_reporte_drr_excel()
        filename = f"DRR_Reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    elif tipo == "multihotel":
        wb = crear_reporte_multihotel_excel()
        filename = f"MultiHotel_Consolidado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        return None
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, filename
