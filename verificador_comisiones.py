"""
verificador_comisiones.py — Yve.01
Cruza Excel de facturas procesadas con comisiones pactadas.
"""
import os, re, glob
from datetime import date
import pandas as pd

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
from tenant_dirs import procesadas_dir as _t_pdir, datos_dir as _t_ddir, reportes_dir as _t_rdir
PROCESADAS_DIR= _t_pdir()
REFERENCIA_DIR= _t_ddir()
REPORTES_DIR  = _t_rdir()
os.makedirs(REPORTES_DIR, exist_ok=True)
COMISIONES_FILE = os.path.join(REFERENCIA_DIR, "comisiones_pactadas.xlsx")
FECHA_HOY = date.today().strftime("%Y%m%d")
REPORTE_SALIDA = os.path.join(REPORTES_DIR, f"verificacion_{FECHA_HOY}.xlsx")
TOLERANCIA = 0.5
NF = "NO_ENCONTRADO"

# Cobrar MENOS de lo pactado no es una discrepancia reclamable: es lo contrario.
# Iba con estado DISCREPANCIA y un importe NEGATIVO, y el panel lo sumaba en
# valor absoluto al total "a devolver". O sea que una OTA que nos cobra 182 EUR
# de MENOS aparecia como 182 EUR que reclamarle. Merece su propio estado: hay
# que mirarlo (a lo mejor el contrato esta mal cargado, o la OTA arrastra un
# ajuste) pero no se reclama, y desde luego no se suma a lo reclamable.
COBRO_DEBAJO = "COBRO_POR_DEBAJO"

try:
    from openpyxl.styles import PatternFill, Font, Alignment
    VERDE    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ROJO     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    USAR_FORMATO = True
except ImportError:
    USAR_FORMATO = False

def cargar_ultimo_excel_procesadas():
    """Facturas OTA de TODOS los dias, consolidadas.

    Antes leia solo el fichero mas reciente, asi que al cambiar de dia cruzaba
    una fraccion de las facturas y las reclamaciones del dia anterior
    desaparecian. La lectura y el deduplicado viven en almacen_datos: es el
    unico sitio a tocar cuando migremos a persistencia.
    """
    from almacen_datos import facturas_ota_para_verificar
    df, rutas = facturas_ota_para_verificar(PROCESADAS_DIR)
    if df is None or df.empty:
        raise FileNotFoundError(f"No hay facturas_procesadas_*.xlsx en {PROCESADAS_DIR}")
    print(f"  Cargando facturas desde: {len(rutas)} fichero(s) · {len(df)} facturas")
    return df, (rutas[-1] if rutas else PROCESADAS_DIR)

def _norm(v):
    """Normaliza un nombre para comparar: sin espacios sobrantes, en minusculas.
    Los vacios, NaN y NO_ENCONTRADO se tratan todos como cadena vacia."""
    s = "" if v is None else str(v)
    if s.strip() in ("", NF, "nan", "None"):
        return ""
    return " ".join(s.split()).lower()


def cargar_comisiones_pactadas():
    if not os.path.exists(COMISIONES_FILE):
        raise FileNotFoundError(f"No se encontro: {COMISIONES_FILE}")
    df = pd.read_excel(COMISIONES_FILE)
    df["OTA_norm"] = df["OTA"].map(_norm)
    # La columna Hotel es OPCIONAL. Una fila sin hotel es la tarifa GENERICA de
    # esa OTA y vale para todos sus hoteles, asi que los ficheros antiguos (que
    # no la tienen) siguen comportandose exactamente igual que antes.
    if "Hotel" in df.columns:
        df["Hotel_norm"] = df["Hotel"].map(_norm)
    else:
        df["Hotel_norm"] = ""
    return df


def buscar_tarifa(comisiones_df, ota_norm, hotel_norm):
    """Tarifa pactada aplicable a (OTA, hotel). Devuelve (fila, origen).

    Prioridad: la tarifa PROPIA del hotel > la generica de la OTA.
    Si la OTA tiene tarifas pero todas son de OTROS hoteles, devuelve
    (None, 'SIN_TARIFA_HOTEL'): antes se cogia la primera fila de la OTA, que
    es como aplicarle a un hotel el porcentaje pactado para otro -- y en un
    grupo con condiciones distintas por hotel eso reclama de mas o de menos.
    """
    de_la_ota = comisiones_df[comisiones_df["OTA_norm"] == ota_norm]
    if de_la_ota.empty:
        return None, "OTA_DESCONOCIDA"
    if hotel_norm:
        propia = de_la_ota[de_la_ota["Hotel_norm"] == hotel_norm]
        if not propia.empty:
            return propia.iloc[0], "hotel"
    generica = de_la_ota[de_la_ota["Hotel_norm"] == ""]
    if not generica.empty:
        return generica.iloc[0], "generica"
    return None, "SIN_TARIFA_HOTEL"

def convertir_porcentaje(valor):
    if valor is None or str(valor).strip() in ("", NF):
        return None
    try:
        return float(str(valor).replace(",", "."))
    except ValueError:
        return None

def convertir_importe(valor):
    if valor is None or str(valor).strip() in ("", NF):
        return None
    try:
        limpio = str(valor).replace("EUR","").replace("USD","").replace(" ","").strip()
        limpio = limpio.replace("\xa0","")
        # Quitar simbolo euro/dollar si quedara
        limpio = limpio.replace("\u20ac","").strip()
        if "," in limpio and "." in limpio:
            if limpio.rfind(".") > limpio.rfind(","):
                limpio = limpio.replace(",","")         # US: 8,300.00 -> 8300.00
            else:
                limpio = limpio.replace(".","").replace(",",".")  # EU: 8.300,00
        elif "," in limpio:
            if re.search(r",\d{3}$", limpio):
                limpio = limpio.replace(",","")
            else:
                limpio = limpio.replace(",",".")
        return float(limpio)
    except ValueError:
        return None

def verificar_factura(fila, comisiones_df):
    ota_nombre = str(fila.get("nombre_ota", NF))
    ota_norm   = _norm(ota_nombre)
    hotel_nombre = str(fila.get("nombre_hotel", NF))
    hotel_norm = _norm(hotel_nombre)
    tarifa, origen = buscar_tarifa(comisiones_df, ota_norm, hotel_norm)
    mercado = NF
    comision_pactada = None
    diferencia = None
    importe_discrepancia = None
    tarifa_aplicada = NF

    if ota_norm == "" or tarifa is None:
        estado = origen if origen == "SIN_TARIFA_HOTEL" else "OTA_DESCONOCIDA"
    else:
        comision_pactada = float(tarifa["Porcentaje_Comision"])
        mercado = tarifa["Mercado"]
        # Deja constancia de QUE tarifa se ha usado: un controller tiene que
        # poder auditar si se aplico la del hotel o la generica de la OTA.
        tarifa_aplicada = (str(tarifa.get("Hotel") or hotel_nombre)
                           if origen == "hotel" else f"{ota_nombre} (genérica)")
        comision_factura = convertir_porcentaje(fila.get("porcentaje_comision"))
        importe_bruto    = convertir_importe(fila.get("importe_bruto"))
        if comision_factura is None:
            estado = "SIN_PORCENTAJE"
        else:
            diferencia = abs(comision_factura - comision_pactada)
            if diferencia < TOLERANCIA:
                estado = "CORRECTO"
            else:
                # El SIGNO decide de que estamos hablando. Positivo: la OTA se
                # ha cobrado mas de lo pactado y hay algo que reclamar.
                # Negativo: se ha cobrado menos, y eso no se reclama.
                estado = ("DISCREPANCIA" if comision_factura > comision_pactada
                          else COBRO_DEBAJO)
                if importe_bruto is not None:
                    importe_discrepancia = round(
                        importe_bruto * (comision_factura - comision_pactada) / 100, 2
                    )
    return {
        "archivo":          fila.get("archivo", NF),
        "numero_factura":   fila.get("numero_factura", NF),
        "fecha":            fila.get("fecha", NF),
        "nombre_ota":       ota_nombre,
        "mercado":          mercado,
        "nombre_hotel":     fila.get("nombre_hotel", NF),
        "periodo_inicio":   fila.get("periodo_inicio", NF),
        "periodo_fin":      fila.get("periodo_fin", NF),
        "importe_bruto":    fila.get("importe_bruto", NF),
        "tarifa_aplicada":  tarifa_aplicada,
        "porcentaje_pactado":       comision_pactada,
        "porcentaje_factura":       fila.get("porcentaje_comision", NF),
        "diferencia_pp":    round(diferencia,2) if diferencia is not None else NF,
        "importe_comision_factura": fila.get("importe_comision", NF),
        "importe_neto":     fila.get("importe_neto", NF),
        "discrepancia_euros": importe_discrepancia if importe_discrepancia is not None else NF,
        "estado":           estado,
        # El hotel viaja con la factura de una etapa a la siguiente.
        #
        # Este diccionario se escribe a mano, campo por campo, asi que lo que
        # no se copie aqui se PIERDE. `hotel_id` no estaba, y eso rompia dos
        # cosas a la vez: la clave de deduplicacion de `almacen_datos` lleva el
        # hotel al final, o sea que esta fila (sin hotel) y la de
        # facturas_procesadas_*.xlsx (con hotel) dejaban de ser la misma
        # factura y salian DUPLICADAS; y encima la fila enriquecida —la que
        # trae estado y discrepancia— se iba a "sin asignar", asi que al elegir
        # un hotel solo quedaba la version cruda, sin analisis.
        #
        # Va al final para no mover ninguna columna de sitio. El formateado del
        # Excel busca "estado" por nombre, no por posicion, asi que no le
        # afecta (comprobado en aplicar_formato_excel).
        "hotel_id":         fila.get("hotel_id", ""),
    }

def aplicar_formato_excel(ws, df):
    if not USAR_FORMATO:
        return
    col_estado = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "estado":
            col_estado = idx
            break
    if col_estado is None:
        return
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        estado = ws.cell(row=row_idx, column=col_estado).value
        fill = {"CORRECTO": VERDE, "DISCREPANCIA": ROJO, "OTA_DESCONOCIDA": AMARILLO, "SIN_PORCENTAJE": AMARILLO,
                "SIN_TARIFA_HOTEL": AMARILLO}.get(estado)
        if fill:
            for cell in row:
                cell.fill = fill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

def generar_resumen(df_resultado):
    total = len(df_resultado)
    resumen = df_resultado["estado"].value_counts().reset_index()
    resumen.columns = ["Estado", "Cantidad"]
    resumen["Porcentaje"] = (resumen["Cantidad"]/total*100).round(1).astype(str)+"%"
    discrepancias = df_resultado[df_resultado["estado"]=="DISCREPANCIA"]
    montos = [v for v in (convertir_importe(r.get("discrepancia_euros")) for _,r in discrepancias.iterrows()) if v is not None]
    total_disc = sum(montos)
    extra = pd.DataFrame([
        {"Estado":"─"*20,"Cantidad":"","Porcentaje":""},
        {"Estado":"TOTAL FACTURAS","Cantidad":total,"Porcentaje":"100%"},
        {"Estado":"DISCREPANCIA TOTAL (EUR)","Cantidad":f"{total_disc:.2f} EUR","Porcentaje":""},
    ])
    return pd.concat([resumen, extra], ignore_index=True)

def main():
    print("="*60)
    print("  Yve.01 — Verificador de Comisiones OTA")
    print("="*60)
    try:
        df_facturas, _ = cargar_ultimo_excel_procesadas()
        df_comisiones  = cargar_comisiones_pactadas()
    except FileNotFoundError as e:
        print(f"\nError: {e}")
        return
    print(f"  Tabla de referencia: {len(df_comisiones)} OTAs cargadas")
    print(f"  Facturas a verificar: {len(df_facturas)}\n")
    resultados = []
    for _, fila in df_facturas.iterrows():
        resultado = verificar_factura(fila, df_comisiones)
        icono = {"CORRECTO":"v","DISCREPANCIA":"X","OTA_DESCONOCIDA":"?","SIN_PORCENTAJE":"~"}.get(resultado["estado"],".")
        disc_str = f"  ({resultado['discrepancia_euros']} EUR)" if resultado["estado"]=="DISCREPANCIA" and resultado["discrepancia_euros"]!=NF else ""
        print(f"  [{icono}] {resultado['archivo']} -> {resultado['estado']}{disc_str}")
        resultados.append(resultado)
    df_resultado = pd.DataFrame(resultados)
    df_resumen   = generar_resumen(df_resultado)
    with pd.ExcelWriter(REPORTE_SALIDA, engine="openpyxl") as writer:
        df_resultado.to_excel(writer, index=False, sheet_name="Detalle")
        df_resumen.to_excel(writer, index=False, sheet_name="Resumen")
        aplicar_formato_excel(writer.sheets["Detalle"], df_resultado)
        for sheet in ["Detalle","Resumen"]:
            ws = writer.sheets[sheet]
            for col in ws.columns:
                ancho = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(ancho+4, 40)
    print("\n"+"─"*60)
    print("  RESUMEN VERIFICACION")
    print("─"*60)
    for _, row in df_resumen.iterrows():
        print(f"  {str(row['Estado']):<30} {str(row['Cantidad']):<12} {row['Porcentaje']}")
    print(f"\nReporte guardado en: {REPORTE_SALIDA}")
    print("="*60)

if __name__ == "__main__":
    main()
