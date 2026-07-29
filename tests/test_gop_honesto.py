"""FASE D · El GOP dice de donde sale, y el inventado no existe.

Tres cosas:

  1. La rama del 22% (`ingresos x 0,22`, "media industria hotelera España") no
     esta en el codigo. Se comprueba sobre el AST y sobre el texto, porque el
     numero podria colarse escrito de otra forma.
  2. `_leer_drr_stats` marca cada periodo con su procedencia, y con un DRR que
     no trae GOP ni forma de derivarlo devuelve N/D — no un numero.
  3. El agregador no deja entrar lo `inventado` en un numero de grupo.
"""
import ast
import os
import sys

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE)

os.environ.setdefault("YVE_TENANT", "fase-d-test")


def test_la_rama_del_22_no_existe():
    fuente = open(os.path.join(BASE, "dashboard.py"), encoding="utf-8").read()
    arbol = ast.parse(fuente)

    # 1) Ningun 0.22 ni 22.0 sueltos en el codigo (los comentarios y docstrings
    #    no son nodos Constant fuera de su sitio, asi que el AST no los ve).
    sospechosos = []
    for nodo in ast.walk(arbol):
        if isinstance(nodo, ast.Constant) and isinstance(nodo.value, float):
            if abs(nodo.value - 0.22) < 1e-9:
                sospechosos.append(f"linea {nodo.lineno}: {nodo.value}")
    assert not sospechosos, f"sigue habiendo un 0.22 en el codigo: {sospechosos}"

    # 2) Ni el texto que lo acompañaba, que es lo que delataba la estimacion.
    #    Se busca en las CADENAS del AST, no en el fichero: el comentario que
    #    explica que la rama se borro menciona el 22% a proposito, y un grep se
    #    lo comeria como si fuera el codigo.
    literales = [n.value for n in ast.walk(arbol)
                 if isinstance(n, ast.Constant) and isinstance(n.value, str)]
    for veneno in ("22.0% ~", "media industria hotelera"):
        malos = [s for s in literales if veneno in s]
        assert not malos, f"sigue estando la cadena {veneno!r}: {malos[:2]}"


def test_procedencia_en_el_resultado():
    """Un DRR sin GOP y sin presupuesto sale N/D, no estimado."""
    import openpyxl
    import tempfile
    import dashboard

    carpeta = tempfile.mkdtemp(prefix="drr_fase_d_")
    ruta = os.path.join(carpeta, "drr_procesado_20260729.xlsx")

    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.title = "Resumen"
    # Ingresos si, GOP no, GOP% no, presupuesto no: el caso exacto en el que
    # antes saltaba el 22%.
    #        nombre            today   mtd    forecast  budget
    filas = [["Total Revenue",  10000, 100000, 120000, None],
             ["Occupancy %",     0.80,   0.78,   0.82, None]]
    for f in filas:
        hoja.append(f)
    wb.create_sheet("Alertas")
    wb.save(ruta)

    stats = dashboard._leer_drr_stats(ruta)
    assert "error" not in stats, stats.get("error")

    gop  = stats["metricas"].get("GOP", {})
    gpct = stats["metricas"].get("GOP %", {})
    proc = stats.get("gop_procedencia", {})

    for periodo in ("today", "mtd", "forecast"):
        assert gop.get(periodo, "N/D") == "N/D", \
            f"{periodo}: GOP deberia ser N/D y es {gop.get(periodo)!r} — ¿ha vuelto la estimacion?"
        assert gpct.get(periodo, "N/D") == "N/D", \
            f"{periodo}: GOP % deberia ser N/D y es {gpct.get(periodo)!r}"
        assert proc.get(periodo) == "sin_datos", \
            f"{periodo}: procedencia {proc.get(periodo)!r}, esperaba 'sin_datos'"

    # Y el caso contrario: si el DRR trae el GOP%, SI se deriva — porque es un
    # dato del propio hotel. Quitar la estimacion inventada no puede llevarse
    # por delante la aritmetica legitima.
    wb2 = openpyxl.Workbook()
    h2 = wb2.active
    h2.title = "Resumen"
    for f in [["Total Revenue", 10000, 100000, 120000, None],
              ["GOP %",          0.30,   0.30,   0.30, None]]:
        h2.append(f)
    wb2.create_sheet("Alertas")
    ruta2 = os.path.join(carpeta, "drr_procesado_20260730.xlsx")
    wb2.save(ruta2)

    stats2 = dashboard._leer_drr_stats(ruta2)
    proc2 = stats2.get("gop_procedencia", {})
    assert proc2.get("mtd") == "derivado", \
        f"con GOP% del hotel deberia derivarse, y sale {proc2.get('mtd')!r}"
    assert stats2["metricas"]["GOP"]["mtd"] != "N/D", "no ha derivado el GOP en euros"

    import shutil
    shutil.rmtree(carpeta, ignore_errors=True)


def test_el_grupo_no_agrega_lo_inventado():
    from agregador_grupo import agregable, PROCEDENCIAS_AGREGABLES

    assert agregable("medido")
    assert agregable("derivado")
    assert not agregable("inventado")
    assert not agregable("sin_datos")
    # Falla en cerrado: una procedencia nueva que nadie ha decidido se queda
    # FUERA del total, no dentro.
    assert not agregable("estimado_por_la_ia")
    assert not agregable(None)
    assert not agregable("")
    assert "inventado" not in PROCEDENCIAS_AGREGABLES


if __name__ == "__main__":
    test_la_rama_del_22_no_existe()
    test_procedencia_en_el_resultado()
    test_el_grupo_no_agrega_lo_inventado()
    print("OK · el GOP dice de donde sale, y el 22% inventado ya no existe")
