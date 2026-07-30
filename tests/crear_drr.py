"""Construye un DRR de prueba, sin dependencias fuera de openpyxl.

EXISTE PARA QUE NO FALTE. `tests/fixtures_drr_fase_e.py` importa
`from crear import construir`, y `tests/crear.py` nunca se commiteó: por eso no
había NINGUNA prueba ejecutable que recorriera el DRR, y por eso el desajuste
INCOME/REVENUE del revenue diario sobrevivió desde el primer commit del lector
hasta que lo destapó una prueba de integración a mano.

El formato es el que lee `lector_drr`:

  DAILY_MASTER  fila 2 col B = la fecha del informe. Las metricas por nombre en
                la columna A, con los valores en C (today), D (mtd),
                F (forecast) y G (budget).
  hojas "1".."N"  Trial Balance de cada dia. D1 = "OK" | "Out of Balance",
                E1 = el descuadre. D3 = la fecha del dia. Las cuentas: col C el
                nombre, H debe, I haber, J total. Las cabeceras de seccion en la
                columna A, y tienen que ser una de `lector_drr.SECTIONS`.
  CtaCble       el plan de cuentas.
"""
import os

# Fila 1-based de cada metrica en DAILY_MASTER. Es el mismo mapa que usa el
# generador del juego de integracion; si `lector_drr` cambia de sitio una
# metrica, este mapa se entera porque la prueba deja de encontrar el valor.
FILA = {
    "Occupancy %": 11, "Rooms Occupied": 12, "ADR": 13, "Revenue PAR": 14,
    "Rooms Revenue": 15, "Food Revenue": 18, "Beverage Revenue": 19,
    "F&B Other": 20, "F&B Revenue Total": 21, "Telephone / Other": 24,
    "Total Revenue": 28, "Spend PAR": 29, "GOP": 31, "GOP %": 32,
}


def construir(ruta, metricas, dias, hotel="Hotel de Prueba",
              fecha_informe="2026-07-03"):
    """Escribe un .xlsm de DRR en `ruta` y la devuelve.

    `metricas`: {nombre: (today, mtd, forecast, budget)}
    `dias`:     [(n, 'YYYY-MM-DD', [(seccion, cuenta, debe, haber)], descuadre)]

    OJO con `seccion`: se escribe TAL CUAL. Es a proposito — la prueba necesita
    poder mandar "INCOME" en un dia y "REVENUE" en otro, porque el lector acepta
    las dos y el panel tiene que tratarlas igual.
    """
    from openpyxl import Workbook
    wb = Workbook()

    ws = wb.active
    ws.title = "DAILY_MASTER"
    ws["A1"] = f"DAILY REVENUE REPORT — {hotel}"
    ws["B2"] = fecha_informe
    ws["A4"] = "KPI METRICS"
    ws["C10"] = "Today"
    ws["D10"] = "MTD"
    ws["F10"] = "Full Month Forecast"
    ws["G10"] = "Budget"
    for met, fila in FILA.items():
        ws.cell(fila, 1, met)
        vals = metricas.get(met)
        if not vals:
            continue
        for col, v in zip((3, 4, 6, 7), vals):
            if v is not None:
                ws.cell(fila, col, v)

    for n, fecha, cuentas, descuadre in dias:
        w = wb.create_sheet(str(n))
        w["A1"] = f"TRIAL BALANCE — DAY {n}"
        w["D1"] = "Out of Balance" if descuadre else "OK"
        w["E1"] = descuadre
        w["C3"] = "ACCOUNT NAME"
        w["D3"] = fecha
        fila = 5
        puesta = None
        td = tc = 0.0
        for seccion, cuenta, debe, haber in cuentas:
            if seccion != puesta:
                w.cell(fila, 1, seccion)
                fila += 1
                puesta = seccion
            w.cell(fila, 3, cuenta)
            w.cell(fila, 8, debe)
            w.cell(fila, 9, haber)
            w.cell(fila, 10, round(debe - haber, 2))
            td += debe
            tc += haber
            fila += 1
        w.cell(fila + 1, 3, "TOTAL")
        w.cell(fila + 1, 8, round(td, 2))
        w.cell(fila + 1, 9, round(tc + descuadre, 2))

    wc = wb.create_sheet("CtaCble")
    wc.append(["Entity", "Department", "Account", "Line Description"])
    for f in [("1010", "ROOMS", "400100", "Room Revenue Transient"),
              ("1010", "ROOMS", "400200", "Room Revenue Group"),
              ("1010", "FB", "410100", "Food Revenue Restaurant"),
              ("1010", "ADMIN", "610100", "Salaries and Wages")]:
        wc.append(list(f))

    os.makedirs(os.path.dirname(os.path.abspath(ruta)), exist_ok=True)
    wb.save(ruta)
    return ruta
