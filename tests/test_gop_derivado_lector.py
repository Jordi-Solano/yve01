# -*- coding: utf-8 -*-
"""Inventario honesto #24: dos derivaciones del GOP que hace el LECTOR (del
GOP% del forecast / del presupuesto) llegaban al panel etiquetadas "Medido".
Ahora el lector deja la fila "GOP (procedencia)" en la hoja Resumen y el
panel etiqueta "derivado" con su origen, con la ~ de siempre.

  python3.12 tests/test_gop_derivado_lector.py
  python3.12 tests/test_gop_derivado_lector.py --sabotaje
"""
import os, subprocess, sys, tempfile
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE); os.chdir(BASE)
sys.path.insert(0, os.path.join(BASE, 'tests'))
SABOTAJE = '--sabotaje' in sys.argv


def main():
    fallos = 0
    def ok(cond, msg):
        nonlocal fallos
        print(f"  {'OK ' if cond else 'FALLA'}  {msg}")
        if not cond: fallos += 1
    import openpyxl
    import lector_drr as L
    import dashboard as D
    from crear_drr import construir
    tmp = tempfile.mkdtemp(prefix='gopd_')
    # DRR con ingresos y GOP forecast, pero SIN GOP today/mtd (formulas sin cachear): el lector lo deriva
    xlsm = construir(os.path.join(tmp, 'drr.xlsm'), {"Total Revenue": (10000, 100000, 120000, 110000), "GOP": (None, None, 36000, 33000), "GOP %": (None, None, 0.30, 0.30), "Occupancy %": (0.8, 0.78, 0.82, 0.8)},
                     [(1, '2026-08-01', [("INCOME", "Rooms", 0, 10000)], 0.0)])
    wb = openpyxl.load_workbook(xlsm, data_only=True)
    met = L.leer_daily_master(wb)
    if SABOTAJE:
        met.pop("__gop_derivado", None)
    der = met.get("__gop_derivado") or {}
    ok(met["GOP"].get("mtd") is not None and der.get("mtd"), f"el lector deriva el GOP MTD y lo anota: {der}")
    # la hoja Resumen lleva la fila de procedencia
    out = openpyxl.Workbook(); ws = out.active; ws.title = "Resumen"
    L.escribir_hoja_resumen(ws, met); out.create_sheet("Alertas")
    ruta = os.path.join(tmp, 'drr_procesado_20260801.xlsx'); out.save(ruta)
    filas = {str(r[0]): r for r in ws.iter_rows(min_row=5, values_only=True) if r and r[0]}
    ok("GOP (procedencia)" in filas and str(filas["GOP (procedencia)"][2]).startswith("derivado"), f"fila 'GOP (procedencia)' en la hoja: {filas.get('GOP (procedencia)')}")
    # y el panel la lee: derivado, no medido
    st = D._leer_drr_stats(ruta)
    proc = st.get("gop_procedencia", {})
    ok(proc.get("mtd") == "derivado" and "forecast" in (proc.get("mtd_origen") or "").lower() or proc.get("mtd") == "derivado", f"panel: GOP MTD = {proc.get('mtd')} ({proc.get('mtd_origen')})")
    ok(str(st["metricas"]["GOP"]["mtd"]).endswith("~"), f"y lleva la ~ de 'derivado': {st['metricas']['GOP']['mtd']}")
    # con GOP medido de verdad sigue siendo 'medido'
    xlsm2 = construir(os.path.join(tmp, 'drr2.xlsm'), {"Total Revenue": (10000, 100000, 120000, 110000), "GOP": (3000, 30000, 36000, 33000), "GOP %": (0.3, 0.3, 0.30, 0.30)}, [(1, '2026-08-01', [("INCOME", "Rooms", 0, 10000)], 0.0)])
    met2 = L.leer_daily_master(openpyxl.load_workbook(xlsm2, data_only=True))
    out2 = openpyxl.Workbook(); ws2 = out2.active; ws2.title = "Resumen"; L.escribir_hoja_resumen(ws2, met2); out2.create_sheet("Alertas")
    ruta2 = os.path.join(tmp, 'drr_procesado_20260802.xlsx'); out2.save(ruta2)
    st2 = D._leer_drr_stats(ruta2)
    ok(st2.get("gop_procedencia", {}).get("mtd") == "medido" and not str(st2["metricas"]["GOP"]["mtd"]).endswith("~"), f"GOP medido de verdad: {st2.get('gop_procedencia', {}).get('mtd')}")
    import shutil; shutil.rmtree(tmp, ignore_errors=True)
    diff = subprocess.run(['git', 'diff', '--name-only', 'HEAD'], capture_output=True, text=True, cwd=BASE).stdout.split()
    ok(not [f for f in diff if f.startswith('oracle_') or f == 'lector_facturas_ap.py'], 'ni oracle_* ni clasificador')
    print()
    if SABOTAJE:
        print('SABOTAJE: se esperaban fallos' if fallos else '*** SABOTAJE SIN EFECTO ***'); sys.exit(0 if fallos else 1)
    print('TODO OK' if not fallos else f'{fallos} FALLOS'); sys.exit(1 if fallos else 0)


if __name__ == '__main__':
    main()
