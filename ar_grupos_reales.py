"""
AR Real Module — Procesa facturas de grupos corporativos (Property Corporate Client)
Estructura real: Rooming list + facturas por subgrupo + BEOs
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
import json
import os

# Paths
UPLOADS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ar_real_data")
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reportes")
os.makedirs(OUTPUT_PATH, exist_ok=True)

print("[AR REAL] Procesando facturas de grupos corporativos...\n")

# ============================================================================
# PASO 1: Leer Rooming List
# ============================================================================
print("1. Leyendo rooming list...")
rooming_file = f"{UPLOADS_PATH}/rooming_demo.xlsx"
wb_rooming = openpyxl.load_workbook(rooming_file, read_only=True, data_only=True)
ws_rooming = wb_rooming["Rooming Template"]

attendees = []
for row in ws_rooming.iter_rows(min_row=4, values_only=True):
    name = row[7]   # NAME
    surname = row[8]  # SURNAME
    checkin = row[9]
    checkout = row[10]
    comments = row[17]  # COMMENTS
    
    if name and surname:
        attendees.append({
            "name": name,
            "surname": surname,
            "checkin": checkin,
            "checkout": checkout,
            "nights": 1,  # Base estimate
            "group": "Subgroup B" if (comments and "Subgroup B" in str(comments)) else "Master Account"
        })

master_count = sum(1 for a in attendees if a['group']=='Master Account')
portugal_count = sum(1 for a in attendees if a['group']=='Subgroup B')

print(f"   ✓ {len(attendees)} asistentes")
print(f"     Master Account: {master_count}")
print(f"     Subgroup B prepay: {portugal_count}\n")

# ============================================================================
# PASO 2: Procesar Factura Subgroup A (xlsm)
# ============================================================================
print("2. Procesando factura Subgroup A...")
poland_file = f"{UPLOADS_PATH}/invoice_subgroup_demo.xlsm"
wb_poland = openpyxl.load_workbook(poland_file, read_only=True, data_only=True)
ws_poland = wb_poland.active

# Extraer invoice data
invoice_no = "INV-001"
invoice_date = datetime(2025, 7, 30)
group_name = "Corporate Subgroup"

# Procesar líneas
deposit = 0
consumed = {}
for row in ws_poland.iter_rows(min_row=24, max_row=27, values_only=True):
    desc = row[4] if len(row) > 4 else None
    qty = row[2] if len(row) > 2 else 1
    price = row[8] if len(row) > 8 else 0
    total = row[9] if len(row) > 9 else 0
    
    if desc:
        desc = str(desc).strip()
        if "DEPOSIT" in desc:
            deposit = float(price) if price else 0
        else:
            consumed[desc] = {
                "qty": qty,
                "unit_price": float(price) if price else 0,
                "total": float(total) if total else 0
            }

total_consumed = sum(v["total"] for v in consumed.values())
balance_poland = deposit - total_consumed  # Positivo = saldo a favor

print(f"   ✓ Invoice {invoice_no}")
print(f"     Deposit: €{deposit:,.2f}")
print(f"     Consumed: €{total_consumed:,.2f}")
print(f"     Balance: €{balance_poland:,.2f} (saldo a favor)\n")

# ============================================================================
# PASO 3: Crear Excel con consolidado
# ============================================================================
print("3. Generando reporte...\n")

output_file = f"{OUTPUT_PATH}/ar_real_grupo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
title_font = Font(bold=True, size=14)
subheader_font = Font(bold=True, size=11)
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ========== SHEET 1: OVERVIEW ==========
ws1 = wb.create_sheet("Overview", 0)
row = 1

ws1[f"A{row}"] = "AR REAL — Corporate Group Event"
ws1[f"A{row}"].font = title_font
row += 1
ws1[f"A{row}"] = "Property | 03-06 July 2025"
ws1[f"A{row}"].font = Font(italic=True)
row += 2

# Tabla resumen
ws1[f"A{row}"] = "Concepto"
ws1[f"B{row}"] = "Cantidad"
ws1[f"C{row}"] = "Unidad"
for col in ["A", "B", "C"]:
    ws1[f"{col}{row}"].font = header_font
    ws1[f"{col}{row}"].fill = header_fill
row += 1

ws1[f"A{row}"] = "Total Attendees"
ws1[f"B{row}"] = len(attendees)
ws1[f"C{row}"] = "personas"
row += 1

ws1[f"A{row}"] = "  - Master Account"
ws1[f"B{row}"] = master_count
ws1[f"C{row}"] = "personas"
row += 1

ws1[f"A{row}"] = "  - Subgroup B Group"
ws1[f"B{row}"] = portugal_count
ws1[f"C{row}"] = "personas"
row += 2

ws1[f"A{row}"] = "Invoices Processed"
ws1[f"A{row}"].font = subheader_font
row += 1

ws1[f"A{row}"] = "Subgroup A (GRUPA EVENT)"
ws1[f"B{row}"] = f"€{total_consumed:,.2f}"
ws1[f"C{row}"] = "consumed"
row += 1

ws1[f"A{row}"] = "Deposit Subgroup A"
ws1[f"B{row}"] = f"€{deposit:,.2f}"
ws1[f"C{row}"] = "received"
row += 1

ws1[f"A{row}"] = "Balance Subgroup A"
ws1[f"B{row}"] = f"€{balance_poland:,.2f}"
ws1[f"C{row}"] = "credit"
ws1[f"B{row}"].font = Font(bold=True, color="008000" if balance_poland > 0 else "FF0000")
row += 2

ws1[f"A{row}"] = "Status"
ws1[f"A{row}"].font = subheader_font
row += 1

ws1[f"A{row}"] = "Invoice:"
ws1[f"B{row}"] = "PROCESSED"
ws1[f"B{row}"].font = Font(color="008000", bold=True)
row += 1

ws1[f"A{row}"] = "Last Updated:"
ws1[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

for col in ["A", "B", "C"]:
    ws1.column_dimensions[col].width = 25

# ========== SHEET 2: ROOMING DETAIL ==========
ws2 = wb.create_sheet("Rooming List", 1)
row = 1

cols = ["Name", "Surname", "Check-In", "Check-Out", "Group", "Payment Type"]
for col_idx, col_name in enumerate(cols, 1):
    cell = ws2.cell(row=row, column=col_idx)
    cell.value = col_name
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

row = 2
for att in attendees:
    ws2.cell(row=row, column=1).value = att["name"]
    ws2.cell(row=row, column=2).value = att["surname"]
    ws2.cell(row=row, column=3).value = att["checkin"].strftime("%Y-%m-%d") if att["checkin"] else ""
    ws2.cell(row=row, column=4).value = att["checkout"].strftime("%Y-%m-%d") if att["checkout"] else ""
    ws2.cell(row=row, column=5).value = att["group"]
    ws2.cell(row=row, column=6).value = "Master Bill" if att["group"] == "Master Account" else "Pre-paid"
    
    for col in range(1, 7):
        ws2.cell(row=row, column=col).border = border
    row += 1

for col_idx, col_name in enumerate(cols, 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = 15

# ========== SHEET 3: INVOICES ==========
ws3 = wb.create_sheet("Invoices", 2)
row = 1

# Header
headers_inv = ["Subgroup", "Invoice #", "Invoice Date", "Description", "Qty", "Unit Price EUR", "Total EUR", "Type"]
for col_idx, col_name in enumerate(headers_inv, 1):
    cell = ws3.cell(row=row, column=col_idx)
    cell.value = col_name
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

row = 2

# Subgroup A lines
ws3.cell(row=row, column=1).value = "Subgroup A"
ws3.cell(row=row, column=2).value = invoice_no
ws3.cell(row=row, column=3).value = invoice_date.strftime("%Y-%m-%d")
ws3.cell(row=row, column=4).value = "DEPOSIT RECEIPT"
ws3.cell(row=row, column=5).value = 1
ws3.cell(row=row, column=6).value = deposit
ws3.cell(row=row, column=7).value = deposit
ws3.cell(row=row, column=8).value = "Deposit"
for col in range(1, 9):
    ws3.cell(row=row, column=col).border = border
row += 1

for desc, data in consumed.items():
    ws3.cell(row=row, column=1).value = "Subgroup A"
    ws3.cell(row=row, column=2).value = invoice_no
    ws3.cell(row=row, column=3).value = invoice_date.strftime("%Y-%m-%d")
    ws3.cell(row=row, column=4).value = desc
    ws3.cell(row=row, column=5).value = data["qty"]
    ws3.cell(row=row, column=6).value = data["unit_price"]
    ws3.cell(row=row, column=7).value = data["total"]
    ws3.cell(row=row, column=8).value = "Service"
    for col in range(1, 9):
        ws3.cell(row=row, column=col).border = border
    row += 1

for col_idx in range(1, 9):
    ws3.column_dimensions[get_column_letter(col_idx)].width = 18

# ========== SHEET 4: RECONCILIATION ==========
ws4 = wb.create_sheet("Reconciliation", 3)
row = 1

ws4[f"A{row}"] = "AR RECONCILIATION"
ws4[f"A{row}"].font = title_font
row += 2

ws4[f"A{row}"] = "Item"
ws4[f"B{row}"] = "EUR"
for col in ["A", "B"]:
    ws4[f"{col}{row}"].font = header_font
    ws4[f"{col}{row}"].fill = header_fill
    ws4[f"{col}{row}"].border = border
row += 1

# Líneas
ws4[f"A{row}"] = "Subgroup A - Deposit Received"
ws4[f"B{row}"] = deposit
row += 1

ws4[f"A{row}"] = "Subgroup A - Services Consumed"
ws4[f"B{row}"] = total_consumed
row += 1

ws4[f"A{row}"] = "Subgroup A - Balance (Credit)"
ws4[f"B{row}"] = balance_poland
ws4[f"A{row}"].font = subheader_font
ws4[f"B{row}"].font = Font(bold=True, color="008000")
row += 2

ws4[f"A{row}"] = "Master Account (62 guests)"
ws4[f"A{row}"].font = subheader_font
row += 1

ws4[f"A{row}"] = "  - Expected rooms revenue (62 x €210)"
ws4[f"B{row}"] = 62 * 210
row += 1

ws4[f"A{row}"] = "  - Subgroup B sub-group (5 x €210)"
ws4[f"B{row}"] = 5 * 210
row += 2

ws4[f"A{row}"] = "TOTAL AR BILLABLE"
ws4[f"B{row}"] = (master_count * 210) + balance_poland
ws4[f"A{row}"].font = Font(bold=True, size=12)
ws4[f"B{row}"].font = Font(bold=True, size=12, color="1F4E78")
ws4[f"A{row}"].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
ws4[f"B{row}"].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

for col in ["A", "B"]:
    ws4.column_dimensions[col].width = 35

# Guardar
wb.save(output_file)
print(f"✓ Reporte guardado: {os.path.basename(output_file)}")
print(f"  Path: {output_file}\n")

print("="*70)
print("RESUMEN AR REAL")
print("="*70)
print(f"Evento: Corporate Group Event")
print(f"Hotel: Property")
print(f"Fechas: 03-06 Julio 2025")
print()
print(f"Rooming Analysis:")
print(f"  Total: {len(attendees)} attendees")
print(f"    Master Account: {master_count} (bill to Corporate Client)")
print(f"    Subgroup B: {portugal_count} (pre-paid)")
print()
print(f"Subgroup A Invoice (INV-001):")
print(f"  Deposit received: €{deposit:,.2f}")
print(f"  Services consumed: €{total_consumed:,.2f}")
print(f"  → Balance: €{balance_poland:,.2f} CREDIT")
print()
print(f"Next steps:")
print(f"  [ ] Process Subgroup B PDF invoices")
print(f"  [ ] Extract BEO (banquet) charges")
print(f"  [ ] Reconcile Master Account final balance")
print(f"  [ ] Create AR summary for accounting")
print("="*70)
