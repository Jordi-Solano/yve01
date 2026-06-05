"""
fb_cost_control.py — Módulo F&B Cost Control para Yve
Calcula: Food Cost % real vs teórico, coste por categoría, inventario, mermas
Fuentes: recetas.xlsx + ventas_fb_diarias.xlsx + inventario.xlsx + mermas.xlsx
"""

import json
import pandas as pd
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DATOS = BASE_DIR / "datos-referencia"
REPORTES = BASE_DIR / "reportes"
try:
    REPORTES.mkdir(exist_ok=True)
except (PermissionError, OSError):
    pass  # Render filesystem is read-only, use /tmp instead

# ─── CARGA DE DATOS ────────────────────────────────────────────────────────────

def cargar_recetas():
    df = pd.read_excel(DATOS / "recetas.xlsx")
    recetas = {}
    for _, row in df.iterrows():
        ingredientes = json.loads(row["ingredientes_json"])
        coste_teorico = sum(i["cantidad"] * i["coste_unitario"] for i in ingredientes)
        recetas[row["id_receta"]] = {
            "nombre": row["nombre"],
            "categoria": row["categoria"],
            "precio_venta": row["precio_venta"],
            "coste_teorico": round(coste_teorico, 4),
            "food_cost_teorico_pct": round(coste_teorico / row["precio_venta"] * 100, 2),
            "ingredientes": ingredientes,
        }
    return recetas

def cargar_ventas():
    return pd.read_excel(DATOS / "ventas_fb_diarias.xlsx")

def cargar_inventario():
    return pd.read_excel(DATOS / "inventario.xlsx")

def cargar_mermas():
    return pd.read_excel(DATOS / "mermas.xlsx")

# ─── CÁLCULO FOOD COST TEÓRICO ─────────────────────────────────────────────────

def calcular_food_cost_teorico(recetas, ventas_df):
    """
    Coste teórico = suma(unidades_vendidas × coste_ingredientes por plato)
    Asume cocina perfecta: sin mermas, sin desviaciones.
    """
    resultados = []
    for _, venta in ventas_df.iterrows():
        rid = venta["id_receta"]
        if rid not in recetas:
            continue
        rec = recetas[rid]
        coste_total_teorico = venta["unidades_vendidas"] * rec["coste_teorico"]
        resultados.append({
            "fecha": venta["fecha"],
            "id_receta": rid,
            "nombre": venta["nombre_plato"],
            "categoria": venta["categoria"],
            "unidades": venta["unidades_vendidas"],
            "precio_venta_unit": rec["precio_venta"],
            "venta_total": venta["total_venta"],
            "coste_teorico_unit": rec["coste_teorico"],
            "coste_total_teorico": round(coste_total_teorico, 2),
            "food_cost_teorico_pct": rec["food_cost_teorico_pct"],
        })
    return pd.DataFrame(resultados)

# ─── CÁLCULO FOOD COST REAL (incluye mermas) ──────────────────────────────────

def calcular_food_cost_real(teorico_df, mermas_df):
    """
    Food Cost Real = Coste Teórico + Mermas del período
    """
    total_teorico = teorico_df["coste_total_teorico"].sum()
    total_ventas = teorico_df["venta_total"].sum()
    total_mermas = mermas_df["coste_merma"].sum() if not mermas_df.empty else 0

    coste_real = total_teorico + total_mermas
    fc_teorico_pct = round(total_teorico / total_ventas * 100, 2) if total_ventas > 0 else 0
    fc_real_pct = round(coste_real / total_ventas * 100, 2) if total_ventas > 0 else 0

    return {
        "total_ventas": round(total_ventas, 2),
        "coste_teorico": round(total_teorico, 2),
        "coste_mermas": round(total_mermas, 2),
        "coste_real": round(coste_real, 2),
        "fc_teorico_pct": fc_teorico_pct,
        "fc_real_pct": fc_real_pct,
        "desviacion_pct": round(fc_real_pct - fc_teorico_pct, 2),
        "alerta": fc_real_pct > 35,  # benchmark hostelería española: 28-35%
    }

# ─── ANÁLISIS POR CATEGORÍA ────────────────────────────────────────────────────

def analizar_por_categoria(teorico_df, mermas_df):
    cats = teorico_df.groupby("categoria").agg(
        total_ventas=("venta_total", "sum"),
        coste_teorico=("coste_total_teorico", "sum"),
        unidades=("unidades", "sum"),
    ).reset_index()

    # Mermas por categoría de ingrediente (aproximado por nombre)
    mermas_cat = {}
    if not mermas_df.empty:
        mermas_por_cat = mermas_df.groupby("categoria")["coste_merma"].sum().to_dict()
        # Mapeo ingrediente→categoría F&B
        mapa_cat = {
            "Mariscos": "Arroces", "Pescados": "Pescados",
            "Carnes": "Carnes", "Verduras": "Ensaladas",
            "Bebidas": "Bebidas", "Postres": "Postres",
            "Pasteleria": "Postres", "Lacteos": "Postres",
            "Panaderia": "Ensaladas", "Secos": "Arroces",
        }
        for ing_cat, coste in mermas_por_cat.items():
            fb_cat = mapa_cat.get(ing_cat, "Otros")
            mermas_cat[fb_cat] = mermas_cat.get(fb_cat, 0) + coste

    resultado = []
    for _, row in cats.iterrows():
        cat = row["categoria"]
        merma_cat = mermas_cat.get(cat, 0)
        coste_real = row["coste_teorico"] + merma_cat
        fc_teorico = round(row["coste_teorico"] / row["total_ventas"] * 100, 2) if row["total_ventas"] > 0 else 0
        fc_real = round(coste_real / row["total_ventas"] * 100, 2) if row["total_ventas"] > 0 else 0
        resultado.append({
            "categoria": cat,
            "total_ventas": round(row["total_ventas"], 2),
            "unidades_vendidas": int(row["unidades"]),
            "coste_teorico": round(row["coste_teorico"], 2),
            "coste_mermas": round(merma_cat, 2),
            "coste_real": round(coste_real, 2),
            "fc_teorico_pct": fc_teorico,
            "fc_real_pct": fc_real,
            "desviacion_pct": round(fc_real - fc_teorico, 2),
            "alerta": fc_real > 35,
        })
    return sorted(resultado, key=lambda x: x["fc_real_pct"], reverse=True)

# ─── ANÁLISIS INVENTARIO ──────────────────────────────────────────────────────

def analizar_inventario(inventario_df, recetas, ventas_df):
    """
    Calcula consumo teórico por ingrediente y lo compara con stock consumido real.
    """
    # Consumo teórico por ingrediente
    consumo_teorico = {}
    for _, venta in ventas_df.iterrows():
        rid = venta["id_receta"]
        if rid not in recetas:
            continue
        for ing in recetas[rid]["ingredientes"]:
            nombre = ing["ingrediente"]
            consumo_teorico[nombre] = consumo_teorico.get(nombre, 0) + (
                ing["cantidad"] * venta["unidades_vendidas"]
            )

    resultado = []
    for _, row in inventario_df.iterrows():
        ing = row["ingrediente"]
        stock_consumido_real = row["stock_inicial_kg_l"] - row["stock_actual_kg_l"]
        consumo_teo = round(consumo_teorico.get(ing, 0), 3)
        desviacion = round(stock_consumido_real - consumo_teo, 3)
        desviacion_pct = round(desviacion / consumo_teo * 100, 1) if consumo_teo > 0 else 0
        valor_stock_actual = round(row["stock_actual_kg_l"] * row["coste_unitario"], 2)

        resultado.append({
            "ingrediente": ing,
            "categoria": row["categoria"],
            "stock_inicial": row["stock_inicial_kg_l"],
            "stock_actual": row["stock_actual_kg_l"],
            "consumido_real": round(stock_consumido_real, 3),
            "consumido_teorico": consumo_teo,
            "desviacion": desviacion,
            "desviacion_pct": desviacion_pct,
            "valor_stock": valor_stock_actual,
            "unidad": row["unidad"],
            "coste_unitario": row["coste_unitario"],
            "alerta": abs(desviacion_pct) > 15,  # misma tolerancia que matching_ap_fb.py
        })
    return resultado

# ─── PLATOS CON MEJOR/PEOR FOOD COST ─────────────────────────────────────────

def ranking_platos(recetas):
    ranking = []
    for rid, rec in recetas.items():
        ranking.append({
            "id": rid,
            "nombre": rec["nombre"],
            "categoria": rec["categoria"],
            "precio_venta": rec["precio_venta"],
            "coste_teorico": rec["coste_teorico"],
            "fc_pct": rec["food_cost_teorico_pct"],
            "margen_bruto": round(rec["precio_venta"] - rec["coste_teorico"], 2),
            "margen_pct": round((rec["precio_venta"] - rec["coste_teorico"]) / rec["precio_venta"] * 100, 2),
        })
    return sorted(ranking, key=lambda x: x["fc_pct"])

# ─── ALERTAS ──────────────────────────────────────────────────────────────────

def generar_alertas(resumen_global, categorias, inventario):
    alertas = []

    if resumen_global["alerta"]:
        alertas.append({
            "nivel": "CRITICO",
            "tipo": "FOOD_COST_GLOBAL",
            "mensaje": f"Food Cost Real {resumen_global['fc_real_pct']}% supera el benchmark (35%). Revisar mermas y desperdicios.",
        })

    for cat in categorias:
        if cat["alerta"]:
            alertas.append({
                "nivel": "AVISO",
                "tipo": "FOOD_COST_CATEGORIA",
                "mensaje": f"Categoría {cat['categoria']}: Food Cost {cat['fc_real_pct']}% — por encima del 35%.",
            })

    for item in inventario:
        if item["alerta"] and item["desviacion"] > 0:
            alertas.append({
                "nivel": "AVISO",
                "tipo": "CONSUMO_EXCESIVO",
                "mensaje": f"{item['ingrediente']}: consumo real {item['consumido_real']} {item['unidad']} vs teórico {item['consumido_teorico']} {item['unidad']} (+{item['desviacion_pct']}%). Posible merma no registrada.",
            })

    return alertas

# ─── GENERAR EXCEL REPORTE ────────────────────────────────────────────────────

def generar_reporte_excel(resumen, categorias, inventario_data, mermas_df, ranking, alertas):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = REPORTES / f"fb_cost_control_{timestamp}.xlsx"

    wb = Workbook()

    VERDE = "FF1DB954"
    ROJO = "FFE05252"
    NARANJA = "FFFF9800"
    AZUL = "FF1A73E8"
    FONDO = "FF0F1117"
    FONDO2 = "FF1C1F2E"
    BLANCO = "FFFFFFFF"
    GRIS = "FF8892A4"

    def estilo_header(ws, fila=1):
        for cell in ws[fila]:
            cell.fill = PatternFill("solid", fgColor=AZUL)
            cell.font = Font(bold=True, color=BLANCO, size=10)
            cell.alignment = Alignment(horizontal="center")

    def autowidth(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

    # ── HOJA 1: Resumen Global ──
    ws1 = wb.active
    ws1.title = "Resumen_Global"
    ws1.append(["MÉTRICA", "VALOR", "REFERENCIA", "ESTADO"])
    estilo_header(ws1)
    data_resumen = [
        ["Total Ventas F&B", f"{resumen['total_ventas']:,.2f} €", "—", "—"],
        ["Coste Teórico", f"{resumen['coste_teorico']:,.2f} €", "—", "—"],
        ["Coste Mermas", f"{resumen['coste_mermas']:,.2f} €", "—", "—"],
        ["Coste Real Total", f"{resumen['coste_real']:,.2f} €", "—", "—"],
        ["Food Cost Teórico %", f"{resumen['fc_teorico_pct']}%", "28-32%", "OK" if resumen['fc_teorico_pct'] <= 32 else "REVISAR"],
        ["Food Cost Real %", f"{resumen['fc_real_pct']}%", "28-35%", "OK" if not resumen['alerta'] else "⚠ ALERTA"],
        ["Desviación (Real-Teórico)", f"+{resumen['desviacion_pct']}%", "<3%", "OK" if resumen['desviacion_pct'] < 3 else "⚠ REVISAR"],
    ]
    for row in data_resumen:
        ws1.append(row)
        cel = ws1.cell(row=ws1.max_row, column=4)
        if "ALERTA" in str(cel.value) or "REVISAR" in str(cel.value):
            cel.font = Font(color=ROJO, bold=True)
        elif "OK" in str(cel.value):
            cel.font = Font(color=VERDE, bold=True)
    autowidth(ws1)

    # ── HOJA 2: Por Categoría ──
    ws2 = wb.create_sheet("Por_Categoria")
    ws2.append(["Categoría","Ventas €","Uds Vendidas","Coste Teórico €","Coste Mermas €","Coste Real €","FC Teórico %","FC Real %","Desviación %","Estado"])
    estilo_header(ws2)
    for cat in categorias:
        estado = "⚠ ALERTA" if cat["alerta"] else "OK"
        ws2.append([
            cat["categoria"], cat["total_ventas"], cat["unidades_vendidas"],
            cat["coste_teorico"], cat["coste_mermas"], cat["coste_real"],
            cat["fc_teorico_pct"], cat["fc_real_pct"], cat["desviacion_pct"], estado
        ])
        cel = ws2.cell(row=ws2.max_row, column=10)
        cel.font = Font(color=ROJO if cat["alerta"] else VERDE, bold=True)
    autowidth(ws2)

    # ── HOJA 3: Inventario ──
    ws3 = wb.create_sheet("Inventario")
    ws3.append(["Ingrediente","Categoría","Stock Inicial","Stock Actual","Unidad","Consumido Real","Consumido Teórico","Desviación","Desv %","Valor Stock €","Estado"])
    estilo_header(ws3)
    for item in inventario_data:
        estado = "⚠ EXCESO" if (item["alerta"] and item["desviacion"] > 0) else ("⚠ BAJO" if (item["alerta"] and item["desviacion"] < 0) else "OK")
        ws3.append([
            item["ingrediente"], item["categoria"], item["stock_inicial"],
            item["stock_actual"], item["unidad"], item["consumido_real"],
            item["consumido_teorico"], item["desviacion"], f"{item['desviacion_pct']}%",
            item["valor_stock"], estado
        ])
        cel = ws3.cell(row=ws3.max_row, column=11)
        cel.font = Font(color=ROJO if item["alerta"] else VERDE, bold=True)
    autowidth(ws3)

    # ── HOJA 4: Mermas ──
    ws4 = wb.create_sheet("Mermas")
    ws4.append(["Fecha","Ingrediente","Categoría","Cantidad","Unidad","Causa","Coste Unit €","Coste Merma €"])
    estilo_header(ws4)
    total_merma = 0
    for _, row in mermas_df.iterrows():
        ws4.append(list(row))
        total_merma += row["coste_merma"]
    ws4.append(["TOTAL","","","","","","",round(total_merma,2)])
    cel_total = ws4.cell(row=ws4.max_row, column=8)
    cel_total.font = Font(bold=True, color=NARANJA)
    autowidth(ws4)

    # ── HOJA 5: Ranking Platos ──
    ws5 = wb.create_sheet("Ranking_Platos")
    ws5.append(["#","Nombre","Categoría","Precio Venta €","Coste Teórico €","Food Cost %","Margen Bruto €","Margen %"])
    estilo_header(ws5)
    for i, plato in enumerate(ranking, 1):
        ws5.append([i, plato["nombre"], plato["categoria"], plato["precio_venta"],
                    plato["coste_teorico"], plato["fc_pct"], plato["margen_bruto"], plato["margen_pct"]])
        cel = ws5.cell(row=ws5.max_row, column=6)
        if plato["fc_pct"] <= 28:
            cel.font = Font(color=VERDE, bold=True)
        elif plato["fc_pct"] <= 35:
            cel.font = Font(color=NARANJA)
        else:
            cel.font = Font(color=ROJO, bold=True)
    autowidth(ws5)

    # ── HOJA 6: Alertas ──
    ws6 = wb.create_sheet("Alertas")
    ws6.append(["Nivel","Tipo","Mensaje"])
    estilo_header(ws6)
    for alerta in alertas:
        ws6.append([alerta["nivel"], alerta["tipo"], alerta["mensaje"]])
        cel = ws6.cell(row=ws6.max_row, column=1)
        cel.font = Font(color=ROJO if alerta["nivel"] == "CRITICO" else NARANJA, bold=True)
    if not alertas:
        ws6.append(["OK", "SIN_ALERTAS", "No se detectaron alertas en el período."])
    autowidth(ws6)

    wb.save(ruta)
    return ruta

# ─── PIPELINE PRINCIPAL ───────────────────────────────────────────────────────

def ejecutar():
    print("=" * 60)
    print("  YVE — F&B COST CONTROL")
    print("=" * 60)

    print("\n📂 Cargando datos...")
    recetas = cargar_recetas()
    ventas_df = cargar_ventas()
    inventario_df = cargar_inventario()
    mermas_df = cargar_mermas()
    print(f"  ✓ {len(recetas)} recetas | {len(ventas_df)} líneas de ventas | {len(inventario_df)} ingredientes | {len(mermas_df)} mermas")

    print("\n🧮 Calculando Food Cost teórico...")
    teorico_df = calcular_food_cost_teorico(recetas, ventas_df)

    print("🧮 Calculando Food Cost real (con mermas)...")
    resumen = calcular_food_cost_real(teorico_df, mermas_df)

    print("📊 Analizando por categoría...")
    categorias = analizar_por_categoria(teorico_df, mermas_df)

    print("📦 Analizando inventario...")
    inventario_data = analizar_inventario(inventario_df, recetas, ventas_df)

    print("🏆 Generando ranking de platos...")
    ranking = ranking_platos(recetas)

    print("🚨 Detectando alertas...")
    alertas = generar_alertas(resumen, categorias, inventario_data)

    print("\n" + "─" * 60)
    print("  RESUMEN GLOBAL — Julio 2025")
    print("─" * 60)
    print(f"  Total Ventas F&B:     {resumen['total_ventas']:>12,.2f} €")
    print(f"  Coste Teórico:        {resumen['coste_teorico']:>12,.2f} € ({resumen['fc_teorico_pct']}%)")
    print(f"  Coste Mermas:         {resumen['coste_mermas']:>12,.2f} €")
    print(f"  Coste Real:           {resumen['coste_real']:>12,.2f} € ({resumen['fc_real_pct']}%)")
    estado = "⚠ POR ENCIMA DEL BENCHMARK" if resumen["alerta"] else "✓ DENTRO DEL BENCHMARK"
    print(f"  Estado:               {estado}")

    print(f"\n  {len(alertas)} alertas detectadas")
    for a in alertas:
        print(f"  [{a['nivel']}] {a['tipo']}: {a['mensaje'][:70]}...")

    print("\n📄 Generando reporte Excel...")
    ruta = generar_reporte_excel(resumen, categorias, inventario_data, mermas_df, ranking, alertas)
    print(f"  ✓ Reporte guardado: {ruta}")
    print("\n✅ FB COST CONTROL COMPLETADO\n")

    return {
        "resumen": resumen,
        "categorias": categorias,
        "inventario": inventario_data,
        "ranking": ranking,
        "alertas": alertas,
    }

if __name__ == "__main__":
    ejecutar()
